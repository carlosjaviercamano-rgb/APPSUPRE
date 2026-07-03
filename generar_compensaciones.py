import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io
import os
import zipfile
import streamlit as st

NIT_ENTIDAD_BANCARIOS = {
    "BANCOLOMBIA": "890903938",
    "DAVIVIENDA":  "860034313",
    "PSE":         "830078512",
    "EFECTY":      "830131993",
    "RECORD":      "111111111",
    "OCCIDENTE":   "860002395",
}
CODIGO_CENTRO_COSTO   = "102"
CODIGO_CUENTA_DEBITO  = "141299011"
CODIGO_CUENTA_CREDITO = "110505017"

ENTIDADES_RECAUDO = {
    "PSE":                    {"nit": "830078512", "cuenta": "138095009"},
    "EFECTY":                 {"nit": "830131993", "cuenta": "138095008"},
    "RECORD":                 {"nit": "800040390", "cuenta": "138095003"},
    "EFECTY-BANCO DE BOGOTA": {"nit": "830131993", "cuenta": "138095010"},
}


def crear_compensaciones(df_area_banco, config, tipo_pago="bancarios"):
    if df_area_banco is None or df_area_banco.empty:
        raise ValueError("No hay datos en Área de Banco para generar compensaciones.")

    df = df_area_banco.copy()
    df["FECHA_NORM"]     = pd.to_datetime(df["FECHA"],           errors="coerce").dt.normalize()
    df["FECHA_DOC_NORM"] = pd.to_datetime(df["FECHA_DOCUMENTO"], errors="coerce").dt.normalize()

    fechas_doc = df["FECHA_DOC_NORM"].dropna().unique()
    if len(fechas_doc) == 0:
        raise ValueError("No se encontraron fechas de documento válidas.")

    hora_str = datetime.now().strftime("%H_%M_%S")
    archivos_generados = []

    ruta_auto = config.get("ruta_compensaciones", "") if config else ""

    for fecha_doc in sorted(fechas_doc):
        df_grupo = df[df["FECHA_DOC_NORM"] == fecha_doc].copy().reset_index(drop=True)
        if df_grupo.empty:
            continue

        fecha_str = pd.Timestamp(fecha_doc).strftime("%d_%m_%Y")
        nombre    = f"COMPENSACION_{fecha_str}_{hora_str}.xlsx"

        if tipo_pago == "bancarios":
            buffer = _generar_bancarios(df_grupo)
        else:
            buffer = _generar_recaudos(df_grupo, fecha_doc)

        # ── Guardar automáticamente en ruta configurada ──────────────────
        if ruta_auto:
            try:
                os.makedirs(ruta_auto, exist_ok=True)
                ruta_completa = os.path.join(ruta_auto, nombre)
                buffer.seek(0)
                with open(ruta_completa, "wb") as f:
                    f.write(buffer.read())
                st.success(f"💾 Guardado en: {ruta_completa}")
            except Exception as e:
                st.warning(f"⚠️ No se pudo guardar automáticamente: {str(e)}")
            buffer.seek(0)

        archivos_generados.append({
            "nombre": nombre,
            "buffer": buffer,
            "fecha":  fecha_str
        })

    if not archivos_generados:
        return "No se generaron compensaciones."

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for arch in archivos_generados:
            arch["buffer"].seek(0)
            zf.writestr(arch["nombre"], arch["buffer"].read())
    zip_buffer.seek(0)

    zip_nombre = f"COMPENSACIONES_{hora_str}.zip"

    st.markdown("#### 📥 Descargar compensaciones generadas:")
    st.download_button(
        label=f"⬇️  Descargar todas las compensaciones ({len(archivos_generados)} archivos)",
        data=zip_buffer,
        file_name=zip_nombre,
        mime="application/zip",
        key="dl_todos_comp",
        type="primary",
        use_container_width=True
    )

    with st.expander("📂 Descargar individualmente"):
        for arch in archivos_generados:
            arch["buffer"].seek(0)
            st.download_button(
                label=f"⬇️  Compensación {arch['fecha']}",
                data=arch["buffer"],
                file_name=arch["nombre"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_comp_{arch['fecha']}"
            )

    return f"{len(archivos_generados)} compensaciones generadas correctamente."


# ══════════════════════════════════════════════════════════════════════════
# LÓGICA BANCARIOS
# ══════════════════════════════════════════════════════════════════════════
def _generar_bancarios(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    _escribir_encabezado(ws)

    consecutivo = 1
    fila_excel  = 2
    filas_validas = [row for _, row in df.iterrows()
                     if str(row.get("CEDULA", "")).strip() not in ("", "nan")]

    for row in filas_validas:
        entidad = str(row.get("ENTIDAD", "")).upper().strip()
        valor   = _num(row.get("VALOR"))
        fra     = _factura(row.get("FRA") or row.get("NUM_FACTURA"))
        fecha_v = _fecha_dt(row.get("FECHA"))

        nit = ""
        for clave, n in NIT_ENTIDAD_BANCARIOS.items():
            if clave in entidad:
                nit = n
                break

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO_COSTO if nit else "",
            nit,
            "NIT" if nit else "",
            CODIGO_CUENTA_DEBITO if nit else "",
            valor,
            fra,
            fecha_v,
            None, None, None, None
        ])
        if fecha_v:
            ws.cell(row=fila_excel, column=8).number_format = "DD/MM/YYYY"

        consecutivo += 1
        fila_excel  += 1

    for row in filas_validas:
        cedula = str(row.get("CEDULA", "")).strip()
        valor  = _num(row.get("VALOR"))

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO_COSTO,
            cedula,
            "CC",
            CODIGO_CUENTA_CREDITO,
            -abs(valor),
            "", "", None, None, None, None
        ])

        consecutivo += 1
        fila_excel  += 1

    _ajustar_anchos(ws)
    return _to_buffer(wb)


# ══════════════════════════════════════════════════════════════════════════
# LÓGICA RECAUDOS
# ══════════════════════════════════════════════════════════════════════════
def _generar_recaudos(df, fecha_doc):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    _escribir_encabezado(ws)

    consecutivo   = 1
    fila_excel    = 2
    fecha_doc_dt  = _fecha_dt(fecha_doc)
    fecha_doc_str = pd.Timestamp(fecha_doc).strftime("%d-%m-%Y") if fecha_doc else ""

    for nombre_entidad, info in ENTIDADES_RECAUDO.items():
        mask  = df["ENTIDAD"].astype(str).str.upper().str.strip() == nombre_entidad.upper()
        total = pd.to_numeric(df.loc[mask, "VALOR"], errors="coerce").sum()

        if total <= 0:
            continue

        detalle = f"{nombre_entidad} COMPENSACION {fecha_doc_str}"

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO_COSTO,
            info["nit"],
            "NIT",
            info["cuenta"],
            total,
            fecha_doc_str,
            fecha_doc_dt,
            None, None, None, detalle
        ])
        if fecha_doc_dt:
            ws.cell(row=fila_excel, column=8).number_format = "DD/MM/YYYY"

        consecutivo += 1
        fila_excel  += 1

    for _, row in df.iterrows():
        cedula = str(row.get("CEDULA", "")).strip()
        if not cedula or cedula == "nan":
            continue

        valor   = _num(row.get("VALOR"))
        recibos = str(row.get("RECIBOS", "")).strip()
        entidad = str(row.get("ENTIDAD", "")).upper().strip()
        fra     = _factura(row.get("FRA") or row.get("NUM_FACTURA"))
        fecha_v = _fecha_dt(row.get("FECHA"))
        detalle = str(row.get("DETALLE", ""))

        if recibos == "NO EXISTE":
            info_ent  = ENTIDADES_RECAUDO.get(entidad, {})
            nit_cred  = info_ent.get("nit", "")
            tipo_dni  = "NIT"
            cuenta    = CODIGO_CUENTA_DEBITO
            fra_out   = fra
            fecha_out = fecha_v
        else:
            nit_cred  = cedula
            tipo_dni  = "CC"
            cuenta    = CODIGO_CUENTA_CREDITO
            fra_out   = ""
            fecha_out = None

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO_COSTO if nit_cred else "",
            nit_cred,
            tipo_dni if nit_cred else "",
            cuenta if nit_cred else "",
            -abs(valor) if valor else "",
            fra_out,
            fecha_out,
            None, None, None, detalle
        ])
        if fecha_out:
            ws.cell(row=fila_excel, column=8).number_format = "DD/MM/YYYY"

        consecutivo += 1
        fila_excel  += 1

    _ajustar_anchos(ws)
    return _to_buffer(wb)


# ══════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════
def _escribir_encabezado(ws):
    encabezados = [
        "Id", "codigoCentroCosto", "dniTercero", "codigoTipoDniTercero",
        "codigoCuenta", "valor", "factura", "fechaVencimiento",
        "codigoImpuesto", "valorBaseImpuesto", "porcentajeImpuesto", "detalle"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")


def _escribir_fila(ws, fila, valores):
    for col_idx, val in enumerate(valores, start=1):
        ws.cell(row=fila, column=col_idx, value=val)


def _ajustar_anchos(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _to_buffer(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _num(val):
    try:
        return float(str(val).replace(",", "") or 0)
    except Exception:
        return 0.0


def _factura(val):
    if not val or str(val).strip() in ("", "nan"):
        return ""
    try:
        f = float(str(val))
        return str(int(f)) if f == int(f) else str(val)
    except Exception:
        return str(val).strip()


def _fecha_dt(val):
    if val is None:
        return None
    try:
        return pd.Timestamp(val).to_pydatetime()
    except Exception:
        return None