import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
import os
import zipfile
import streamlit as st

EMPRESAS = {
    "Movicap":       "ruta_movicap",
    "Suprecartera":  "ruta_suprecartera",
    "Suprecredito":  "ruta_suprecredito",
    "TuCredito":     "ruta_tucredito",
}

METODO_PAGO = {
    "Movicap":       "0041",
    "TuCredito":     "0041",
    "Suprecredito":  "0041",
    "Suprecartera":  "0041",
}


def crear_planos(df_sheet1, config, df_area_banco, tipo_pago,
                 cedulas_excluidas=None, cedulas_solo_cuota=None,
                 cedulas_inmovilizadas=None):
    if df_sheet1 is None or df_sheet1.empty:
        raise ValueError("No hay datos en Sheet1 para generar planos.")

    fecha_doc = None
    if df_area_banco is not None and "FECHA" in df_area_banco.columns:
        fechas = pd.to_datetime(df_area_banco["FECHA"], errors="coerce").dropna()
        if not fechas.empty:
            fecha_doc = fechas.max()

    fecha_str = fecha_doc.strftime("%d_%m_%Y") if fecha_doc else datetime.now().strftime("%d_%m_%Y")
    hora_str  = datetime.now().strftime("%H_%M_%S")
    tipo_str  = "PAGOS_BANCARIOS" if tipo_pago == "bancarios" else "PAGOS_RECAUDOS"

    archivos_generados = []

    excluidas     = [str(c).strip() for c in (cedulas_excluidas or [])]
    solo_cuota    = [str(c).strip() for c in (cedulas_solo_cuota or [])]
    inmovilizadas = {str(k).strip(): v for k, v in (cedulas_inmovilizadas or {}).items()}

    for empresa in EMPRESAS:
        df_emp = df_sheet1[
            df_sheet1["COMPANY"].astype(str).str.strip().str.upper() == empresa.upper()
        ].copy()

        if excluidas and "IDEN" in df_emp.columns:
            df_emp = df_emp[~df_emp["IDEN"].astype(str).str.strip().isin(excluidas)]

        df_emp = df_emp.reset_index(drop=True)
        if df_emp.empty:
            continue

        df_cash     = _construir_cash_receipt(df_emp, empresa, solo_cuota, inmovilizadas)
        df_services = _construir_services(df_emp, df_cash, inmovilizadas)
        df_payment  = _construir_payment_method(df_emp, df_cash, empresa)

        # ── Partir en bloques de máx 150 filas ───────────────────────────
        LIMITE = 150
        total_filas = len(df_cash)
        n_bloques   = max(1, -(-total_filas // LIMITE))  # ceil division

        clave_ruta = EMPRESAS.get(empresa, "")
        ruta_auto  = config.get(clave_ruta, "") if config else ""

        for bloque_idx in range(n_bloques):
            inicio = bloque_idx * LIMITE
            fin    = inicio + LIMITE

            df_cash_b     = df_cash.iloc[inicio:fin].reset_index(drop=True)
            df_emp_b      = df_emp.iloc[inicio:fin].reset_index(drop=True)
            df_services_b = _construir_services(df_emp_b, df_cash_b, inmovilizadas)
            df_payment_b  = _construir_payment_method(df_emp_b, df_cash_b, empresa)

            # Renumerar IDs desde 1 en cada bloque
            df_cash_b["id"] = range(1, len(df_cash_b) + 1)
            df_services_b   = _construir_services(df_emp_b, df_cash_b, inmovilizadas)
            df_payment_b    = _construir_payment_method(df_emp_b, df_cash_b, empresa)

            sufijo = f"_{bloque_idx + 1}" if n_bloques > 1 else ""
            nombre = f"{tipo_str}_{empresa}_{fecha_str}_{hora_str}{sufijo}.xlsx"
            buffer = _generar_excel(df_cash_b, df_services_b, df_payment_b)

            # ── Guardar automáticamente en ruta configurada ───────────────
            if ruta_auto:
                try:
                    os.makedirs(ruta_auto, exist_ok=True)
                    ruta_completa = os.path.join(ruta_auto, nombre)
                    buffer.seek(0)
                    with open(ruta_completa, "wb") as f:
                        f.write(buffer.read())
                    st.success(f"💾 {empresa}{sufijo}: guardado en {ruta_completa}")
                except Exception as e:
                    st.warning(f"⚠️ {empresa}{sufijo}: no se pudo guardar — {str(e)}")
                buffer.seek(0)

            archivos_generados.append({
                "nombre":  nombre,
                "buffer":  buffer,
                "empresa": f"{empresa}{sufijo}"
            })

    if not archivos_generados:
        return "No se encontraron datos para ninguna empresa."

    st.markdown("#### 📥 Descargar planos generados:")
    for arch in archivos_generados:
        arch["buffer"].seek(0)
        st.download_button(
            label=f"⬇️  Descargar {arch['empresa']}",
            data=arch["buffer"],
            file_name=arch["nombre"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{arch['empresa']}"
        )

    return f"{len(archivos_generados)} planos generados correctamente."


# ══════════════════════════════════════════════════════════════════════════
# CONSTRUCCIÓN DE HOJAS
# ══════════════════════════════════════════════════════════════════════════

def _construir_cash_receipt(df, empresa, solo_cuota=None, inmovilizadas=None):
    solo_cuota    = solo_cuota or []
    inmovilizadas = inmovilizadas or {}
    rows = []
    consecutivo = 1
    for idx, row in df.iterrows():
        cedula = row.get("IDEN")
        if not cedula or str(cedula).strip() == "" or str(cedula) == "nan":
            continue

        cuota      = _num(row.get("CUOTA"))
        diferencia = _num(row.get("DIFERENCIA"))
        valor_cb   = _num(row.get("VALOR_CB"))

        valor_pago = diferencia if valor_cb > 0 and diferencia > 0 else cuota

        ced_str          = str(cedula).strip()
        es_solo_cuota    = ced_str in solo_cuota
        es_inmovilizada  = ced_str in inmovilizadas

        if es_inmovilizada:
            data_im   = inmovilizadas[ced_str]
            suma_serv = sum(float(v or 0) for v in data_im["servicios"].values())
            valor_pago = float(data_im["cuota_total"]) - suma_serv

        aplica = 0 if (es_solo_cuota or es_inmovilizada) else 1

        rows.append({
            "id":                          consecutivo,
            "codigoTipoDocumento":         "DC",
            "codigoCentroCosto":           104,
            "fechaDocumento":              _fecha(row.get("FECHA_DOCUMENTO")),
            "fechaPago":                   _fecha(row.get("FECHA")),
            "detalle":                     row.get("DETALLE", ""),
            "codigoTipoDni":               "CC",
            "dni":                         ced_str,
            "factura":                     _factura(row.get("NUM_FACTURA", "")),
            "valorPago":                   valor_pago,
            "aplicaInteresMoratorio":      aplica,
            "aplicaDescuentoProntoPago":   aplica,
            "aplicaGestionCobranza":       aplica,
        })
        consecutivo += 1

    return pd.DataFrame(rows)


def _construir_services(df, df_cash, inmovilizadas=None):
    inmovilizadas = inmovilizadas or {}
    rows = []
    for (idx, row_emp), (_, row_cash) in zip(df.iterrows(), df_cash.iterrows()):
        cedula = row_cash.get("dni")
        if not cedula:
            continue
        ced_str = str(cedula).strip()

        valor_cb       = _num(row_emp.get("VALOR_CB"))
        inmovilizacion = _num(row_emp.get("INMOVILIZACION"))
        otros_gastos   = _num(row_emp.get("OTROS_GASTOS"))
        id_doc         = row_cash.get("id", 1)
        factura        = row_emp.get("NUM_FACTURA", "")

        if valor_cb > 0:
            rows.append({"idDocumento": id_doc, "codigoServicio": 586325,
                         "valor": valor_cb, "factura": factura,
                         "fechaVencimiento": None, "dni": cedula, "codigoTipoDni": "CC"})
        elif inmovilizacion > 0:
            rows.append({"idDocumento": id_doc, "codigoServicio": 19051,
                         "valor": inmovilizacion, "factura": factura,
                         "fechaVencimiento": None, "dni": cedula, "codigoTipoDni": "CC"})
        elif otros_gastos > 0:
            rows.append({"idDocumento": id_doc, "codigoServicio": 11111,
                         "valor": otros_gastos, "factura": factura,
                         "fechaVencimiento": None, "dni": cedula, "codigoTipoDni": "CC"})

        if ced_str in inmovilizadas:
            data_im = inmovilizadas[ced_str]
            for cod_s, val_s in data_im["servicios"].items():
                if float(val_s or 0) > 0:
                    rows.append({
                        "idDocumento":      id_doc,
                        "codigoServicio":   cod_s,
                        "valor":            float(val_s),
                        "factura":          factura,
                        "fechaVencimiento": None,
                        "dni":              cedula,
                        "codigoTipoDni":    "CC",
                    })

    return pd.DataFrame(rows)


def _construir_payment_method(df, df_cash, empresa, inmovilizadas=None):
    metodo = METODO_PAGO.get(empresa, "0040")
    rows = []
    for (idx, row_emp), (_, row_cash) in zip(df.iterrows(), df_cash.iterrows()):
        cedula = row_cash.get("dni")
        if not cedula:
            continue
        rows.append({
            "idDocumento":      row_cash.get("id", 1),
            "codigoMetodoPago": metodo,
            "valor":            _num(row_emp.get("CUOTA")),
            "voucher":          None,
            "dni":              cedula,
            "codigoTipoDni":    "CC",
        })

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════
# GENERAR EXCEL CON FORMATO
# ══════════════════════════════════════════════════════════════════════════

def _generar_excel(df_cash, df_services, df_payment):
    wb = openpyxl.Workbook()
    _escribir_hoja(wb.active, "CashReceipt", df_cash)
    _escribir_hoja(wb.create_sheet("Services"), "Services", df_services)
    _escribir_hoja(wb.create_sheet("PaymentMethod"), "PaymentMethod", df_payment)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _escribir_hoja(ws, titulo, df):
    ws.title = titulo
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align

    cols_fecha = [col for col in df.columns if "fecha" in col.lower()]
    idx_fechas = [list(df.columns).index(c) + 1 for c in cols_fecha]

    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            if col_idx in idx_fechas and value:
                try:
                    cell.value = pd.Timestamp(value).to_pydatetime()
                    cell.number_format = "DD/MM/YYYY"
                except Exception:
                    pass

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


# ══════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════

def _factura(val):
    if not val or str(val).strip() == "" or str(val) == "nan":
        return ""
    try:
        f = float(str(val))
        if f == int(f):
            return str(int(f))
        return str(val)
    except Exception:
        return str(val).strip()


def _num(val):
    try:
        return float(str(val).replace(",", "") or 0)
    except Exception:
        return 0.0


def _fecha(val):
    if val is None:
        return None
    try:
        return pd.Timestamp(val).to_pydatetime()
    except Exception:
        return None