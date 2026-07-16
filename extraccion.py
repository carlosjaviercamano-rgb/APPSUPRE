import pandas as pd
import openpyxl
from datetime import datetime

# ─── Hojas por grupo ───────────────────────────────────────────────────────
HOJAS_BANCARIOS = [
    "BANCOLOMBIA SUPRECREDITO",
    "DAVIVIENDA SUPRECREDITO",
    "DAVIVIENDA SUPRECREDITO ",   # con espacio al final
]

HOJAS_RECAUDOS = [
    "OCCIDENTE SUPRECREDITO 2026",
    "RECORD"
]

# Entidades que viven en cada hoja
ENTIDADES_POR_HOJA = {
    "OCCIDENTE SUPRECREDITO 2026": ["EFECTY", "PSE", "EFECTY-BANCO DE BOGOTA"],
    "RECORD": ["RECORD"]
}

# ─── Mapeo de columnas del libro fuente ────────────────────────────────────
# El libro tiene: A=FECHAINGRESO, B=ENTIDAD, C=NOMBRE, D=CEDULA,
#                 E=DOCUMENTODEAPROBACION, F=TIPODETRANSACIÓN,
#                 G=REFERENCIA, H=VALOR, I=FRA, J=RECIBOS/RECIBO

COL_FECHA    = 0   # A
COL_ENTIDAD  = 1   # B
COL_NOMBRE   = 2   # C
COL_CEDULA   = 3   # D
COL_DOC      = 4   # E
COL_TIPO     = 5   # F
COL_REF      = 6   # G
COL_VALOR    = 7   # H
COL_FRA      = 8   # I
COL_RECIBO   = 9   # J


def leer_hoja(archivo, nombre_hoja):
    """Lee una hoja del libro como DataFrame sin fórmulas."""
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        return None

    ws = wb[nombre_hoja]
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            datos.append(row)
    wb.close()

    if not datos:
        return pd.DataFrame()

    # Detectar número de columnas de la hoja
    max_cols = max(len(r) for r in datos)
    columnas_hoja = [f"COL_{i}" for i in range(max_cols)]

    df = pd.DataFrame(datos, columns=columnas_hoja)
    return df


def extraer_pagos_bancarios(archivo, archivo_corresponsal):
    """
    Extrae datos de las hojas de PAGOS BANCARIOS.
    Filtro: columna D (CEDULA) con valor Y columna J (RECIBOS) vacía.
    Copia columnas A, B, D, H, I → FECHA, ENTIDAD, CEDULA, VALOR, NUM_FACTURA
    De BANCOLOMBIA además copia columna F → T_TRANSACCION
    Proceso CORRESPONSAL: filtra solo REINCIDENTES si T_TRANSACCION = CONSIGNACION CORRESPONSAL CB
    """
    # Cargar cédulas corresponsales
    cedulas_corresponsal = cargar_corresponsal(archivo_corresponsal)

    frames = []

    for nombre_hoja in HOJAS_BANCARIOS:
        df = leer_hoja(archivo, nombre_hoja)
        if df is None or df.empty:
            continue

        # Renombrar columnas relevantes
        rename = {}
        if df.shape[1] > COL_FECHA:    rename[f"COL_{COL_FECHA}"]   = "FECHA"
        if df.shape[1] > COL_ENTIDAD:  rename[f"COL_{COL_ENTIDAD}"] = "ENTIDAD"
        if df.shape[1] > COL_CEDULA:   rename[f"COL_{COL_CEDULA}"]  = "CEDULA"
        if df.shape[1] > COL_TIPO:     rename[f"COL_{COL_TIPO}"]    = "T_TRANSACCION_SRC"
        if df.shape[1] > COL_VALOR:    rename[f"COL_{COL_VALOR}"]   = "VALOR"
        if df.shape[1] > COL_FRA:      rename[f"COL_{COL_FRA}"]     = "NUM_FACTURA"
        if df.shape[1] > COL_RECIBO:   rename[f"COL_{COL_RECIBO}"]  = "RECIBO_SRC"
        df = df.rename(columns=rename)

        # ── Filtro rosado: cedula con valor Y recibo vacío ──────────────
        tiene_cedula  = df["CEDULA"].notna() & (df["CEDULA"].astype(str).str.strip() != "")
        sin_recibo    = df["RECIBO_SRC"].isna() | (df["RECIBO_SRC"].astype(str).str.strip() == "")
        df = df[tiene_cedula & sin_recibo].copy()

        if df.empty:
            continue

        # ── Construir fila destino (orden AREA DE BANCO) ──────────────
        df_out = pd.DataFrame()
        df_out["FECHA"]           = df["FECHA"]   if "FECHA"   in df.columns else None
        df_out["ENTIDAD"]         = df["ENTIDAD"] if "ENTIDAD" in df.columns else nombre_hoja
        df_out["CEDULA"]          = df["CEDULA"]
        df_out["VALOR"]           = df["VALOR"]   if "VALOR"   in df.columns else None
        df_out["FRA"]             = df["NUM_FACTURA"] if "NUM_FACTURA" in df.columns else None
        df_out["RECIBOS"]         = None
        df_out["FECHA_DOCUMENTO"] = None
        df_out["T_TRANSACCION"]   = None   # interno, no visible

        # T_TRANSACCION: solo de BANCOLOMBIA (col F), resto vacío
        if "BANCOLOMBIA" in nombre_hoja.upper() and "T_TRANSACCION_SRC" in df.columns:
            df_out["T_TRANSACCION"] = df["T_TRANSACCION_SRC"]

        # REINCIDENTES_CB es la versión visual de T_TRANSACCION
        df_out["REINCIDENTES_CB"] = None
        df_out["COMPENSACION"]    = None

        # Alias interno para compatibilidad con alistar.py
        df_out["NUM_FACTURA"]     = df_out["FRA"]

        frames.append(df_out)

    if not frames:
        raise ValueError("No se encontraron datos con los filtros aplicados en las hojas de Pagos Bancarios.")

    df_final = pd.concat(frames, ignore_index=True)

    # ── Proceso CORRESPONSAL ────────────────────────────────────────────
    # 1. Limpiar T_TRANSACCION: dejar solo "CONSIGNACION CORRESPONSAL CB"
    mask_consig = df_final["T_TRANSACCION"].astype(str).str.upper().str.strip() == "CONSIGNACION CORRESPONSAL CB"
    df_final.loc[~mask_consig, "T_TRANSACCION"] = None

    # 2. Identificar PRIMERA VEZ: tiene CONSIGNACION pero cédula NO está en CORRESPONSAL
    mask_tiene_consig = df_final["T_TRANSACCION"].notna()
    cedulas_str = df_final["CEDULA"].astype(str).str.strip()
    mask_primera_vez = mask_tiene_consig & (~cedulas_str.isin(cedulas_corresponsal))

    # 3. Borrar T_TRANSACCION de los PRIMERA VEZ (quedan en blanco)
    df_final.loc[mask_primera_vez, "T_TRANSACCION"] = None

    # 4. Reflejar en REINCIDENTES_CB (columna visual)
    df_final["REINCIDENTES_CB"] = df_final["T_TRANSACCION"]

    # Calcular fecha documento = fecha más reciente de col FECHA
    fecha_doc = None
    fechas = pd.to_datetime(df_final["FECHA"], errors="coerce").dropna()
    if not fechas.empty:
        fecha_doc = fechas.max()
    df_final["FECHA_DOCUMENTO"] = fecha_doc

    n_total       = len(df_final)
    n_corresponsal = mask_tiene_consig.sum()
    n_primera_vez  = mask_primera_vez.sum()
    n_reincidente  = n_corresponsal - n_primera_vez

    resumen = (f"{n_total} registros extraídos. "
               f"Corresponsal: {n_corresponsal} ({n_reincidente} reincidentes, "
               f"{n_primera_vez} primera vez eliminados).")

    return df_final, resumen


def extraer_pagos_recaudos(archivo, fechas_filtro, entidades_filtro=None):
    """
    Extrae datos de las hojas de PAGOS POR RECAUDOS.
    Filtro 1: columna A (FECHA) en lista de fechas seleccionadas.
    Filtro 2: columna B (ENTIDAD) en lista de entidades seleccionadas.
    FECHA_DOCUMENTO = la misma fecha de cada registro.
    """
    if not fechas_filtro:
        raise ValueError("Debes seleccionar al menos una fecha para los Pagos por Recaudos.")

    fechas_dt = [pd.Timestamp(f).normalize() for f in fechas_filtro]
    frames = []

    # Determinar qué hojas procesar según entidades seleccionadas
    if entidades_filtro:
        hojas_a_procesar = []
        for hoja, entidades_hoja in ENTIDADES_POR_HOJA.items():
            # Incluir hoja si al menos una entidad seleccionada pertenece a ella
            if any(e in entidades_hoja for e in entidades_filtro):
                hojas_a_procesar.append(hoja)
    else:
        hojas_a_procesar = HOJAS_RECAUDOS

    for nombre_hoja in hojas_a_procesar:
        df = leer_hoja(archivo, nombre_hoja)
        if df is None or df.empty:
            continue

        rename = {}
        if df.shape[1] > COL_FECHA:   rename[f"COL_{COL_FECHA}"]   = "FECHA"
        if df.shape[1] > COL_ENTIDAD: rename[f"COL_{COL_ENTIDAD}"] = "ENTIDAD"
        if df.shape[1] > COL_CEDULA:  rename[f"COL_{COL_CEDULA}"]  = "CEDULA"
        if df.shape[1] > COL_VALOR:   rename[f"COL_{COL_VALOR}"]   = "VALOR"
        if df.shape[1] > COL_FRA:     rename[f"COL_{COL_FRA}"]     = "NUM_FACTURA"
        if df.shape[1] > COL_RECIBO:  rename[f"COL_{COL_RECIBO}"]  = "RECIBO_SRC"
        df = df.rename(columns=rename)

        # ── Filtro 1: por múltiples fechas ──────────────────────────────
        df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
        df = df[df["FECHA"].dt.normalize().isin(fechas_dt)].copy()

        if df.empty:
            continue

        # ── Filtro 2: por entidad (columna B ya renombrada a ENTIDAD) ──────
        if entidades_filtro and "ENTIDAD" in df.columns:
            entidades_hoja    = ENTIDADES_POR_HOJA.get(nombre_hoja, [])
            entidades_aplicar = [e for e in entidades_filtro if e in entidades_hoja]
            if entidades_aplicar:
                mask = df["ENTIDAD"].astype(str).str.upper().str.strip().isin(
                    [e.upper() for e in entidades_aplicar]
                )
                df = df[mask].copy()

        if df.empty:
            continue

        df_out = pd.DataFrame()
        df_out["FECHA"]           = df["FECHA"]
        df_out["ENTIDAD"]         = df["ENTIDAD"] if "ENTIDAD" in df.columns else nombre_hoja
        df_out["CEDULA"]          = df["CEDULA"]  if "CEDULA"  in df.columns else None
        df_out["VALOR"]           = df["VALOR"]   if "VALOR"   in df.columns else None
        df_out["FRA"]             = df["NUM_FACTURA"] if "NUM_FACTURA" in df.columns else None
        df_out["RECIBOS"]         = None
        df_out["T_TRANSACCION"]   = None
        df_out["REINCIDENTES_CB"] = None
        df_out["COMPENSACION"]    = None

        # FECHA_DOCUMENTO = la misma fecha de cada registro (no la más reciente)
        df_out["FECHA_DOCUMENTO"] = df["FECHA"]

        # Alias interno
        df_out["NUM_FACTURA"] = df_out["FRA"]

        frames.append(df_out)

    if not frames:
        fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_filtro)
        raise ValueError(f"No se encontraron registros para las fechas {fechas_str} en las hojas de Recaudos.")

    if not frames:
        fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_filtro)
        ent_str = ", ".join(entidades_filtro) if entidades_filtro else "todas"
        raise ValueError(f"No se encontraron registros para las fechas {fechas_str} y entidades: {ent_str}.")

    df_final = pd.concat(frames, ignore_index=True)
    fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_filtro)
    ent_str = ", ".join(entidades_filtro) if entidades_filtro else "todas las entidades"
    resumen = f"{len(df_final)} registros extraídos. Fechas: {fechas_str}. Entidades: {ent_str}."
    return df_final, resumen


def cargar_corresponsal(archivo):
    """Carga las cédulas del archivo CORRESPONSAL como un set de strings."""
    try:
        df = pd.read_excel(archivo, header=None)
        cedulas = df.iloc[:, 0].dropna().astype(str).str.strip()
        return set(cedulas)
    except Exception:
        return set()


# ══════════════════════════════════════════════════════════════════════════
# EXTRACCIÓN CARGUE BANCO
# ══════════════════════════════════════════════════════════════════════════

# Valores a excluir en DAVIVIENDA columna H (CONCEPTO)
EXCLUIR_DAVIVIENDA = [
    "Redeban BreB", "BANSUP ESTABLECIMIEN",
    "Compras y Pagos PSE", "Multiabonos", "BTA PROCESOS ESP."
]

# Valores a excluir en DAVIVIENDA columna G (TIPOMOVIMIENTO)
EXCLUIR_DAVIVIENDA_TIPOMOV = ["Nota Débito"]

# Columnas del libro para cargue banco
# BANCOLOMBIA: A=fecha, B=entidad, C=cedula, D=?, E=fra, F=?, G=?, H=valor
# DAVIVIENDA:  A=fecha, B=entidad, C=cedula, D=?, E=fra, F=?, G=tipomovimiento, H=concepto, I=valor, J=fra

def extraer_cargue_banco(archivo, fechas_filtro):
    """
    Extrae movimientos bancarios para Cargue Banco.
    BANCOLOMBIA: filtro fecha col A + col H > 0
    DAVIVIENDA:  filtro fecha col A + col G (TIPOMOVIMIENTO) != Nota Débito
                 + col H (CONCEPTO) no en exclusiones + col C vacía
    """
    if not fechas_filtro:
        raise ValueError("Debes seleccionar al menos una fecha.")

    fechas_dt = [pd.Timestamp(f).normalize() for f in fechas_filtro]
    frames = []

    # ── BANCOLOMBIA ─────────────────────────────────────────────────────
    # Columnas: A=0(FECHA) B=1(ENTIDAD) H=7(VALOR) I=8(FRA)
    df_banc = leer_hoja(archivo, "BANCOLOMBIA SUPRECREDITO")
    if df_banc is not None and not df_banc.empty:
        df_banc.columns = [f"COL_{i}" for i in range(df_banc.shape[1])]

        # Filtro 1: fecha columna A
        df_banc["COL_0"] = pd.to_datetime(df_banc["COL_0"], errors="coerce")
        df_banc = df_banc[df_banc["COL_0"].dt.normalize().isin(fechas_dt)].copy()

        # Filtro 2: columna H > 0
        if df_banc.shape[1] > 7:
            df_banc["COL_7"] = pd.to_numeric(df_banc["COL_7"], errors="coerce").fillna(0)
            df_banc = df_banc[df_banc["COL_7"] > 0].copy()

        if not df_banc.empty:
            df_out = pd.DataFrame()
            df_out["FECHA"]   = df_banc["COL_0"]
            df_out["ENTIDAD"] = df_banc["COL_1"] if df_banc.shape[1] > 1 else "BANCOLOMBIA"
            df_out["VALOR"]   = df_banc["COL_7"] if df_banc.shape[1] > 7 else 0
            df_out["FRA"]     = df_banc["COL_8"] if df_banc.shape[1] > 8 else None
            frames.append(df_out)

    # ── DAVIVIENDA ──────────────────────────────────────────────────────
    # Columnas: A=0(FECHA) B=1(ENTIDAD) C=2(CEDULA) G=6(TIPOMOVIMIENTO)
    #           H=7(CONCEPTO) I=8(VALOR) J=9(FRA)
    df_davi = leer_hoja(archivo, "DAVIVIENDA SUPRECREDITO")
    if df_davi is not None and not df_davi.empty:
        df_davi.columns = [f"COL_{i}" for i in range(df_davi.shape[1])]

        # Filtro 1: fecha columna A
        df_davi["COL_0"] = pd.to_datetime(df_davi["COL_0"], errors="coerce")
        df_davi = df_davi[df_davi["COL_0"].dt.normalize().isin(fechas_dt)].copy()

        # Filtro 2: columna G (TIPOMOVIMIENTO) != Nota Débito
        if df_davi.shape[1] > 6:
            excluir_tipomov = [e.upper().strip() for e in EXCLUIR_DAVIVIENDA_TIPOMOV]
            mask_tipomov = df_davi["COL_6"].astype(str).str.upper().str.strip().isin(excluir_tipomov)
            df_davi = df_davi[~mask_tipomov].copy()

        # Filtro 3: columna H (CONCEPTO) no en exclusiones
        if df_davi.shape[1] > 7:
            excluir_upper = [e.upper().strip() for e in EXCLUIR_DAVIVIENDA]
            mask_excluir = df_davi["COL_7"].astype(str).str.upper().str.strip().isin(excluir_upper)
            df_davi = df_davi[~mask_excluir].copy()

        # Filtro 4: columna C vacía
        if df_davi.shape[1] > 2:
            mask_vacia = df_davi["COL_2"].isna() | (df_davi["COL_2"].astype(str).str.strip().isin(["", "nan"]))
            df_davi = df_davi[mask_vacia].copy()

        if not df_davi.empty:
            df_out2 = pd.DataFrame()
            df_out2["FECHA"]   = df_davi["COL_0"]
            df_out2["ENTIDAD"] = df_davi["COL_1"] if df_davi.shape[1] > 1 else "DAVIVIENDA"
            df_out2["VALOR"]   = pd.to_numeric(df_davi["COL_8"], errors="coerce") if df_davi.shape[1] > 8 else 0
            df_out2["FRA"]     = df_davi["COL_9"] if df_davi.shape[1] > 9 else None
            frames.append(df_out2)

    if not frames:
        fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_filtro)
        raise ValueError(f"No se encontraron movimientos para las fechas: {fechas_str}")

    df_final = pd.concat(frames, ignore_index=True)
    fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_filtro)
    return df_final, f"{len(df_final)} movimientos extraídos para: {fechas_str}"