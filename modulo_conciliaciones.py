import streamlit as st
import pandas as pd
import io


def render():
    st.markdown("""
    <div class="module-header">
        <div class="module-icon">🔍</div>
        <div>
            <h1>Conciliaciones</h1>
            <p>Conciliación bancaria, QR & Credibanco y cuentas puentes/transitorias</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Submódulos ───────────────────────────────────────────────────────
    if "submodulo_conciliacion" not in st.session_state:
        st.session_state.submodulo_conciliacion = None

    sub = st.session_state.submodulo_conciliacion

    if sub is None:
        _render_menu_submodulos()
    elif sub == "bancaria":
        _render_volver()
        render_bancaria()
    elif sub == "qr_credibanco":
        _render_volver()
        render_qr_credibanco()
    elif sub == "puentes":
        _render_volver()
        render_puentes()


def _render_menu_submodulos():
    st.markdown("### Selecciona el tipo de conciliación")
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;cursor:pointer;">
            <div style="font-size:2.5rem">🏦</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Conciliación Bancaria</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Libro Auxiliar + Extracto bancario</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_bancaria", use_container_width=True, type="primary"):
            st.session_state.submodulo_conciliacion = "bancaria"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;cursor:pointer;">
            <div style="font-size:2.5rem">📱</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Conciliación QR & Credibanco</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Libro Auxiliar + Libro de Banco</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_qr", use_container_width=True, type="primary"):
            st.session_state.submodulo_conciliacion = "qr_credibanco"
            st.rerun()

    with col3:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;cursor:pointer;">
            <div style="font-size:2.5rem">🔄</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Cuentas Puentes/Transitorias</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Solo Libro Auxiliar</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_puentes", use_container_width=True, type="primary"):
            st.session_state.submodulo_conciliacion = "puentes"
            st.rerun()


def _render_volver():
    if st.button("← Volver a Conciliaciones", key="btn_volver_conciliacion"):
        st.session_state.submodulo_conciliacion = None
        st.rerun()
    st.markdown("<br>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
# SECCIÓN COMPARTIDA: LIBRO AUXILIAR
# ══════════════════════════════════════════════════════════════════════════

COLUMNAS_AUXILIAR = [
    "Source.Name", "id", "empresa", "centrocosto", "fecha", "tipodocumento",
    "nrodocumento", "codigocuenta", "nombrecuenta", "descripcion", "factura",
    "fechavencimiento", "identificacion", "tercero", "valor", "baseimpuesto",
    "porcentajeimpuesto", "saldoanterior", "debito", "credito", "saldoactual",
    "usuariocreador", "usuarioactualizador"
]


def render_cargue_auxiliares(key_prefix=""):
    """
    Sección reutilizable para cargar y unir auxiliares.
    Retorna el DataFrame del libro auxiliar unificado o None.
    """
    st.markdown("#### 📂 Libro Auxiliar")
    st.caption("Sube hasta 15 auxiliares individuales. La app los unirá automáticamente.")

    key_archivos = f"{key_prefix}_auxiliares"
    if key_archivos not in st.session_state:
        st.session_state[key_archivos] = []

    # Uploader múltiple
    archivos = st.file_uploader(
        "Selecciona los auxiliares (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key=f"up_{key_prefix}_auxiliares",
        label_visibility="collapsed"
    )

    if archivos:
        # Limitar a 15
        if len(archivos) > 15:
            st.warning("⚠️ Máximo 15 auxiliares. Se tomarán los primeros 15.")
            archivos = archivos[:15]
        st.session_state[key_archivos] = archivos

    archivos_cargados = st.session_state.get(key_archivos, [])

    if archivos_cargados:
        st.success(f"✅ {len(archivos_cargados)} auxiliar(es) cargado(s):")
        for arch in archivos_cargados:
            st.caption(f"   📄 {arch.name}")

        key_df = f"{key_prefix}_df_auxiliar"

        if st.button("🔗 Unir Auxiliares", key=f"btn_unir_{key_prefix}",
                     type="primary", use_container_width=True):
            with st.spinner("Uniendo auxiliares..."):
                try:
                    df_unido = _unir_auxiliares(archivos_cargados)
                    st.session_state[key_df] = df_unido
                    st.success(f"✅ Libro Auxiliar creado: {len(df_unido):,} registros de {len(archivos_cargados)} auxiliar(es).")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")

        # Mostrar preview y descarga si ya está unido
        df_auxiliar = st.session_state.get(key_df)
        if df_auxiliar is not None:
            st.markdown("---")
            c1, c2, c3 = st.columns(3)
            c1.metric("Total registros", f"{len(df_auxiliar):,}")
            c2.metric("Empresas", df_auxiliar["empresa"].nunique() if "empresa" in df_auxiliar.columns else 0)
            c3.metric("Auxiliares", df_auxiliar["Source.Name"].nunique() if "Source.Name" in df_auxiliar.columns else 0)

            with st.expander("👁️ Vista previa"):
                st.dataframe(df_auxiliar.head(20), use_container_width=True)

            # Botón descargar libro auxiliar con formato de decimales
            import openpyxl
            from openpyxl.styles import numbers
            buf = io.BytesIO()

            # Escribir con openpyxl para aplicar formato
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_auxiliar.to_excel(writer, index=False, sheet_name="AUXILIARES")
                ws = writer.sheets["AUXILIARES"]

                # Columnas numéricas con 2 decimales
                cols_decimales = ["valor", "baseimpuesto", "saldoanterior",
                                  "debito", "credito", "saldoactual"]
                # Encontrar índices de esas columnas
                header = [c.value for c in ws[1]]
                for col_idx, col_name in enumerate(header, start=1):
                    if col_name in cols_decimales:
                        for row in ws.iter_rows(min_row=2, min_col=col_idx,
                                                max_col=col_idx, max_row=ws.max_row):
                            for cell in row:
                                cell.number_format = '#,##0.00'

            buf.seek(0)
            st.download_button(
                label="⬇️  Descargar Libro Auxiliar unificado",
                data=buf.getvalue(),
                file_name="LIBRO_AUXILIAR.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_auxiliar_{key_prefix}"
            )
            return df_auxiliar

    return None


def _unir_auxiliares(archivos):
    """Une múltiples auxiliares en un solo DataFrame."""
    frames = []
    for archivo in archivos:
        try:
            df = pd.read_excel(archivo, sheet_name=0)
            df.columns = [c.strip().lower() for c in df.columns]

            # Obtener nombre empresa del archivo
            empresa = ""
            if "empresa" in df.columns and not df.empty:
                empresa = str(df["empresa"].iloc[0]).strip()

            # Agregar Source.Name con el nombre de la empresa
            df.insert(0, "Source.Name", empresa if empresa else archivo.name)

            frames.append(df)
        except Exception as e:
            raise ValueError(f"Error leyendo {archivo.name}: {str(e)}")

    if not frames:
        raise ValueError("No se pudieron leer los auxiliares.")

    df_final = pd.concat(frames, ignore_index=True)
    return df_final


# ══════════════════════════════════════════════════════════════════════════
# SUBMÓDULO 1: CONCILIACIÓN BANCARIA
# ══════════════════════════════════════════════════════════════════════════
def render_bancaria():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">🏦 Conciliación Bancaria</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Libro Auxiliar + Extracto bancario</p>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📂 1. Cargar Archivos", "🔍 2. Conciliar"])

    with tab1:
        df_aux = render_cargue_auxiliares("bancaria")
        if df_aux is not None:
            st.session_state["bancaria_df_auxiliar"] = df_aux

        st.markdown("---")
        st.markdown("#### 🏦 Extracto Bancario")
        extracto = st.file_uploader(
            "Sube el extracto bancario",
            type=["xlsx", "xls", "csv"],
            key="up_extracto_bancario",
            label_visibility="collapsed"
        )
        if extracto:
            st.session_state["bancaria_extracto"] = extracto
            st.success(f"✅ {extracto.name}")

    with tab2:
        st.info("🚧 Lógica de conciliación — próximamente. Primero carga los archivos en la pestaña 1.")


# ══════════════════════════════════════════════════════════════════════════
# SUBMÓDULO 2: CONCILIACIÓN QR & CREDIBANCO
# ══════════════════════════════════════════════════════════════════════════
def render_qr_credibanco():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">📱 Conciliación QR & Credibanco</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Libro Auxiliar + Libro de Banco</p>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📂 1. Cargar Archivos", "🔍 2. Conciliar"])

    with tab1:
        df_aux = render_cargue_auxiliares("qr")
        if df_aux is not None:
            st.session_state["qr_df_auxiliar"] = df_aux

        st.markdown("---")
        st.markdown("#### 📖 Libro de Banco")
        if st.session_state.get("archivo_libro"):
            st.success("✅ Libro de Banco ya cargado desde Aplicación y Compensación.")
        else:
            libro = st.file_uploader(
                "Sube el Libro de Banco",
                type=["xlsx", "xlsm"],
                key="up_libro_qr",
                label_visibility="collapsed"
            )
            if libro:
                st.session_state["qr_libro_banco"] = libro
                st.success(f"✅ {libro.name}")

    with tab2:
        st.info("🚧 Lógica de conciliación — próximamente. Primero carga los archivos en la pestaña 1.")


# ══════════════════════════════════════════════════════════════════════════
# SUBMÓDULO 3: CUENTAS PUENTES / TRANSITORIAS
# ══════════════════════════════════════════════════════════════════════════
def render_puentes():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">🔄 Cuentas Puentes / Transitorias</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Solo Libro Auxiliar</p>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📂 1. Cargar Archivos", "🔍 2. Conciliar"])

    with tab1:
        df_aux = render_cargue_auxiliares("puentes")
        if df_aux is not None:
            st.session_state["puentes_df_auxiliar"] = df_aux

    with tab2:
        st.info("🚧 Lógica de conciliación — próximamente. Primero carga los archivos en la pestaña 1.")
