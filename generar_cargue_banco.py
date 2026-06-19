import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io
import streamlit as st

# Configuración por entidad
ENTIDADES_CONFIG = {
    "BANCOLOMBIA": {
        "nit":           "890903938",
        "cuenta_debito": "112005001",
    },
    "DAVIVIENDA": {
        "nit":           "860034313",
        "cuenta_debito": "111005005",
    }
}

CUENTA_CREDITO    = "141299011"
CODIGO_CENTRO     = "102"


def crear_cargue_banco(df_movimientos, tipo_pago="bancarios"):
    """
    Genera archivos de cargue banco (Items) con débitos y créditos.
    Un archivo por fecha de movimiento.
    """
    if df_movimientos is None or df_movimientos.empty:
        raise ValueError("No hay movimientos para generar el cargue.")

    df = df_movimientos.copy()
    df["FECHA_NORM"] = pd.to_datetime(df["FECHA"], errors="coerce").dt.normalize()
    fechas_unicas = df["FECHA_NORM"].dropna().unique()

    hora_str = datetime.now().strftime("%H_%M_%S")
    archivos_generados = []

    for fecha in sorted(fechas_unicas):
        df_fecha = df[df["FECHA_NORM"] == fecha].copy().reset_index(drop=True)
        if df_fecha.empty:
            continue

        fecha_str = pd.Timestamp(fecha).strftime("%d_%m_%Y")
        nombre    = f"CARGUE_BANCO_{fecha_str}_{hora_str}.xlsx"
        buffer    = _generar_excel(df_fecha)
        archivos_generados.append({
            "nombre": nombre,
            "buffer": buffer,
            "fecha":  fecha_str
        })

    if not archivos_generados:
        return "No se generaron archivos de cargue."

    # Mostrar botones de descarga
    st.markdown("#### 📥 Descargar archivos de cargue:")
    for arch in archivos_generados:
        arch["buffer"].seek(0)
        st.download_button(
            label=f"⬇️  Cargue {arch['fecha']}",
            data=arch["buffer"].read(),
            file_name=arch["nombre"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_cargue_{arch['fecha']}"
        )

    return f"{len(archivos_generados)} archivos de cargue generados."


def _generar_excel(df):
    """
    Genera el Excel con DÉBITOS seguidos de CRÉDITOS.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Encabezados
    encabezados = [
        "Id", "codigoCentroCosto", "dniTercero", "codigoTipoDniTercero",
        "codigoCuenta", "valor", "factura", "fechaVencimiento",
        "codigoImpuesto", "valorBaseImpuesto", "porcentajeImpuesto", "detalle"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    consecutivo = 1
    fila_excel  = 2

    # ── DÉBITOS ─────────────────────────────────────────────────────────
    for _, row in df.iterrows():
        entidad = str(row.get("ENTIDAD", "")).upper().strip()
        config  = ENTIDADES_CONFIG.get(entidad, {})
        nit     = config.get("nit", "")
        cuenta  = config.get("cuenta_debito", "")
        valor   = _num(row.get("VALOR"))
        fecha   = _fecha_dt(row.get("FECHA"))
        fecha_str = _fecha_str(row.get("FECHA"))
        detalle = f"{entidad} {fecha_str}".strip()

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO if nit else "",
            nit,
            "NIT" if nit else "",
            cuenta,
            valor,
            "",      # factura vacía en débito
            None,    # fechaVencimiento vacía en débito
            None, None, None,
            detalle
        ])

        consecutivo += 1
        fila_excel  += 1

    # ── CRÉDITOS ─────────────────────────────────────────────────────────
    for _, row in df.iterrows():
        entidad = str(row.get("ENTIDAD", "")).upper().strip()
        config  = ENTIDADES_CONFIG.get(entidad, {})
        nit     = config.get("nit", "")
        valor   = _num(row.get("VALOR"))
        fra     = _factura(row.get("FRA"))
        fecha   = _fecha_dt(row.get("FECHA"))

        _escribir_fila(ws, fila_excel, [
            consecutivo,
            CODIGO_CENTRO if nit else "",
            nit,
            "NIT" if nit else "",
            CUENTA_CREDITO,
            -abs(valor) if valor else "",
            fra,     # factura en crédito
            fecha,   # fechaVencimiento en crédito
            None, None, None,
            ""
        ])

        # Formato fecha
        if fecha:
            ws.cell(row=fila_excel, column=8).number_format = "DD/MM/YYYY"

        consecutivo += 1
        fila_excel  += 1

    # Ajustar anchos
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _escribir_fila(ws, fila, valores):
    for col_idx, val in enumerate(valores, start=1):
        ws.cell(row=fila, column=col_idx, value=val)


def _num(val):
    try:
        return round(float(str(val).replace(",", "") or 0), 2)
    except Exception:
        return 0.0


def _factura(val):
    if not val or str(val).strip() in ("", "nan"):
        return ""
    try:
        f = float(str(val))
        return str(int(f)) if f == int(f) else str(val)
    except Exception:
        return str(val).strip()


def _fecha_dt(val):
    if val is None:
        return None
    try:
        return pd.Timestamp(val).to_pydatetime()
    except Exception:
        return None


def _fecha_str(val):
    if val is None:
        return ""
    try:
        return pd.Timestamp(val).strftime("%d-%m-%Y")
    except Exception:
        return str(val)[:10] if val else ""
