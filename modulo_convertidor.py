import streamlit as st
import pandas as pd
import re
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def render():
    st.markdown("""
    <div class="module-header">
        <div class="module-icon">🔄</div>
        <div>
            <h1>Convertidor de Archivos</h1>
            <p>Conversión de reportes bancarios a Excel estructurado</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if "submodulo_convertidor" not in st.session_state:
        st.session_state.submodulo_convertidor = None

    sub = st.session_state.submodulo_convertidor

    if sub is None:
        _render_menu()
    elif sub == "efecty_record":
        _render_volver()
        render_efecty_record()
    elif sub == "bancolombia":
        _render_volver()
        render_bancolombia()


def _render_menu():
    st.markdown("### Selecciona el tipo de conversión")
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">🧾</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Efecty / Record</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Reporte .lst → Excel estructurado</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_efecty", use_container_width=True, type="primary"):
            st.session_state.submodulo_convertidor = "efecty_record"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">🏦</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Bancolombia</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Extracto PDF → Excel estructurado</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_bancolombia", use_container_width=True, type="primary"):
            st.session_state.submodulo_convertidor = "bancolombia"
            st.rerun()

    with col3:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">➕</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Próximamente</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Nuevos convertidores</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.button("Próximamente", key="btn_prox2", use_container_width=True, disabled=True)


def _render_volver():
    if st.button("← Volver a Convertidor", key="btn_volver_convertidor"):
        st.session_state.submodulo_convertidor = None
        st.rerun()
    st.markdown("<br>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
# CONVERTIDOR EFECTY / RECORD — ARCHIVO .LST
# ══════════════════════════════════════════════════════════════════════════

def render_efecty_record():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">🧾 Efecty / Record — Reporte .lst</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Convierte el reporte de recaudo en formato .lst a Excel estructurado</p>
    </div>
    """, unsafe_allow_html=True)

    archivo = st.file_uploader(
        "Sube el archivo de reporte (.lst)",
        type=["lst"],
        key="up_lst",
        label_visibility="collapsed"
    )

    if archivo is None:
        st.info("📂 Sube el archivo .lst para continuar.")
        return

    # Parsear
    try:
        contenido = archivo.read().decode("latin-1")
        datos     = _parsear_lst(contenido)
    except Exception as e:
        st.error(f"❌ Error al leer el archivo: {str(e)}")
        return

    if not datos["secciones"]:
        st.error("❌ No se encontraron secciones válidas en el archivo. Verifica el formato.")
        return

    # Mostrar resumen
    st.success(f"✅ Archivo parseado: **{datos['codigo']} - {datos['empresa']}** | Fecha: {datos['fecha']}")
    st.markdown("#### 📊 Resumen por entidad")

    resumen_rows = []
    for sec in datos["secciones"]:
        resumen_rows.append({
            "Entidad":              sec["entidad"],
            "Operaciones":          len(sec["registros"]),
            "Valor Total Compensado": sec["valor_total"],
            "Valor Adicional Total":  sec["valor_adicional_total"],
        })
    df_resumen = pd.DataFrame(resumen_rows)
    df_resumen["Valor Total Compensado"] = df_resumen["Valor Total Compensado"].apply(lambda x: f"${x:,.2f}")
    df_resumen["Valor Adicional Total"]  = df_resumen["Valor Adicional Total"].apply(lambda x: f"${x:,.2f}")
    st.dataframe(df_resumen, use_container_width=True, hide_index=True)

    # Generar Excel
    import re as _re
    nombre_raw  = archivo.name.replace(".lst", "").replace(".LST", "")
    m_fecha_arch = _re.search(r"(\d{8})$", nombre_raw)
    fecha_arch   = m_fecha_arch.group(1) if m_fecha_arch else nombre_raw
    nombre_base  = nombre_raw
    nombre_xlsx  = f"EFECTY{fecha_arch}.xlsx"

    excel_bytes = _generar_excel_lst(datos, nombre_base)

    # Guardar automáticamente en carpeta fija
    ruta_auto = r"C:\Users\ASUS\Desktop\BANCOS\CONVERTIDOR DE ARCHIVO"
    if ruta_auto:
        try:
            import os
            os.makedirs(ruta_auto, exist_ok=True)
            ruta_completa = os.path.join(ruta_auto, nombre_xlsx)
            with open(ruta_completa, "wb") as f:
                f.write(excel_bytes)
            st.success(f"💾 Guardado automáticamente en: {ruta_completa}")
        except Exception as e:
            st.warning(f"⚠️ No se pudo guardar automáticamente: {str(e)}")

    st.download_button(
        label=f"⬇️  Descargar Excel — {nombre_xlsx}",
        data=excel_bytes,
        file_name=nombre_xlsx,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="dl_lst_excel"
    )


# ══════════════════════════════════════════════════════════════════════════
# PARSEO DEL ARCHIVO .LST
# ══════════════════════════════════════════════════════════════════════════

def _parsear_lst(contenido):
    lines = contenido.replace("\r", "").split("\n")

    # Encabezado general
    codigo  = ""
    empresa = ""
    fecha   = ""
    for line in lines[:5]:
        m = re.match(r"REPORTE RECAUDO:\s+(\S+)\s+-\s+(.+?)\s+FECHA RECAUDO:\s+(\S+)", line)
        if m:
            codigo  = m.group(1).strip()
            empresa = m.group(2).strip()
            fecha   = m.group(3).strip()
            break

    # Regex para detectar inicio de sección
    re_entidad = re.compile(r"^ENTIDAD DE ABONO:\s+(.+?)\s{2,}CTA PPAL:", re.IGNORECASE)
    # Regex para líneas de detalle
    re_detalle = re.compile(
        r"^(\d{15,20})\s+\$([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})\s+(\d+)\s+(\d+)\s+(\d{4}-\d{2}-\d{2})"
    )

    secciones      = []
    entidad_actual = None
    registros      = []

    for line in lines:
        m_ent = re_entidad.match(line)
        if m_ent:
            # Guardar sección anterior
            if entidad_actual is not None:
                secciones.append(_cerrar_seccion(entidad_actual, registros))
            entidad_actual = m_ent.group(1).strip()
            registros      = []
            continue

        if entidad_actual:
            m_det = re_detalle.match(line.strip())
            if m_det:
                registros.append({
                    "FACTURA":          m_det.group(1),
                    "VALOR":            float(m_det.group(2).replace(",", "")),
                    "VALOR ADICIONAL":  float(m_det.group(3).replace(",", "")),
                    "CAJERO":           m_det.group(4),
                    "TRANSAC.":         m_det.group(5),
                    "FECHA DE PAGO":    m_det.group(6),
                })

    # Cerrar última sección
    if entidad_actual is not None:
        secciones.append(_cerrar_seccion(entidad_actual, registros))

    return {
        "codigo":   codigo,
        "empresa":  empresa,
        "fecha":    fecha,
        "secciones": secciones,
    }


def _cerrar_seccion(entidad, registros):
    valor_total    = sum(r["VALOR"]           for r in registros)
    val_adic_total = sum(r["VALOR ADICIONAL"] for r in registros)
    return {
        "entidad":              entidad,
        "registros":            registros,
        "valor_total":          round(valor_total, 2),
        "valor_adicional_total": round(val_adic_total, 2),
    }


# ══════════════════════════════════════════════════════════════════════════
# GENERACIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════

def _generar_excel_lst(datos, nombre_base):
    wb = openpyxl.Workbook()

    # Estilos comunes
    font_base    = Font(name="Arial", size=10)
    font_bold    = Font(name="Arial", size=10, bold=True)
    font_titulo  = Font(name="Arial", size=12, bold=True)
    font_header  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header  = PatternFill("solid", fgColor="1F4E78")
    fill_total   = PatternFill("solid", fgColor="D9E1F2")
    fill_resumen_header = PatternFill("solid", fgColor="1F4E78")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center")
    fmt_moneda   = '$#,##0.00'
    fmt_num      = '#,##0'

    thin = Side(style="thin", color="AAAAAA")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    secciones  = datos["secciones"]
    fecha_rec  = datos["fecha"]

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"

    # Título
    ws_res.merge_cells("A1:D1")
    c = ws_res["A1"]
    c.value     = f"REPORTE RECAUDO: {datos['codigo']} - {datos['empresa']}    FECHA: {fecha_rec}"
    c.font      = font_titulo
    c.alignment = align_left

    # Encabezados resumen
    hdrs_res = ["ENTIDAD", "OPERACIONES", "VALOR TOTAL COMPENSADO", "VALOR ADICIONAL TOTAL"]
    for ci, h in enumerate(hdrs_res, 1):
        cell = ws_res.cell(row=3, column=ci, value=h)
        cell.font      = font_header
        cell.fill      = fill_resumen_header
        cell.alignment = align_center
        cell.border    = borde

    # Una fila por sección — con referencias a hojas
    for ri, sec in enumerate(secciones, 4):
        nombre_hoja = _nombre_hoja_seguro(sec["entidad"])
        n_reg       = len(sec["registros"])
        fila_total_op  = n_reg + 4   # fila donde está TOTAL OPERACIONES en la hoja
        fila_total_val = n_reg + 5   # fila donde está VALOR TOTAL COMPENSADO
        fila_total_adi = n_reg + 6   # fila donde está VALOR TOTAL RECAUDO ADICIONAL

        ws_res.cell(row=ri, column=1, value=sec["entidad"]).font   = font_base
        ws_res.cell(row=ri, column=2,
                    value=f"='{nombre_hoja}'!B{fila_total_op}").font  = font_base
        ws_res.cell(row=ri, column=3,
                    value=f"='{nombre_hoja}'!B{fila_total_val}").font = font_base
        ws_res.cell(row=ri, column=4,
                    value=f"='{nombre_hoja}'!B{fila_total_adi}").font = font_base

        ws_res.cell(row=ri, column=2).number_format = fmt_num
        ws_res.cell(row=ri, column=3).number_format = fmt_moneda
        ws_res.cell(row=ri, column=4).number_format = fmt_moneda

        for ci in range(1, 5):
            ws_res.cell(row=ri, column=ci).border    = borde
            ws_res.cell(row=ri, column=ci).alignment = align_center if ci > 1 else align_left

    # Fila GRAN TOTAL
    fila_gt = 4 + len(secciones)
    ws_res.cell(row=fila_gt, column=1, value="GRAN TOTAL").font = font_bold
    if len(secciones) > 0:
        rng_op  = f"B4:B{fila_gt-1}"
        rng_val = f"C4:C{fila_gt-1}"
        rng_adi = f"D4:D{fila_gt-1}"
        ws_res.cell(row=fila_gt, column=2, value=f"=SUM({rng_op})").number_format  = fmt_num
        ws_res.cell(row=fila_gt, column=3, value=f"=SUM({rng_val})").number_format = fmt_moneda
        ws_res.cell(row=fila_gt, column=4, value=f"=SUM({rng_adi})").number_format = fmt_moneda

    for ci in range(1, 5):
        c = ws_res.cell(row=fila_gt, column=ci)
        c.fill      = fill_total
        c.font      = font_bold
        c.border    = borde
        c.alignment = align_center if ci > 1 else align_left

    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 26
    ws_res.column_dimensions["D"].width = 26

    # ── Hojas por entidad ─────────────────────────────────────────────────
    for sec in secciones:
        nombre_hoja = _nombre_hoja_seguro(sec["entidad"])
        ws = wb.create_sheet(title=nombre_hoja)

        # Título fila 1
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value     = f"{sec['entidad']} — Fecha Recaudo: {fecha_rec}"
        c.font      = font_titulo
        c.alignment = align_left

        # Encabezados fila 3
        cols = ["FACTURA", "VALOR", "VALOR ADICIONAL", "CAJERO", "TRANSAC.", "FECHA DE PAGO"]
        anchos = [24, 16, 18, 10, 10, 16]
        for ci, (h, w) in enumerate(zip(cols, anchos), 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font      = font_header
            cell.fill      = fill_header
            cell.alignment = align_center
            cell.border    = borde
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Datos desde fila 4
        for ri, reg in enumerate(sec["registros"], 4):
            fac = reg["FACTURA"]
            try:
                fac = int(fac)
            except Exception:
                pass
            fecha_pago = reg["FECHA DE PAGO"]
            try:
                from datetime import datetime as _dt
                fecha_pago = _dt.strptime(fecha_pago, "%Y-%m-%d").strftime("%d-%m-%Y")
            except Exception:
                pass
            vals = [
                fac,
                reg["VALOR"],
                reg["VALOR ADICIONAL"],
                reg["CAJERO"],
                reg["TRANSAC."],
                fecha_pago,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = font_base
                cell.border    = borde
                cell.alignment = align_center if ci != 1 else align_left
                if ci in (1, 2, 3, 4, 5):
                    cell.number_format = "General"

        n_reg      = len(sec["registros"])
        fila_datos_inicio = 4
        fila_datos_fin    = fila_datos_inicio + n_reg - 1
        fila_tot_op  = fila_datos_fin + 2
        fila_tot_val = fila_datos_fin + 3
        fila_tot_adi = fila_datos_fin + 4

        # Totales
        totales = [
            (fila_tot_op,  "TOTAL OPERACIONES",              f"=COUNTA(A{fila_datos_inicio}:A{fila_datos_fin})", fmt_num),
            (fila_tot_val, "VALOR TOTAL COMPENSADO",         f"=SUM(B{fila_datos_inicio}:B{fila_datos_fin})",    fmt_moneda),
            (fila_tot_adi, "VALOR TOTAL RECAUDO ADICIONAL",  f"=SUM(C{fila_datos_inicio}:C{fila_datos_fin})",    fmt_moneda),
        ]
        for fila_t, label, formula, fmt in totales:
            ws.cell(row=fila_t, column=1, value=label).font      = font_bold
            ws.cell(row=fila_t, column=1).fill                   = fill_total
            ws.cell(row=fila_t, column=1).border                 = borde
            ws.cell(row=fila_t, column=1).alignment              = align_left
            c2 = ws.cell(row=fila_t, column=2, value=formula)
            c2.font           = font_bold
            c2.fill           = fill_total
            c2.border         = borde
            c2.alignment      = align_center
            c2.number_format  = fmt
            for ci in range(3, 7):
                c3 = ws.cell(row=fila_t, column=ci)
                c3.fill   = fill_total
                c3.border = borde

        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _nombre_hoja_seguro(nombre):
    """Limpia el nombre para usarlo como hoja de Excel (máx 31 chars, sin caracteres especiales)."""
    invalidos = r'[]\/?*:'
    for ch in invalidos:
        nombre = nombre.replace(ch, "")
    return nombre[:31]


# ══════════════════════════════════════════════════════════════════════════
# CONVERTIDOR BANCOLOMBIA — EXTRACTO PDF
# ══════════════════════════════════════════════════════════════════════════

def render_bancolombia():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">🏦 Bancolombia — Extracto PDF</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Convierte el extracto bancario PDF a Excel estructurado</p>
    </div>
    """, unsafe_allow_html=True)

    archivo = st.file_uploader(
        "Sube el archivo de extracto (.pdf)",
        type=["pdf"],
        key="up_banc_pdf",
        label_visibility="collapsed"
    )

    if archivo is None:
        st.info("📂 Sube el archivo PDF para continuar.")
        return

    try:
        datos     = _parsear_pdf_bancolombia(archivo)
    except Exception as e:
        st.error(f"❌ Error al leer el PDF: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return

    if not datos["registros"]:
        st.error("❌ No se encontraron movimientos en el archivo.")
        return

    st.success(
        f"✅ **{datos['empresa']}** | Cuenta: {datos['num_cuenta']} | "
        f"Fecha: {datos['fecha']} | **{len(datos['registros'])} movimientos**"
    )

    # Preview
    st.markdown("#### 📊 Vista previa")
    import pandas as pd
    df_prev = pd.DataFrame(datos["registros"])
    st.dataframe(df_prev.head(20), use_container_width=True, hide_index=True)

    # Nombre archivo
    nombre_xlsx = _nombre_bancolombia(archivo.name, datos["fecha"])
    excel_bytes = _generar_excel_bancolombia(datos)

    # Guardar automáticamente
    ruta_auto = r"C:\Users\ASUS\Desktop\BANCOS\CONVERTIDOR DE ARCHIVO"
    try:
        import os
        os.makedirs(ruta_auto, exist_ok=True)
        ruta_completa = os.path.join(ruta_auto, nombre_xlsx)
        with open(ruta_completa, "wb") as f:
            f.write(excel_bytes)
        st.success(f"💾 Guardado automáticamente en: {ruta_completa}")
    except Exception as e:
        st.warning(f"⚠️ No se pudo guardar automáticamente: {str(e)}")

    st.download_button(
        label=f"⬇️  Descargar Excel — {nombre_xlsx}",
        data=excel_bytes,
        file_name=nombre_xlsx,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="dl_banc_excel"
    )


def _nombre_bancolombia(nombre_archivo, fecha_datos):
    """
    ZIP_16600000605_000000901347233_20260704_08120932.pdf
    → BANCOLOMBIA0605_20260704.xlsx
    Toma los últimos 4 dígitos del segundo segmento y la fecha del tercer segmento.
    """
    import re
    nombre = nombre_archivo.replace(".pdf", "").replace(".PDF", "")
    partes = nombre.split("_")
    # partes[0]=ZIP, partes[1]=num_cuenta, partes[2]=nit, partes[3]=fecha
    if len(partes) >= 4:
        num_cuenta = partes[1][-4:]   # últimos 4 dígitos
        fecha_arch = partes[3][:8]    # YYYYMMDD
        return f"BANCOLOMBIA{num_cuenta}_{fecha_arch}.xlsx"
    return f"BANCOLOMBIA_{nombre}.xlsx"


def _parsear_pdf_bancolombia(archivo):
    """Parsea el PDF de extracto Bancolombia y retorna dict con registros."""
    import pdfplumber
    import re

    RE_VALOR = re.compile(r'(-?[\d,]+\.\d{2})$')
    RE_FECHA = re.compile(r'^(\d{4}/\d{2}/\d{2})\s+')
    RE_REFS  = re.compile(r'\s+(\d{7,15})\s+(\d{7,15})\s*$')
    RE_REF1  = re.compile(r'\s+(\d{7,15})\s*$')
    RE_ORPHAN = re.compile(r'^\d{3,5}$')  # líneas sueltas de 3-5 dígitos (teléfono NEQUI)

    empresa = num_cuenta = fecha = tipo_cuenta = ""

    archivo.seek(0)
    contenido = archivo.read()

    with pdfplumber.open(io.BytesIO(contenido)) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    # Extraer metadata del encabezado
    for line in all_lines[:10]:
        m = re.search(r'Empresa:\s*(.+?)\s+Número de Cuenta:\s*(\d+)', line)
        if m:
            empresa     = m.group(1).strip()
            num_cuenta  = m.group(2).strip()
        m2 = re.search(r'Fecha y Hora Actual:\s*(\d{2}-\d{2}-\d{4})', line)
        if m2:
            fecha = m2.group(1)
        m3 = re.search(r'Tipo de cuenta:\s*(.+?)(?:\s{2}|$)', line)
        if m3:
            tipo_cuenta = m3.group(1).strip()

    # Parsear movimientos
    header_found = False
    registros = []

    for line in all_lines:
        line_s = line.strip()

        if 'FECHA' in line_s and 'DESCRIPCIÓN' in line_s and 'VALOR' in line_s:
            header_found = True
            continue

        if not header_found or not line_s:
            continue
        if line_s.startswith('Página') or line_s.startswith('Saldo'):
            continue
        if RE_ORPHAN.match(line_s):
            continue  # línea suelta de número de NEQUI

        if not RE_FECHA.match(line_s):
            continue

        # Extraer fecha
        m_f = RE_FECHA.match(line_s)
        fecha_mov = m_f.group(1).replace('/', '-')
        resto = line_s[m_f.end():]

        # Extraer valor
        m_v = RE_VALOR.search(resto)
        if not m_v:
            continue
        valor = float(m_v.group(1).replace(',', ''))
        resto = resto[:m_v.start()].strip()

        # Extraer referencias
        ref1 = ref2 = ''
        m_r2 = RE_REFS.search(resto)
        if m_r2:
            ref1  = m_r2.group(1)
            ref2  = m_r2.group(2)
            resto = resto[:m_r2.start()].strip()
        else:
            m_r1 = RE_REF1.search(resto)
            if m_r1:
                ref1  = m_r1.group(1)
                resto = resto[:m_r1.start()].strip()

        # Separar descripcion y sucursal/canal
        partes = re.split(r'\s{2,}', resto, maxsplit=1)
        descripcion = partes[0].strip()
        sucursal    = partes[1].strip() if len(partes) > 1 else ''

        registros.append({
            'FECHA':           fecha_mov,
            'DESCRIPCION':     descripcion,
            'SUCURSAL/CANAL':  sucursal,
            'REFERENCIA 1':    ref1,
            'REFERENCIA 2':    ref2,
            'DOCUMENTO':       '',
            'VALOR':           valor,
        })

    return {
        "empresa":    empresa,
        "num_cuenta": num_cuenta,
        "fecha":      fecha,
        "tipo_cuenta": tipo_cuenta,
        "registros":  registros,
    }


def _generar_excel_bancolombia(datos):
    """Genera el Excel con todos los movimientos en una sola hoja."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    font_base   = Font(name="Arial", size=10)
    font_bold   = Font(name="Arial", size=10, bold=True)
    font_titulo = Font(name="Arial", size=12, bold=True)
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="1F4E78")
    fill_total  = PatternFill("solid", fgColor="D9E1F2")
    align_c     = Alignment(horizontal="center", vertical="center")
    align_l     = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="AAAAAA")
    borde       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título fila 1
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = (f"{datos['empresa']} | Cuenta: {datos['num_cuenta']} "
                   f"({datos['tipo_cuenta']}) | Fecha: {datos['fecha']}")
    c.font      = font_titulo
    c.alignment = align_l

    # Encabezados fila 3
    cols   = ["FECHA", "DESCRIPCIÓN", "SUCURSAL/CANAL", "REFERENCIA 1",
              "REFERENCIA 2", "DOCUMENTO", "VALOR"]
    anchos = [14, 45, 25, 16, 16, 14, 16]
    for ci, (h, w) in enumerate(zip(cols, anchos), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = align_c
        cell.border    = borde
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Datos desde fila 4
    for ri, reg in enumerate(datos["registros"], 4):
        vals = [
            reg["FECHA"],
            reg["DESCRIPCION"],
            reg["SUCURSAL/CANAL"],
            reg["REFERENCIA 1"],
            reg["REFERENCIA 2"],
            reg["DOCUMENTO"],
            reg["VALOR"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = font_base
            cell.border    = borde
            cell.alignment = align_c if ci != 2 else align_l
            if ci == 7:
                cell.number_format = "General"

    # Fila total
    n   = len(datos["registros"])
    ft  = n + 5
    ws.cell(row=ft, column=1, value="TOTAL MOVIMIENTOS").font  = font_bold
    ws.cell(row=ft, column=1).fill                             = fill_total
    ws.cell(row=ft, column=1).border                           = borde
    ws.cell(row=ft, column=7,
            value=f"=SUM(G4:G{n+3})").number_format            = "General"
    ws.cell(row=ft, column=7).font                             = font_bold
    ws.cell(row=ft, column=7).fill                             = fill_total
    ws.cell(row=ft, column=7).border                           = borde
    ws.cell(row=ft, column=2,
            value=f"=COUNTA(A4:A{n+3})").number_format         = "General"
    ws.cell(row=ft, column=2).font                             = font_bold
    ws.cell(row=ft, column=2).fill                             = fill_total
    ws.cell(row=ft, column=2).border                           = borde
    for ci in range(3, 8):
        c2 = ws.cell(row=ft, column=ci)
        if ci not in (7, 2):
            c2.fill   = fill_total
            c2.border = borde

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()