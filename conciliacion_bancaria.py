import pandas as pd
import numpy as np
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

# ══════════════════════════════════════════════════════════════════════════
# CATÁLOGO DE EMPRESAS Y CUENTAS
# ══════════════════════════════════════════════════════════════════════════
EMPRESAS_CUENTAS = {
    "SUPRECREDITO": [
        ("BANCOLOMBIA AHORRO No 0605",  112005001),
        ("DAVIVIENDA CTA CTE No 7563",  111005005),
        ("OCCIDENTE CTA CTE No 5816",   111005002),
        ("BOGOTA CT CTE No 0165",       111005007),
    ],
    "SUPREMOTOS": [
        ("BANCOLOMBIA AHORRO CTA No 7140", 112005001),
        ("DAVIVIENDA CTA CTE No 7571",     111005005),
        ("OCCIDENTE CTA CTE No 9045",      111005002),
        ("BOGOTA CTA CTE No 9696",         111005007),
    ],
    "AmanecerDePascua":     [("AMANECER DE PASCUA CTA CTE 5004",  111005002)],
    "AnochecerDePascua":    [("ANOCHECER DE PASCUA CTA CTE 4999", 111005002)],
    "MañanaDePascua":       [("MAÑANA DE PASCUA CTA CTE 7481",    111005002)],
    "GriGroup":             [("GRI CTA CTE 1525",                  111005002)],
    "Movicap":              [("MOVICAPSAS CTE 4395",               111005002),
                             ("AHO MOVICAP 2050",                  112005002)],
    "ConfianzaGlobal":      [("OCCIDENTE AHORRO CTA No 9890",      112005001)],
    "SeguroConfianzaGlobal":[("AHO SEGUROS CONFIANZA 2449",        112005002)],
    "Suprogreso":           [("BANCOLOMBIA CT AHO 9260",           111005001)],
}

MESES = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
         "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]


# ══════════════════════════════════════════════════════════════════════════
# PARSERS DE EXTRACTOS BANCARIOS
# ══════════════════════════════════════════════════════════════════════════

def detectar_y_parsear_extracto(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith('.csv'):
        return _parsear_bancolombia_csv(archivo)
    elif nombre.endswith('.xlsx'):
        return _parsear_davivienda_xlsx(archivo)
    elif nombre.endswith('.xls'):
        return _parsear_xls(archivo)
    else:
        raise ValueError(f"Formato no reconocido: {archivo.name}")


def _parsear_bancolombia_csv(archivo):
    content = archivo.read().decode('utf-8', errors='replace')
    rows = []
    for line in content.strip().split('\n'):
        parts = [p.strip() for p in line.split(',')]
        if len(parts) < 8:
            continue
        try:
            fecha_str = parts[3].strip()
            if len(fecha_str) == 8:
                fecha = datetime.strptime(fecha_str, '%Y%m%d')
            else:
                continue
            valor    = float(parts[5].strip().replace(' ', ''))
            concepto = parts[7].strip()
            debito   = round(abs(valor), 2) if valor < 0 else 0.0
            credito  = round(valor, 2)      if valor > 0 else 0.0
            rows.append({"FECHA": fecha, "DEBITO": debito, "CREDITO": credito, "CONCEPTO": concepto})
        except Exception:
            continue
    return pd.DataFrame(rows)


def _parsear_davivienda_xlsx(archivo):
    df = pd.read_excel(archivo, sheet_name=0, header=0)
    rows = []
    for _, row in df.iterrows():
        try:
            fecha    = pd.to_datetime(row.iloc[0], dayfirst=True)
            concepto = str(row.iloc[2]).strip()
            tipo     = str(row.iloc[3]).strip().lower()
            val_str  = str(row.iloc[7]).replace('$','').replace('.','').replace(',','.').strip()
            valor    = float(val_str)
            es_debito  = 'débito'  in tipo or 'debito'  in tipo
            es_credito = 'crédito' in tipo or 'credito' in tipo or 'deposito especial' in tipo
            debito   = round(valor, 2) if es_debito  else 0.0
            credito  = round(valor, 2) if es_credito else 0.0
            rows.append({"FECHA": fecha, "DEBITO": debito, "CREDITO": credito, "CONCEPTO": concepto})
        except Exception:
            continue
    return pd.DataFrame(rows)


def _parsear_xls(archivo):
    import xlrd
    content = archivo.read()
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    fila_datos = None
    tipo = None
    for i in range(min(30, ws.nrows)):
        row_str = ' '.join(str(v) for v in ws.row_values(i)).lower()
        if 'fecha movimiento' in row_str or ('débitos' in row_str and i < 10):
            fila_datos = i + 1; tipo = 'v1'; break
        if 'débitos' in row_str and i >= 20:
            fila_datos = i + 1; tipo = 'v2'; break

    if fila_datos is None:
        raise ValueError("No se pudo detectar el formato XLS.")

    col_cre_v2 = 13
    if tipo == 'v2' and ws.nrows > fila_datos:
        hdr = ws.row_values(fila_datos - 1)
        if len(hdr) > 14:
            if str(hdr[13]).strip() == "" and "cr" in str(hdr[14]).lower():
                col_cre_v2 = 14

    rows = []
    for i in range(fila_datos, ws.nrows):
        row = ws.row_values(i)
        try:
            if tipo == 'v1':
                fecha    = _parse_fecha_str(str(row[0]))
                concepto = str(row[2]).strip()
                deb_str  = str(row[4]).replace('$','').replace(',','').strip()
                cre_str  = str(row[5]).replace('$','').replace(',','').strip()
                debito   = round(float(deb_str or 0), 2)
                credito  = round(float(cre_str or 0), 2)
            else:
                fecha    = _parse_fecha_str(str(row[1]))
                concepto = str(row[4]).strip()
                debito   = round(float(row[12] or 0), 2)
                credito  = round(float(row[col_cre_v2] or 0) if len(row) > col_cre_v2 else 0, 2)
            if not fecha or (debito == 0 and credito == 0):
                continue
            rows.append({"FECHA": fecha, "DEBITO": debito, "CREDITO": credito, "CONCEPTO": concepto})
        except Exception:
            continue
    return pd.DataFrame(rows)


def _parse_fecha_str(s):
    s = s.strip()
    for fmt in ['%Y/%m/%d', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(s, fmt)
        except:
            continue
    return None


# ══════════════════════════════════════════════════════════════════════════
# PASO 1: FILTRAR DATOS (rápido, sin conciliar)
# ══════════════════════════════════════════════════════════════════════════

def filtrar_datos(archivos_aux, archivo_extracto, empresa, codigo_cuenta, mes_idx):
    """
    Solo filtra y lee los datos necesarios.
    Retorna (df_banco, df_aux, df_aux_original, resumen_filtro) o lanza excepción.
    """
    # 1. Leer y filtrar auxiliares
    mes_num = mes_idx + 1
    frames_aux = []
    frames_orig = []

    for f in archivos_aux:
        f.seek(0)
        try:
            df = pd.read_excel(f, sheet_name=0, header=0)
        except Exception:
            continue

        # Limpiar columnas NaN al inicio
        df = df.loc[:, ~df.columns.astype(str).str.match(r'^Unnamed|^nan$')]
        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

        # Detectar columna empresa
        col_emp = next((c for c in df.columns if 'empresa' in c), None)
        col_cta = next((c for c in df.columns if 'codigocuenta' in c or 'codigo_cuenta' in c), None)
        col_fec = next((c for c in df.columns if c == 'fecha'), None)

        if not col_emp or not col_cta or not col_fec:
            continue

        # Filtro empresa (case-insensitive)
        mask_emp = df[col_emp].astype(str).str.upper().str.strip() == empresa.upper().strip()
        df = df[mask_emp]

        # Filtro cuenta
        try:
            mask_cta = pd.to_numeric(df[col_cta], errors='coerce').astype('Int64') == int(codigo_cuenta)
            df = df[mask_cta]
        except Exception:
            continue

        # Filtro mes
        try:
            df[col_fec] = pd.to_datetime(df[col_fec], dayfirst=True, errors='coerce')
            mask_mes = df[col_fec].dt.month == mes_num
            df = df[mask_mes]
        except Exception:
            continue

        if len(df) == 0:
            continue

        frames_orig.append(df.copy())

        # Normalizar para conciliación
        col_deb = next((c for c in df.columns if c == 'debito'), None)
        col_cre = next((c for c in df.columns if c == 'credito'), None)
        col_des = next((c for c in df.columns if 'descripcion' in c or 'descripci' in c), None)

        df_norm = pd.DataFrame({
            "fecha":       df[col_fec].values,
            "descripcion": df[col_des].astype(str).values if col_des else [""] * len(df),
            "debito":      pd.to_numeric(df[col_deb], errors='coerce').fillna(0).round(2).values if col_deb else np.zeros(len(df)),
            "credito":     pd.to_numeric(df[col_cre], errors='coerce').fillna(0).round(2).values if col_cre else np.zeros(len(df)),
        })
        frames_aux.append(df_norm)

    if not frames_aux:
        raise ValueError(f"No se encontraron datos para {empresa} / cuenta {codigo_cuenta} / mes {MESES[mes_idx]}.")

    df_aux      = pd.concat(frames_aux,  ignore_index=True)
    df_aux_orig = pd.concat(frames_orig, ignore_index=True)

    # 2. Parsear extracto bancario
    archivo_extracto.seek(0)
    df_banco = detectar_y_parsear_extracto(archivo_extracto)

    if df_banco.empty:
        raise ValueError("El extracto bancario no tiene movimientos válidos.")

    resumen_filtro = {
        "n_banco": len(df_banco),
        "n_aux":   len(df_aux),
        "total_deb_banco":  round(df_banco["DEBITO"].sum(),  2),
        "total_cre_banco":  round(df_banco["CREDITO"].sum(), 2),
        "total_deb_aux":    round(df_aux["debito"].sum(),    2),
        "total_cre_aux":    round(df_aux["credito"].sum(),   2),
    }

    return df_banco, df_aux, df_aux_orig, resumen_filtro


# ══════════════════════════════════════════════════════════════════════════
# PASO 2: CONCILIAR (rápido con merge en vez de loops anidados)
# ══════════════════════════════════════════════════════════════════════════

def conciliar(df_banco, df_aux):
    """
    Cruce usando merge por valor exacto — mucho más rápido que loops anidados.
    banco.DEBITO  ↔ aux.credito  (lo que sale del banco entra en aux)
    banco.CREDITO ↔ aux.debito   (lo que entra al banco sale en aux)
    """
    df_b = df_banco.copy().reset_index(drop=True)
    df_a = df_aux.copy().reset_index(drop=True)

    df_b["DEBITO"]  = pd.to_numeric(df_b["DEBITO"],  errors="coerce").fillna(0).round(2)
    df_b["CREDITO"] = pd.to_numeric(df_b["CREDITO"], errors="coerce").fillna(0).round(2)
    df_a["debito"]  = pd.to_numeric(df_a["debito"],  errors="coerce").fillna(0).round(2)
    df_a["credito"] = pd.to_numeric(df_a["credito"], errors="coerce").fillna(0).round(2)

    df_b["_idx_b"] = df_b.index
    df_a["_idx_a"] = df_a.index

    b_match = set()
    a_match = set()

    def _match_por_valor_fecha(df_banco_sub, col_banco, df_aux_sub, col_aux):
        """
        Para cada movimiento del auxiliar busca en el banco el match de
        valor exacto con fecha más reciente (desempate por fecha DESC).
        """
        from collections import defaultdict

        # Forzar copia y conversión de FECHA a datetime para poder ordenar
        b_sub = df_banco_sub.copy()
        b_sub["FECHA"] = pd.to_datetime(b_sub["FECHA"], errors="coerce")

        # Construir lookup: valor → lista de (fecha, idx_b) ordenados por fecha DESC
        lookup = defaultdict(list)
        for _, rb in b_sub.iterrows():
            ib  = int(rb["_idx_b"])
            val = round(float(rb[col_banco]), 2)
            fec = rb["FECHA"]
            lookup[val].append((fec, ib))

        # Ordenar cada lista por fecha DESC
        for val in lookup:
            lookup[val].sort(key=lambda x: x[0] if pd.notna(x[0]) else pd.Timestamp.min,
                             reverse=True)

        pares = []
        for _, ra in df_aux_sub.iterrows():
            ia  = int(ra["_idx_a"])
            val = round(float(ra[col_aux]), 2)
            candidatos = lookup.get(val, [])
            for i, (fec, ib) in enumerate(candidatos):
                if ib not in b_match and ia not in a_match:
                    pares.append((ib, ia))
                    b_match.add(ib)
                    a_match.add(ia)
                    candidatos.pop(i)  # eliminar para no reutilizar
                    break
        return pares

    # ── DEBUG: mostrar fechas del banco antes del cruce ─────────────────
    import streamlit as st
    st.warning("🔍 DEBUG fechas banco (primeras 10 filas):")
    st.dataframe(df_b[["FECHA","DEBITO","CREDITO","CONCEPTO"]].head(10))
    st.caption(f"Tipo FECHA: {df_b['FECHA'].dtype}")

    # ── Cruce 1: banco DÉBITO ↔ aux CRÉDITO ──────────────────────────────
    b_deb_sub = df_b[df_b["DEBITO"]  > 0]
    a_cre_sub = df_a[df_a["credito"] > 0]
    _match_por_valor_fecha(b_deb_sub, "DEBITO", a_cre_sub, "credito")

    # ── Cruce 2: banco CRÉDITO ↔ aux DÉBITO ──────────────────────────────
    b_cre_sub = df_b[(df_b["CREDITO"] > 0) & (~df_b["_idx_b"].isin(b_match))]
    a_deb_sub = df_a[(df_a["debito"]  > 0) & (~df_a["_idx_a"].isin(a_match))]
    _match_por_valor_fecha(b_cre_sub, "CREDITO", a_deb_sub, "debito")

    # ── Reclasificaciones auxiliar (no cruzado: deb↔cre mismo valor) ─────
    a_no_match = df_a[~df_a["_idx_a"].isin(a_match)].copy()
    a_reclass  = set()

    a_nd = a_no_match[a_no_match["debito"]  > 0][["_idx_a", "debito"]].rename(columns={"debito":  "_val"})
    a_nc = a_no_match[a_no_match["credito"] > 0][["_idx_a", "credito"]].rename(columns={"credito": "_val"})
    merged_rc = pd.merge(a_nd, a_nc, on="_val", suffixes=("_d", "_c"), how="inner")

    seen_d, seen_c = set(), set()
    for _, row in merged_rc.iterrows():
        id_, ic = int(row["_idx_a_d"]), int(row["_idx_a_c"])
        if id_ not in seen_d and ic not in seen_c and id_ not in a_reclass and ic not in a_reclass:
            a_reclass.add(id_); a_reclass.add(ic)
            seen_d.add(id_);    seen_c.add(ic)

    # ── Totales ───────────────────────────────────────────────────────────
    saldo_aux_deb   = round(df_a["debito"].sum(),  2)
    saldo_aux_cred  = round(df_a["credito"].sum(), 2)
    saldo_banco_deb  = round(df_b["DEBITO"].sum(),  2)
    saldo_banco_cred = round(df_b["CREDITO"].sum(), 2)

    # ── Construir partidas conciliatorias ─────────────────────────────────
    partidas = []

    # 2A: Reclasificaciones
    a_rd = a_no_match[a_no_match["_idx_a"].isin(a_reclass) & (a_no_match["debito"]  > 0)]
    a_rc = a_no_match[a_no_match["_idx_a"].isin(a_reclass) & (a_no_match["credito"] > 0)]

    val_rd = a_rd[["_idx_a", "debito"]].rename(columns={"debito": "_val"})
    val_rc = a_rc[["_idx_a", "credito"]].rename(columns={"credito": "_val"})
    pares_rc = pd.merge(val_rd, val_rc, on="_val", suffixes=("_d", "_c"))

    done_d, done_c = set(), set()
    for _, row in pares_rc.iterrows():
        id_, ic = int(row["_idx_a_d"]), int(row["_idx_a_c"])
        if id_ in done_d or ic in done_c:
            continue
        done_d.add(id_); done_c.add(ic)
        val = float(row["_val"])
        row_d = df_a.loc[id_]
        row_c = df_a.loc[ic]
        partidas.append({
            "tipo": "RECLASIFICACION",
            "conc_aux": str(row_d.get("descripcion", "")),
            "fecha_aux": row_d.get("fecha"),
            "deb": -val, "cred": 0,
            "conc_banco": "", "fecha_banco": None,
        })
        partidas.append({
            "tipo": "RECLASIFICACION",
            "conc_aux": "",
            "fecha_aux": row_c.get("fecha"),
            "deb": 0, "cred": -val,
            "conc_banco": str(row_c.get("descripcion", "")),
            "fecha_banco": row_c.get("fecha"),
        })

    # 2B: Solo banco (falta en auxiliar)
    for ib in df_b[~df_b["_idx_b"].isin(b_match)]["_idx_b"]:
        row_b = df_b.loc[ib]
        bd = float(row_b["DEBITO"])
        bc = float(row_b["CREDITO"])
        conc  = str(row_b["CONCEPTO"])
        fecha = row_b["FECHA"]
        if bd > 0:
            partidas.append({"tipo": "SOLO_BANCO", "conc_aux": "", "fecha_aux": None,
                             "deb": 0, "cred": bd, "conc_banco": conc, "fecha_banco": fecha})
        else:
            partidas.append({"tipo": "SOLO_BANCO", "conc_aux": conc, "fecha_aux": fecha,
                             "deb": bc, "cred": 0, "conc_banco": "", "fecha_banco": None})

    # 2C: Solo auxiliar (está de más)
    for ia in df_a[~df_a["_idx_a"].isin(a_match) & ~df_a["_idx_a"].isin(a_reclass)]["_idx_a"]:
        row_a = df_a.loc[ia]
        ad = float(row_a["debito"])
        ac = float(row_a["credito"])
        conc  = str(row_a.get("descripcion", ""))
        fecha = row_a.get("fecha")
        if ad > 0:
            partidas.append({"tipo": "SOLO_AUX", "conc_aux": conc, "fecha_aux": fecha,
                             "deb": -ad, "cred": 0, "conc_banco": "", "fecha_banco": None})
        else:
            partidas.append({"tipo": "SOLO_AUX", "conc_aux": "", "fecha_aux": None,
                             "deb": 0, "cred": -ac, "conc_banco": conc, "fecha_banco": fecha})

    # ── Totales finales ───────────────────────────────────────────────────
    total_conc_deb  = round(saldo_aux_deb  + sum(p["deb"]  for p in partidas), 2)
    total_conc_cred = round(saldo_aux_cred + sum(p["cred"] for p in partidas), 2)
    diferencia_deb  = round(total_conc_deb  - saldo_banco_cred, 2)
    diferencia_cred = round(total_conc_cred - saldo_banco_deb,  2)

    resumen = {
        "saldo_aux_deb":     saldo_aux_deb,
        "saldo_aux_cred":    saldo_aux_cred,
        "saldo_banco_deb":   saldo_banco_deb,
        "saldo_banco_cred":  saldo_banco_cred,
        "total_conc_deb":    total_conc_deb,
        "total_conc_cred":   total_conc_cred,
        "diferencia_deb":    diferencia_deb,
        "diferencia_cred":   diferencia_cred,
        "n_reclasif":        sum(1 for p in partidas if p["tipo"] == "RECLASIFICACION"),
        "n_solo_banco":      sum(1 for p in partidas if p["tipo"] == "SOLO_BANCO"),
        "n_solo_aux":        sum(1 for p in partidas if p["tipo"] == "SOLO_AUX"),
        "conciliados_banco": len(b_match),
        "conciliados_aux":   len(a_match),
    }

    return partidas, df_b.drop(columns=["_idx_b"]), df_a.drop(columns=["_idx_a"]), resumen


# ══════════════════════════════════════════════════════════════════════════
# GENERACIÓN DEL EXCEL (3 hojas: CONCILIACION, BANCO, AUXILIAR)
# ══════════════════════════════════════════════════════════════════════════

def generar_excel(partidas, df_banco, df_aux, resumen, empresa, cuenta, mes, df_aux_original=None):
    wb = openpyxl.Workbook()
    _hoja_conciliacion(wb, partidas, resumen, empresa, cuenta, mes)
    _hoja_banco(wb, df_banco)
    _hoja_auxiliar(wb, df_aux, df_aux_original)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _hoja_conciliacion(wb, partidas, resumen, empresa, cuenta, mes):
    ws = wb.active
    ws.title = "CONCILIACION"

    azul     = PatternFill("solid", fgColor="BDD7EE")
    verde    = PatternFill("solid", fgColor="C6E0B4")
    amarillo = PatternFill("solid", fgColor="FFEB9C")
    rojo     = PatternFill("solid", fgColor="FFC7CE")
    bold     = Font(bold=True)
    fmt_num  = '#,##0.00;(#,##0.00);"-"'
    fmt_date = "DD/MM/YYYY"

    # Encabezados fila 1
    hdrs = ["OBSERVACION", "CONCEPTO", "FECHA", "DEBITO", "CREDITO",
            "FECHA", "CONCEPTO", "OBSERVACION"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=col, value=h).font = bold

    # Fila 2: SALDO AUXILIAR
    ws.cell(row=2, column=2, value="SALDO AUXILIAR")
    c4 = ws.cell(row=2, column=4, value=resumen["saldo_aux_deb"])
    c4.number_format = fmt_num
    c5 = ws.cell(row=2, column=5, value=resumen["saldo_aux_cred"])
    c5.number_format = fmt_num
    ws.cell(row=2, column=7, value="SALDO AUXILIAR")
    for col in range(1, 9):
        ws.cell(row=2, column=col).fill = azul
        ws.cell(row=2, column=col).font = bold

    # Partidas
    row = 3
    for p in partidas:
        if p["conc_aux"]:
            ws.cell(row=row, column=2, value=p["conc_aux"])
        if p["fecha_aux"]:
            try:
                ws.cell(row=row, column=3, value=pd.Timestamp(p["fecha_aux"]).to_pydatetime())
                ws.cell(row=row, column=3).number_format = fmt_date
            except Exception:
                pass
        if p["deb"] != 0:
            ws.cell(row=row, column=4, value=p["deb"]).number_format = fmt_num
        if p["cred"] != 0:
            ws.cell(row=row, column=5, value=p["cred"]).number_format = fmt_num
        if p["fecha_banco"]:
            try:
                ws.cell(row=row, column=6, value=pd.Timestamp(p["fecha_banco"]).to_pydatetime())
                ws.cell(row=row, column=6).number_format = fmt_date
            except Exception:
                pass
        if p["conc_banco"]:
            ws.cell(row=row, column=7, value=p["conc_banco"])
        row += 1

    # TOTAL AUXILIAR CONCILIADO
    ws.cell(row=row, column=2, value="TOTAL AUXILIAR CONCILIADO")
    ws.cell(row=row, column=4, value=resumen["total_conc_deb"]).number_format  = fmt_num
    ws.cell(row=row, column=5, value=resumen["total_conc_cred"]).number_format = fmt_num
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = verde
        ws.cell(row=row, column=col).font = bold
    row += 1

    # SALDO BANCO
    ws.cell(row=row, column=2, value="SALDO BANCO")
    ws.cell(row=row, column=4, value=resumen["saldo_banco_cred"]).number_format = fmt_num
    ws.cell(row=row, column=5, value=resumen["saldo_banco_deb"]).number_format  = fmt_num
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = amarillo
        ws.cell(row=row, column=col).font = bold
    row += 1

    # DIFERENCIA
    ws.cell(row=row, column=2, value="DIFERENCIA")
    ws.cell(row=row, column=4, value=resumen["diferencia_deb"]).number_format  = fmt_num
    ws.cell(row=row, column=5, value=resumen["diferencia_cred"]).number_format = fmt_num
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = rojo
        ws.cell(row=row, column=col).font = bold

    # Anchos
    for col, w in zip("ABCDEFGH", [18, 50, 14, 16, 16, 14, 50, 18]):
        ws.column_dimensions[col].width = w


def _hoja_banco(wb, df_banco):
    ws = wb.create_sheet("BANCO")
    headers = ["FECHA", "DEBITO", "CREDITO", "CONCEPTO"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for r, row in enumerate(df_banco.itertuples(index=False), 2):
        try:
            ws.cell(row=r, column=1, value=row.FECHA).number_format = "DD/MM/YYYY"
        except:
            ws.cell(row=r, column=1, value=str(row.FECHA))
        ws.cell(row=r, column=2, value=row.DEBITO)
        ws.cell(row=r, column=3, value=row.CREDITO)
        ws.cell(row=r, column=4, value=row.CONCEPTO)


def _hoja_auxiliar(wb, df_aux, df_aux_original=None):
    ws  = wb.create_sheet("AUXILIAR")
    df  = df_aux_original if df_aux_original is not None else df_aux
    bold = Font(bold=True)
    for col, h in enumerate(df.columns, 1):
        ws.cell(row=1, column=col, value=h).font = bold
    for r, row in enumerate(df.itertuples(index=False), 2):
        for col, val in enumerate(row, 1):
            try:
                ws.cell(row=r, column=col, value=val)
            except Exception:
                ws.cell(row=r, column=col, value=str(val))
