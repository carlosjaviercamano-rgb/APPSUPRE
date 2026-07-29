import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io
import os
import streamlit as st

from conciliacion_bancaria import EMPRESAS_CUENTAS

# NIT de cada banco (aplica igual para débito y crédito)
NIT_BANCOS = {
    "BANCOLOMBIA": "890903938",
    "DAVIVIENDA":  "860034313",
    "OCCIDENTE":   "890300279",
    "BOGOTA":      "860002964",
}

TIPO_TERCERO_POR_BANCO = {
    "OCCIDENTE":   "PIT",
    "BANCOLOMBIA": "NIT",
    "DAVIVIENDA":  "NIT",
    "BOGOTA":      "NIT",
}
CENTRO_COSTO   = "102"
LIMITE_MOVIMIENTOS = 480

# Cuenta de débito según el concepto de cargue detectado
CUENTAS_DEBITO_GASTOS = {
    "GMF 4X1000": "539595002",
    "IVA":        "240805002",
    "COMISION":   "530515001",
    "SOBRE GIRO": "530520002",
}

# Palabras clave para detectar el concepto de cargue a partir del
# concepto tal como viene del banco. Se va ampliando por banco a medida
# que aparecen nuevas variantes de texto.
_PALABRAS_CLAVE = [
    (("SERVICIO", "COMISION", "COMISIÓN", "COMIS"),          "COMISION"),
    (("IVA",),                                                "IVA"),
    (("GMF", "GRAVAMEN", "4X1000", "IMPTO GOBIERNO"),         "GMF 4X1000"),
    (("GIRO",),                                                "SOBRE GIRO"),
]


def detectar_concepto_cargue(concepto_banco):
    """Detecta el CONCEPTO CARGUE a partir del texto del concepto del banco."""
    texto = str(concepto_banco).upper()
    for palabras, concepto in _PALABRAS_CLAVE:
        if any(p in texto for p in palabras):
            return concepto
    return None


def obtener_cuenta_credito(empresa, banco):
    """
    Busca en EMPRESAS_CUENTAS la cuenta bancaria de la empresa cuyo
    nombre coincide con el banco seleccionado.
    """
    cuentas_emp = EMPRESAS_CUENTAS.get(empresa, [])
    banco_up = banco.upper()
    for nombre_cuenta, codigo in cuentas_emp:
        if banco_up in nombre_cuenta.upper():
            return str(codigo)
    return None


def crear_gastos_bancarios(df_tabla, empresa, banco, config=None):
    """
    Genera el archivo de Gastos Bancarios (hoja Items, débitos y créditos)
    a partir de la tabla pegada por el usuario (FECHA, VALOR, CONCEPTO).

    No se agrupa por fecha: todos los movimientos pegados se procesan
    como un solo lote y solo se dividen en varios archivos si superan
    los 480 movimientos (igual que el límite de 150 filas de los planos
    de Aplicación de Pagos, pero con este tope específico).
    """
    if df_tabla is None or df_tabla.empty:
        raise ValueError("No hay filas en la tabla para generar el archivo.")

    nit_banco = NIT_BANCOS.get(banco.upper())
    if not nit_banco:
        raise ValueError(f"No se encontró el NIT configurado para el banco '{banco}'.")
    tipo_tercero = TIPO_TERCERO_POR_BANCO.get(banco.upper(), "NIT")

    cuenta_credito = obtener_cuenta_credito(empresa, banco)
    if not cuenta_credito:
        raise ValueError(
            f"No se encontró una cuenta de '{banco}' configurada para la empresa '{empresa}'."
        )

    df = df_tabla.copy()
    # Descartar filas totalmente vacías (sin fecha ni valor ni concepto)
    df = df[
        df["FECHA"].astype(str).str.strip().ne("") |
        df["VALOR"].astype(str).str.strip().ne("") |
        df["CONCEPTO"].astype(str).str.strip().ne("")
    ].reset_index(drop=True)
    if df.empty:
        raise ValueError("No hay filas con datos para generar el archivo.")

    # Parsear fecha y valor tal como se pegaron, sin depender de un tipo
    # de columna estricto (evita que Streamlit reemplace fechas/valores).
    df["_FECHA_PARSED"] = _parsear_fecha_mixta(df["FECHA"])
    df["_VALOR_PARSED"] = df["VALOR"].apply(_num)
    df["_CONCEPTO_CARGUE"] = df["CONCEPTO"].apply(detectar_concepto_cargue)

    filas_sin_fecha    = int(df["_FECHA_PARSED"].isna().sum())
    filas_sin_concepto = df.loc[df["_CONCEPTO_CARGUE"].isna(), "CONCEPTO"].tolist()

    df_validas = df[df["_CONCEPTO_CARGUE"].notna() & df["_FECHA_PARSED"].notna()].reset_index(drop=True)

    if filas_sin_fecha:
        st.warning(f"⚠️ {filas_sin_fecha} fila(s) no tienen una fecha válida y se excluyeron.")
    if filas_sin_concepto:
        st.warning(
            "⚠️ No se pudo determinar el concepto de cargue para "
            f"{len(filas_sin_concepto)} fila(s) (no contienen 'servicio/comisión', "
            "'iva', 'gmf/gravamen' ni 'giro'). Esas filas no se incluyeron en el "
            "archivo:\n" + "\n".join(f"- {c}" for c in filas_sin_concepto[:10])
        )

    if df_validas.empty:
        raise ValueError("Ninguna fila válida pudo procesarse (revisa fecha y concepto).")

    total = len(df_validas)
    n_bloques = max(1, -(-total // LIMITE_MOVIMIENTOS))  # ceil division
    hora_str = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    archivos_generados = []

    for bloque_idx in range(n_bloques):
        inicio = bloque_idx * LIMITE_MOVIMIENTOS
        fin = inicio + LIMITE_MOVIMIENTOS
        df_bloque = df_validas.iloc[inicio:fin].reset_index(drop=True)

        filas_debito  = []
        filas_credito = []

        for _, row in df_bloque.iterrows():
            valor        = row["_VALOR_PARSED"]
            concepto_cargue = row["_CONCEPTO_CARGUE"]
            concepto_src = str(row.get("CONCEPTO", "") or "").strip()
            fecha_str    = _fecha_str(row["_FECHA_PARSED"])
            detalle = f"{banco.upper()} {fecha_str} {concepto_src}".strip()
            cuenta_debito = CUENTAS_DEBITO_GASTOS[concepto_cargue]

            cod_impuesto = None
            valor_base   = None
            pct_impuesto = None
            if concepto_cargue == "IVA":
                cod_impuesto = "01"
                pct_impuesto = 0.19
                valor_base   = round(valor / 0.19, 7) if valor else 0

            filas_debito.append({
                "codigoCentroCosto": CENTRO_COSTO,
                "dniTercero": nit_banco, "codigoTipoDniTercero": tipo_tercero,
                "codigoCuenta": cuenta_debito, "valor": valor,
                "factura": "", "fechaVencimiento": None,
                "codigoImpuesto": cod_impuesto, "valorBaseImpuesto": valor_base,
                "porcentajeImpuesto": pct_impuesto, "detalle": detalle,
            })
            filas_credito.append({
                "codigoCentroCosto": CENTRO_COSTO,
                "dniTercero": nit_banco, "codigoTipoDniTercero": tipo_tercero,
                "codigoCuenta": cuenta_credito, "valor": -abs(valor),
                "factura": "", "fechaVencimiento": None,
                "codigoImpuesto": None, "valorBaseImpuesto": None,
                "porcentajeImpuesto": None, "detalle": detalle,
            })

        # Numerar el Id de forma secuencial según el orden final en el archivo
        # (todos los débitos primero, luego todos los créditos)
        todas_las_filas = filas_debito + filas_credito
        for idx, fila in enumerate(todas_las_filas, start=1):
            fila["Id"] = idx

        sufijo = f"_{bloque_idx + 1}" if n_bloques > 1 else ""
        nombre = f"GASTOS_BANCARIOS_{banco.upper()}_{empresa}_{hora_str}{sufijo}.xlsx"
        buffer = _generar_excel(filas_debito + filas_credito)

        ruta_auto = config.get("ruta_gastos_bancarios", "") if config else ""
        if ruta_auto:
            try:
                os.makedirs(ruta_auto, exist_ok=True)
                ruta_completa = os.path.join(ruta_auto, nombre)
                buffer.seek(0)
                with open(ruta_completa, "wb") as f:
                    f.write(buffer.read())
                st.success(f"💾 {nombre}: guardado en {ruta_completa}")
            except Exception as e:
                st.warning(f"⚠️ {nombre}: no se pudo guardar automáticamente — {str(e)}")
            buffer.seek(0)

        archivos_generados.append({"nombre": nombre, "buffer": buffer})

    st.markdown("#### 📥 Descargar archivo(s) generado(s):")
    for arch in archivos_generados:
        arch["buffer"].seek(0)
        st.download_button(
            label=f"⬇️  Descargar {arch['nombre']}",
            data=arch["buffer"],
            file_name=arch["nombre"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_gastos_{arch['nombre']}",
        )

    return f"{total} movimiento(s) procesados en {len(archivos_generados)} archivo(s)."


def _generar_excel(filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    encabezados = [
        "Id", "codigoCentroCosto", "dniTercero", "codigoTipoDniTercero",
        "codigoCuenta", "valor", "factura", "fechaVencimiento",
        "codigoImpuesto", "valorBaseImpuesto", "porcentajeImpuesto", "detalle"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for fila_idx, fila in enumerate(filas, start=2):
        for col_idx, col_name in enumerate(encabezados, start=1):
            ws.cell(row=fila_idx, column=col_idx, value=fila.get(col_name))
        if fila.get("fechaVencimiento"):
            ws.cell(row=fila_idx, column=8).number_format = "DD/MM/YYYY"

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _num(val):
    """
    Convierte a float respetando el valor tal cual se pegó.
    Soporta '7500,55' (coma decimal), '7.500,55' (miles con punto,
    decimal con coma) y '7500.55' (punto decimal).
    """
    s = str(val).strip()
    if s in ("", "nan", "None"):
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except Exception:
        return 0.0


def _parsear_fecha_mixta(serie):
    """
    Parsea fechas pegadas tal cual, soportando DD/MM/YYYY (Excel local)
    e ISO YYYY-MM-DD, sin depender de un tipo de columna estricto de
    Streamlit (evita que se reemplacen por la fecha de hoy).
    """
    s = serie.astype(str).str.strip()
    es_iso = s.str.match(r"^\d{4}-\d{1,2}-\d{1,2}")
    resultado = pd.Series(pd.NaT, index=serie.index, dtype="datetime64[ns]")
    if es_iso.any():
        resultado.loc[es_iso] = pd.to_datetime(s[es_iso], errors="coerce")
    if (~es_iso).any():
        resultado.loc[~es_iso] = pd.to_datetime(s[~es_iso], dayfirst=True, errors="coerce")
    return resultado


def _fecha_str(val):
    if val is None:
        return ""
    try:
        return pd.Timestamp(val).strftime("%d-%m-%Y")
    except Exception:
        return str(val)[:10] if val else ""