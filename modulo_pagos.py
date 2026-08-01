import streamlit as st
import pandas as pd
import os

# ─── Estilos del módulo ────────────────────────────────────────────────────
STYLES = """
<style>
/* Selector de tipo de pago */
.tipo-pago-card {
    border: 2px solid #e2e8f0;
    border-radius: 10px;
    padding: 1.2rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.15s;
    background: white;
}
.tipo-pago-card:hover { border-color: #1e40af; background: #eff6ff; }
.tipo-pago-card.selected { border-color: #1e40af; background: #eff6ff; }
.tipo-pago-card .tp-icon { font-size: 2rem; }
.tipo-pago-card .tp-title { font-weight: 600; color: #1e293b; font-size: 0.95rem; margin-top: 0.4rem; }
.tipo-pago-card .tp-sub { color: #64748b; font-size: 0.78rem; margin-top: 0.2rem; }

/* Tabla interactiva */
.tabla-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 0.8rem;
}
.tabla-header h3 { margin: 0; font-size: 1rem; color: #1e293b; }
.badge-count {
    background: #dbeafe;
    color: #1e40af;
    font-size: 0.78rem;
    font-weight: 600;
    padding: 0.2rem 0.6rem;
    border-radius: 20px;
}

/* Pasos del proceso */
.paso-container {
    display: flex;
    align-items: flex-start;
    gap: 1rem;
    margin-bottom: 1.5rem;
}
.paso-num {
    background: #1e40af;
    color: white;
    width: 28px;
    height: 28px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.8rem;
    font-weight: 700;
    flex-shrink: 0;
    margin-top: 0.1rem;
}
.paso-content h4 { margin: 0 0 0.2rem 0; font-size: 0.92rem; color: #1e293b; }
.paso-content p  { margin: 0; font-size: 0.82rem; color: #64748b; }

/* Botones de acción */
.btn-accion {
    background: #1e40af;
    color: white;
    border: none;
    border-radius: 8px;
    padding: 0.55rem 1.4rem;
    font-weight: 600;
    font-size: 0.88rem;
    cursor: pointer;
}
.btn-secundario {
    background: white;
    color: #1e40af;
    border: 1.5px solid #1e40af;
    border-radius: 8px;
    padding: 0.5rem 1.2rem;
    font-weight: 600;
    font-size: 0.88rem;
    cursor: pointer;
}

/* Alerta info */
.alerta-info {
    background: #eff6ff;
    border-left: 4px solid #1e40af;
    border-radius: 4px;
    padding: 0.8rem 1rem;
    font-size: 0.85rem;
    color: #1e40af;
    margin: 0.5rem 0;
}
</style>
"""

# ─── Columnas de AREA DE BANCO ─────────────────────────────────────────────
COLUMNAS_AREA_BANCO = [
    "ENTIDAD", "FECHA", "CEDULA", "VALOR", "T_TRANSACCION",
    "NUM_FACTURA", "CUOTA", "RECIBO", "DIFERENCIA",
    "VALOR_CB", "INMOVILIZACION", "OTROS_GASTOS",
    "OBSERVACION", "CORRESPONSAL", "FECHA_DOCUMENTO"
]

def init_estado():
    defaults = {
        "tipo_pago": None,
        "fecha_recaudo": None,
        "df_area_banco": None,
        "df_sheet1": None,
        "archivo_libro": None,
        "archivo_clientes": None,
        "archivo_corresponsal": None,
        "paso_actual": 1,
        "planos_generados": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def render():
    init_estado()
    st.markdown(STYLES, unsafe_allow_html=True)

    st.markdown("""
    <div class="module-header">
        <div class="module-icon">💳</div>
        <div>
            <h1>Aplicación y Compensación de Pagos</h1>
            <p>Diaria y pendiente por ID</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if "submodulo_pagos" not in st.session_state:
        st.session_state.submodulo_pagos = None

    sub = st.session_state.submodulo_pagos

    if sub is None:
        _render_menu_submodulos_pagos()
        return

    if st.button("← Volver a Aplicación y Compensación", key="btn_volver_pagos"):
        st.session_state.submodulo_pagos = None
        st.rerun()
    st.markdown("<br>", unsafe_allow_html=True)

    if sub == "diaria":
        st.markdown("### 📅 Aplicación y Compensación Diaria")
        tab1, tab2, tab3, tab4 = st.tabs([
            "📥 1. Carga de Archivos",
            "📊 2. Tabla de Pagos",
            "🔗 3. Alistar Información",
            "📁 4. Generar Archivos"
        ])

        with tab1:
            render_carga_archivos()
        with tab2:
            render_tabla_pagos()
        with tab3:
            render_alistar_informacion()
        with tab4:
            render_generar_archivos()

    elif sub == "pendiente":
        render_pendiente_por_id()


def _render_menu_submodulos_pagos():
    st.markdown("### Selecciona el tipo de aplicación")
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">📅</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Aplicación y Compensación Diaria</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Carga, tabla de pagos, alistar información y generar archivos</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_diaria", use_container_width=True, type="primary"):
            st.session_state.submodulo_pagos = "diaria"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">🆔</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem;font-size:1rem">
                Aplicación y Compensación Pendiente por ID</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Pagos antes marcados NO EXISTE, ya identificados</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_pendiente", use_container_width=True, type="primary"):
            st.session_state.submodulo_pagos = "pendiente"
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════
# TAB 1 — CARGA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════
GITHUB_RAW = "https://raw.githubusercontent.com/carlosjaviercamano-rgb/APPSUPRE/main"

def _cargar_desde_github(nombre_archivo):
    import requests
    import io
    url = f"{GITHUB_RAW}/{nombre_archivo}"
    resp = requests.get(url)
    if resp.status_code == 200:
        return io.BytesIO(resp.content)
    return None


def render_carga_archivos():
    st.markdown("### Archivos necesarios para el proceso")
    st.markdown('<div class="alerta-info">📌 Los archivos de Clientes Activos y Corresponsal se cargan automáticamente desde el repositorio. Solo debes subir el Libro de Banco cada vez que vayas a trabajar.</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    if st.session_state.archivo_clientes is None:
        archivo = _cargar_desde_github("CLIENTES_ACTIVOS.xlsx")
        if archivo:
            st.session_state.archivo_clientes = archivo

    if st.session_state.archivo_corresponsal is None:
        archivo = _cargar_desde_github("CORRESPONSAL.xlsx")
        if archivo:
            st.session_state.archivo_corresponsal = archivo

    st.markdown("**📂 Libro de Banco**")
    st.caption("Descarga automáticamente desde SharePoint o súbelo manualmente.")

    # ── Opción SharePoint (solo versión local) ────────────────────────────
    try:
        from sharepoint_connector import render_sharepoint_widget, tiene_token
        col_sp1, col_sp2 = st.columns([3, 1])
        with col_sp1:
            if tiene_token():
                st.success("✅ SharePoint conectado — puedes descargar el Libro automáticamente.")
            else:
                st.info("🔗 Conecta SharePoint para descargar el Libro automáticamente.")
        with col_sp2:
            if st.button("🔗 SharePoint", key="btn_abrir_sp", use_container_width=True):
                st.session_state["mostrar_sp"] = not st.session_state.get("mostrar_sp", False)
                st.rerun()

        if st.session_state.get("mostrar_sp", False):
            render_sharepoint_widget()
    except ImportError:
        pass

    # ── Carga automática desde OneDrive local ────────────────────────────────────
    RUTA_ONEDRIVE = (
        "C:\\Users\\ASUS\\OneDrive - SUPRECREDITO SAS\\"
        "Departamento administrativo y financiero - LIBRO DE TRABAJO\\"
        "LIBRO BANCOS 2025.xlsx"
    )
    if st.button("📂 Cargar desde OneDrive", type="primary",
                 use_container_width=True, key="btn_onedrive"):
        try:
            import io as _io
            with open(RUTA_ONEDRIVE, "rb") as f:
                contenido = f.read()
            archivo_od      = _io.BytesIO(contenido)
            archivo_od.name = "LIBRO BANCOS 2025.xlsx"
            st.session_state.archivo_libro = archivo_od
            st.success("✅ Libro de Banco cargado desde OneDrive automáticamente.")
            st.rerun()
        except Exception as e:
            st.error(f"❌ No se pudo cargar: {str(e)}")

    st.markdown("**⬆️ O súbelo manualmente:**")
    archivo_libro = st.file_uploader(
        "Subir libro de banco",
        type=["xlsx", "xlsm"],
        key="up_libro",
        label_visibility="collapsed"
    )
    if archivo_libro:
        st.session_state.archivo_libro = archivo_libro
        st.success(f"✅ {archivo_libro.name}")

    st.markdown("---")
    libro_ok        = st.session_state.archivo_libro is not None
    clientes_ok     = st.session_state.archivo_clientes is not None
    corresponsal_ok = st.session_state.archivo_corresponsal is not None

    c1, c2, c3 = st.columns(3)
    c1.metric("Libro de Banco",   "✅ Cargado"    if libro_ok        else "⚠️ Pendiente")
    c2.metric("Clientes Activos", "✅ Automático" if clientes_ok     else "⚠️ No encontrado")
    c3.metric("Corresponsal",     "✅ Automático" if corresponsal_ok else "⚠️ No encontrado")

    if libro_ok and clientes_ok and corresponsal_ok:
        st.success("🎉 Todo listo. Ve a la pestaña **2. Tabla de Pagos** para continuar.")
    elif not clientes_ok or not corresponsal_ok:
        st.warning("⚠️ No se pudieron cargar los archivos de referencia desde GitHub.")

    with st.expander("🔧 Actualizar archivos de referencia"):
        st.caption("Solo usa esto si necesitas reemplazar manualmente Clientes Activos o Corresponsal.")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**👥 Clientes Activos**")
            nuevo_clientes = st.file_uploader(
                "Reemplazar clientes activos",
                type=["xlsx"],
                key="up_clientes_manual",
                label_visibility="collapsed"
            )
            if nuevo_clientes:
                st.session_state.archivo_clientes = nuevo_clientes
                st.success(f"✅ {nuevo_clientes.name}")
        with col2:
            st.markdown("**🔑 Corresponsal**")
            nuevo_corresponsal = st.file_uploader(
                "Reemplazar corresponsal",
                type=["xlsx"],
                key="up_corresponsal_manual",
                label_visibility="collapsed"
            )
            if nuevo_corresponsal:
                st.session_state.archivo_corresponsal = nuevo_corresponsal
                st.success(f"✅ {nuevo_corresponsal.name}")


# ══════════════════════════════════════════════════════════════════════════
# TAB 2 — TABLA DE PAGOS
# ══════════════════════════════════════════════════════════════════════════
def render_tabla_pagos():
    from extraccion import extraer_pagos_bancarios, extraer_pagos_recaudos

    if not st.session_state.archivo_libro:
        st.warning("⚠️ Primero carga el Libro de Banco en la pestaña **1. Carga de Archivos**.")
        return

    st.markdown("### Selecciona el tipo de pago a trabajar")

    col1, col2 = st.columns(2)
    with col1:
        sel1 = st.session_state.tipo_pago == "bancarios"
        if st.button(
            "🏦  PAGOS BANCARIOS\nBancolombia · Davivienda · Suprecredito",
            key="sel_bancarios",
            use_container_width=True,
            type="primary" if sel1 else "secondary"
        ):
            st.session_state.tipo_pago = "bancarios"
            st.session_state.df_area_banco = None
            st.rerun()

    with col2:
        sel2 = st.session_state.tipo_pago == "recaudos"
        if st.button(
            "📋  PAGOS POR RECAUDOS\nOccidente · Suprecredito 2026 · Record",
            key="sel_recaudos",
            use_container_width=True,
            type="primary" if sel2 else "secondary"
        ):
            st.session_state.tipo_pago = "recaudos"
            st.session_state.df_area_banco = None
            st.rerun()

    if not st.session_state.tipo_pago:
        st.info("👆 Selecciona el tipo de pago para continuar.")
        return

    st.markdown("---")

    if st.session_state.tipo_pago == "recaudos":
        st.markdown("**📅 Fechas de trabajo** (máximo 5 fechas)")
        st.caption("Útil para fines de semana o días acumulados.")

        if "fechas_recaudo" not in st.session_state:
            st.session_state.fechas_recaudo = [None]

        col_add, col_del = st.columns([1, 1])
        with col_add:
            if len(st.session_state.fechas_recaudo) < 5:
                if st.button("➕ Agregar fecha", key="add_fecha"):
                    st.session_state.fechas_recaudo.append(None)
                    st.rerun()
        with col_del:
            if len(st.session_state.fechas_recaudo) > 1:
                if st.button("➖ Quitar fecha", key="del_fecha"):
                    st.session_state.fechas_recaudo.pop()
                    st.rerun()

        fechas_cols = st.columns(len(st.session_state.fechas_recaudo))
        for i, col in enumerate(fechas_cols):
            with col:
                from datetime import date
                val = st.session_state.fechas_recaudo[i] or date.today()
                fecha_sel = st.date_input(
                    f"Fecha {i+1}",
                    value=val,
                    key=f"fecha_recaudo_{i}"
                )
                st.session_state.fechas_recaudo[i] = fecha_sel

        fechas_validas = [f for f in st.session_state.fechas_recaudo if f is not None]
        fechas_str = ", ".join(f.strftime("%d/%m/%Y") for f in fechas_validas)
        st.caption(f"Fechas seleccionadas: **{fechas_str}**")
        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("**🏢 Filtro por entidad recaudadora**")
        tipo_entidad = st.radio(
            "Tipo de filtro",
            ["Todas", "Elegir por entidad"],
            horizontal=True,
            key="tipo_filtro_entidad",
            label_visibility="collapsed"
        )
        st.session_state["filtro_entidad_tipo"] = tipo_entidad

        ENTIDADES_DISPONIBLES = ["EFECTY", "PSE", "EFECTY-BANCO DE BOGOTA", "RECORD"]

        if tipo_entidad == "Elegir por entidad":
            if "entidades_recaudo" not in st.session_state or st.session_state.get("filtro_entidad_tipo_prev") != "Elegir por entidad":
                st.session_state.entidades_recaudo = []
            st.session_state["filtro_entidad_tipo_prev"] = "Elegir por entidad"

            for i, ent in enumerate(st.session_state.entidades_recaudo):
                ya_sel = [e for j, e in enumerate(st.session_state.entidades_recaudo) if j != i]
                opciones = [e for e in ENTIDADES_DISPONIBLES if e not in ya_sel]
                col_ent, col_del2 = st.columns([4, 1])
                with col_ent:
                    sel = st.selectbox(
                        f"Entidad {i+1}",
                        opciones,
                        index=opciones.index(ent) if ent in opciones else 0,
                        key=f"entidad_sel_{i}"
                    )
                    st.session_state.entidades_recaudo[i] = sel
                with col_del2:
                    if st.button("✕", key=f"del_ent_{i}", help="Quitar esta entidad"):
                        st.session_state.entidades_recaudo.pop(i)
                        st.rerun()

            if len(st.session_state.entidades_recaudo) < 3:
                if st.button("➕ Agregar entidad", key="add_entidad"):
                    ya_sel = st.session_state.entidades_recaudo
                    primera_libre = next((e for e in ENTIDADES_DISPONIBLES if e not in ya_sel), None)
                    if primera_libre:
                        st.session_state.entidades_recaudo.append(primera_libre)
                    st.rerun()

            entidades_validas = [e for e in st.session_state.entidades_recaudo if e]
            if entidades_validas:
                st.caption(f"Entidades seleccionadas: **{', '.join(entidades_validas)}**")
            else:
                st.warning("⚠️ Agrega al menos una entidad.")
        else:
            st.session_state.entidades_recaudo = list(ENTIDADES_DISPONIBLES)
            st.session_state["filtro_entidad_tipo_prev"] = "Todas"
            entidades_validas = list(ENTIDADES_DISPONIBLES)

        st.markdown("<br>", unsafe_allow_html=True)

    tipo_label = "Pagos Bancarios" if st.session_state.tipo_pago == "bancarios" else "Pagos por Recaudos"
    if st.button(f"⬇️  Extraer {tipo_label} del Libro de Banco", type="primary", use_container_width=True):
        with st.spinner("Leyendo el libro de banco y aplicando filtros..."):
            try:
                if st.session_state.tipo_pago == "bancarios":
                    df, resumen = extraer_pagos_bancarios(
                        st.session_state.archivo_libro,
                        st.session_state.archivo_corresponsal
                    )
                else:
                    fechas_sel = [f for f in st.session_state.get("fechas_recaudo", [None]) if f is not None]
                    entidades_sel = st.session_state.get("entidades_recaudo", None)
                    df, resumen = extraer_pagos_recaudos(
                        st.session_state.archivo_libro,
                        fechas_sel,
                        entidades_sel
                    )
                st.session_state.df_area_banco = df
                st.success(f"✅ Extracción completada. {resumen}")
            except Exception as e:
                st.error(f"❌ Error en la extracción: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.df_area_banco is not None:
        df = st.session_state.df_area_banco
        st.markdown("---")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total registros", len(df))
        c2.metric("Corresponsales", df["T_TRANSACCION"].notna().sum() if "T_TRANSACCION" in df.columns else 0)
        c3.metric("Entidades", df["ENTIDAD"].nunique() if "ENTIDAD" in df.columns else 0)
        if "VALOR" in df.columns:
            total = pd.to_numeric(df["VALOR"], errors="coerce").sum()
            c4.metric("Valor total", f"${total:,.0f}")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="tabla-header"><h3>📊 Área de Banco</h3></div>', unsafe_allow_html=True)

        cols_visibles = ["FECHA", "ENTIDAD", "CEDULA", "VALOR", "FRA",
                         "RECIBOS", "FECHA_DOCUMENTO", "REINCIDENTES_CB", "COMPENSACION"]
        cols_mostrar = [c for c in cols_visibles if c in df.columns]

        df_show = df[cols_mostrar].copy()
        for col in df_show.columns:
            if col in ["FECHA", "FECHA_DOCUMENTO"]:
                df_show[col] = pd.to_datetime(df_show[col], errors="coerce")
            elif col == "VALOR":
                df_show[col] = pd.to_numeric(df_show[col], errors="coerce")
            else:
                df_show[col] = df_show[col].astype(str).replace("None", "").replace("nan", "")

        df_editado = st.data_editor(
            df_show,
            use_container_width=True,
            num_rows="fixed",
            key="tabla_area_banco",
            column_config={
                "FECHA":           st.column_config.DateColumn("Fecha",           format="DD/MM/YYYY"),
                "FECHA_DOCUMENTO": st.column_config.DateColumn("Fecha Documento", format="DD/MM/YYYY"),
                "VALOR":           st.column_config.NumberColumn("Valor",         format="$%d"),
                "REINCIDENTES_CB": st.column_config.TextColumn("Reincidentes CB"),
                "COMPENSACION":    st.column_config.TextColumn("Compensación"),
            }
        )
        for col in df.columns:
            if col not in cols_mostrar:
                df_editado[col] = df[col].values
        st.session_state.df_area_banco = df_editado

        st.caption("Puedes editar celdas directamente en la tabla si necesitas hacer ajustes manuales.")


# ══════════════════════════════════════════════════════════════════════════
# FUNCIÓN: REEMPLAZO DE CÉDULAS NO ENCONTRADAS
# ══════════════════════════════════════════════════════════════════════════
def _mostrar_reemplazo_cedulas():
    from collections import defaultdict

    no_encontradas = st.session_state.get("pendiente_reemplazo", [])

    st.markdown("---")
    st.markdown("### ⚠️ Cédulas no encontradas en base de clientes")
    st.markdown("Puedes reemplazar cédulas incorrectas o registrar clientes nuevos.")

    if "clientes_temporales" not in st.session_state:
        st.session_state["clientes_temporales"] = []
    if "cedulas_nuevas_set" not in st.session_state:
        st.session_state["cedulas_nuevas_set"] = set()

    # Cédulas convenio que deben mostrarse individualmente (no agrupar)
    CEDULAS_CONVENIO = {"11237"}

    # Agrupar normalmente excepto las de convenio → una entrada por movimiento
    cedulas_agrupadas = defaultdict(list)
    for item in no_encontradas:
        if item["cedula"] in CEDULAS_CONVENIO:
            # Clave única por índice para que no se agrupen
            cedulas_agrupadas[f"{item['cedula']}__idx__{item['idx']}"].append(item)
        else:
            cedulas_agrupadas[item["cedula"]].append(item)

    reemplazos = {}
    df_banco = st.session_state.df_area_banco

    for i, (cedula_key, items) in enumerate(cedulas_agrupadas.items()):
        # Recuperar cédula real (sin el sufijo __idx__ para convenios)
        cedula = cedula_key.split("__idx__")[0]
        n_pagos     = len(items)
        valores     = []
        for item in items:
            try:
                val = df_banco.at[item["idx"], "VALOR"]
                valores.append(float(str(val).replace(",", "") or 0))
            except Exception:
                valores.append(0)

        pagos_str   = f"{n_pagos} pago" if n_pagos == 1 else f"{n_pagos} pagos"
        valores_str = " | ".join(f"${v:,.0f}" for v in valores)
        es_nuevo    = cedula in st.session_state["cedulas_nuevas_set"]

        # Obtener entidad(es) de los movimientos de esta cédula
        entidades_ced = []
        for item in items:
            try:
                ent = str(df_banco.at[item["idx"], "ENTIDAD"]).strip()
                if ent and ent not in ("nan", "None", "") and ent not in entidades_ced:
                    entidades_ced.append(ent)
            except Exception:
                pass
        entidad_str = " · ".join(entidades_ced) if entidades_ced else ""

        with st.container():
            header = f"**Cédula {cedula}**"
            if entidad_str:
                header += f" | {entidad_str}"
            header += f" | {pagos_str}: {valores_str}"
            st.markdown(header)

            if not es_nuevo:
                col_inp, col_btn = st.columns([3, 1])
                with col_inp:
                    nueva = st.text_input(
                        "Reemplazar por (vacío = NO EXISTE):",
                        value=st.session_state.get(f"remp_val_{cedula}", ""),
                        key=f"remp_{i}",
                        placeholder="Dejar vacío = NO EXISTE"
                    )
                    st.session_state[f"remp_val_{cedula}"] = nueva
                with col_btn:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("➕ Cliente nuevo", key=f"btn_nuevo_{cedula}_{i}"):
                        st.session_state["cedulas_nuevas_set"].add(cedula)
                        st.rerun()
                for item in items:
                    reemplazos[item["idx"]] = {
                        "cedula_original": cedula,
                        "cedula_nueva":    nueva.strip()
                    }
            else:
                st.success("✅ Registrando como cliente nuevo")
                col_c, col_f, col_x = st.columns([2, 2, 1])
                with col_c:
                    company_sel = st.selectbox(
                        "Company",
                        ["Suprecartera", "Suprecreditos", "Movicap", "TuCredito"],
                        key=f"company_{cedula}_{i}"
                    )
                with col_f:
                    factura_sel = st.text_input(
                        "Num Factura",
                        value=st.session_state.get(f"factura_val_{cedula}", ""),
                        key=f"factura_{cedula}_{i}",
                        placeholder="0000000000"
                    )
                    st.session_state[f"factura_val_{cedula}"] = factura_sel
                with col_x:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("✕", key=f"btn_cancel_{cedula}_{i}"):
                        st.session_state["cedulas_nuevas_set"].discard(cedula)
                        st.rerun()

                existente = next((c for c in st.session_state["clientes_temporales"] if c["iden"] == cedula), None)
                if existente:
                    existente["company"]     = company_sel
                    existente["num_factura"] = factura_sel
                else:
                    st.session_state["clientes_temporales"].append({
                        "company":     company_sel,
                        "iden":        cedula,
                        "num_factura": factura_sel
                    })

                for item in items:
                    reemplazos[item["idx"]] = {
                        "cedula_original": cedula,
                        "cedula_nueva":    ""
                    }

            st.markdown("---")

    if st.button("✅ Confirmar y continuar alistamiento",
                 use_container_width=True, type="primary",
                 key="btn_confirmar_reemplazo"):

        df_banco = st.session_state.df_area_banco.copy()
        if "cedulas_ya_revisadas" not in st.session_state:
            st.session_state["cedulas_ya_revisadas"] = set()

        for idx, info in reemplazos.items():
            cedula_orig = info["cedula_original"]
            if cedula_orig in st.session_state["cedulas_nuevas_set"]:
                st.session_state["cedulas_ya_revisadas"].add(cedula_orig)
            elif info["cedula_nueva"]:
                df_banco.at[idx, "CEDULA"] = info["cedula_nueva"]
                st.session_state["cedulas_reemplazadas"][idx] = info
            else:
                st.session_state["cedulas_ya_revisadas"].add(cedula_orig)

        st.session_state.df_area_banco       = df_banco
        st.session_state["pendiente_reemplazo"] = None
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════
# TAB 3 — ALISTAR INFORMACIÓN
# ══════════════════════════════════════════════════════════════════════════
def render_alistar_informacion():
    from alistar import alistar_informacion

    if st.session_state.df_area_banco is None:
        st.warning("⚠️ Primero extrae los datos en la pestaña **2. Tabla de Pagos**.")
        return

    if not st.session_state.archivo_clientes:
        st.warning("⚠️ Primero carga el archivo de Clientes Activos en la pestaña **1. Carga de Archivos**.")
        return

    st.markdown("### Cruce con base de clientes activos")
    st.markdown("""
    <div class="alerta-info">
    🔗 Este proceso cruza las cédulas del Área de Banco con la base de clientes activos,
    generando la tabla final con entidad, fecha, company, cédula, factura, cuota y demás campos.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_btn1, col_btn2 = st.columns([3, 1])
    with col_btn1:
        ejecutar = st.button("🔗  Alistar Información", type="primary", use_container_width=True)
    with col_btn2:
        if st.button("🔄  Limpiar", use_container_width=True):
            st.session_state.df_sheet1 = None
            st.session_state.df_sheet1_base = None
            st.session_state.df_sheet1_alertas = None
            st.session_state.distribuciones_confirmadas = None
            st.session_state["pendiente_reemplazo"] = None
            st.session_state["cedulas_ya_revisadas"] = set()
            st.session_state["cedulas_reemplazadas"] = {}
            st.rerun()

    if st.session_state.get("pendiente_reemplazo"):
        _mostrar_reemplazo_cedulas()
        return

    if ejecutar:
        with st.spinner("Validando cédulas en base de clientes..."):
            try:
                st.session_state["no_encontradas"] = []
                st.session_state["cedulas_reemplazadas"] = {}

                df_clientes = pd.read_excel(
                    st.session_state.archivo_clientes, sheet_name="sheet1"
                )
                df_clientes.columns = [c.strip().upper() for c in df_clientes.columns]
                cedulas_activas = set(df_clientes["IDEN"].astype(str).str.strip().tolist())

                df_banco = st.session_state.df_area_banco.copy()
                no_encontradas = []
                for idx, row in df_banco.iterrows():
                    cedula = str(row.get("CEDULA", "")).strip()
                    if cedula and cedula != "nan" and cedula not in cedulas_activas:
                        no_encontradas.append({"idx": idx, "cedula": cedula})

                ya_revisadas = st.session_state.get("cedulas_ya_revisadas", set())
                no_encontradas_nuevas = [
                    item for item in no_encontradas
                    if item["cedula"] not in ya_revisadas
                ]

                if no_encontradas_nuevas:
                    st.session_state["pendiente_reemplazo"] = no_encontradas_nuevas
                    st.rerun()
                    return

            except Exception as e:
                st.error(f"❌ Error validando cédulas: {str(e)}")
                return

        with st.spinner("Cruzando información y procesando escenarios..."):
            try:
                st.session_state["no_encontradas"] = []

                df_resultado, alertas = alistar_informacion(
                    st.session_state.df_area_banco,
                    st.session_state.archivo_clientes
                )
                st.session_state.df_sheet1_base    = df_resultado
                st.session_state.df_sheet1_alertas = alertas
                st.session_state.distribuciones_confirmadas = None

                no_encontradas = st.session_state.get("no_encontradas", [])
                if no_encontradas:
                    df_banco = st.session_state.df_area_banco.copy()
                    for idx2 in no_encontradas:
                        if idx2 in df_banco.index:
                            df_banco.at[idx2, "RECIBOS"] = "NO EXISTE"
                            if "COMPENSACION" in df_banco.columns:
                                if df_banco.at[idx2, "COMPENSACION"] is None or str(df_banco.at[idx2, "COMPENSACION"]) in ("nan", "None", ""):
                                    df_banco.at[idx2, "COMPENSACION"] = ""
                    st.session_state.df_area_banco = df_banco
                    st.warning(f"⚠️ {len(no_encontradas)} cédula(s) no encontradas en clientes activos. Marcadas como 'NO EXISTE' en la tabla.")

                if not alertas:
                    st.session_state.df_sheet1 = df_resultado
                    st.success(f"✅ Información alistada. {len(df_resultado)} registros procesados.")

            except Exception as e:
                st.error(f"❌ Error al alistar: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.get("df_sheet1_alertas"):
        alertas = st.session_state.df_sheet1_alertas
        from alistar import _resolver_escenarios_multifactura
        try:
            df_extra = _resolver_escenarios_multifactura(alertas)
        except Exception:
            df_extra = None

        if st.session_state.get("distribuciones_confirmadas") is not None:
            df_base  = st.session_state.get("df_sheet1_base", pd.DataFrame())
            df_extra = st.session_state["distribuciones_confirmadas"]
            df_final = pd.concat([df_base, df_extra], ignore_index=True)
            st.session_state.df_sheet1 = df_final
            st.success(f"✅ Información alistada. {len(df_final)} registros procesados.")

    if st.session_state.df_sheet1 is not None:
        df = st.session_state.df_sheet1
        st.markdown("---")

        df_banco       = st.session_state.df_area_banco
        total_extraido = pd.to_numeric(df_banco["VALOR"], errors="coerce").sum() if df_banco is not None and "VALOR" in df_banco.columns else 0
        total_alistado = pd.to_numeric(df["CUOTA"], errors="coerce").sum() if "CUOTA" in df.columns else 0

        total_no_encontrado = 0
        if df_banco is not None and "RECIBOS" in df_banco.columns and "VALOR" in df_banco.columns:
            mask_no = df_banco["RECIBOS"].astype(str).str.strip() == "NO EXISTE"
            total_no_encontrado = pd.to_numeric(df_banco.loc[mask_no, "VALOR"], errors="coerce").sum()

        diferencia = total_extraido - total_alistado - total_no_encontrado

        st.markdown("#### 📊 Cuadre de Control")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total extraído",  f"${total_extraido:,.0f}")
        c2.metric("✅ Alistado",      f"${total_alistado:,.0f}")
        c3.metric("❌ No encontrado", f"${total_no_encontrado:,.0f}")
        if abs(diferencia) < 1:
            c4.metric("Diferencia", "$0", delta="✅ Cuadra", delta_color="normal")
        else:
            c4.metric("Diferencia", f"${diferencia:,.0f}", delta="⚠️ Revisar", delta_color="inverse")

        st.markdown("---")
        c1, c2, c3 = st.columns(3)
        c1.metric("Registros alistados", len(df))
        if "COMPANY" in df.columns:
            c2.metric("Empresas", df["COMPANY"].nunique())
        if "CUOTA" in df.columns:
            c3.metric("Total cuotas", f"${total_alistado:,.0f}")

        st.markdown("<br>", unsafe_allow_html=True)
        st.data_editor(
            df,
            use_container_width=True,
            num_rows="fixed",
            key="tabla_sheet1",
            column_config={
                "FECHA": st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"),
                "FECHA_DOCUMENTO": st.column_config.DateColumn("Fecha Doc.", format="DD/MM/YYYY"),
                "CUOTA": st.column_config.NumberColumn("Cuota", format="$%d"),
            }
        )


# ══════════════════════════════════════════════════════════════════════════
# TAB 4 — GENERAR ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════
def render_generar_archivos():
    if st.session_state.df_sheet1 is None:
        st.warning("⚠️ Primero completa el paso **3. Alistar Información**.")
        return

    st.markdown("### Generación de archivos finales")

    col1, col2 = st.columns(2)

    for key, default in [
        ("cedulas_excluidas_planos", []),
        ("cedulas_solo_cuota", []),
        ("cedulas_inmovilizadas", {}),
        ("df_inmov_cruce", None),
        ("inmov_cruce_stats", {}),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── 1. Excluir cédulas de planos ─────────────────────────────────────
    st.markdown("#### 🚫 Excluir cédulas de planos")
    st.caption("Las cédulas excluidas no aparecerán en los planos pero sí en las compensaciones.")

    col_exc, col_add = st.columns([3, 1])
    with col_exc:
        nueva_exc = st.text_input("Cédula a excluir:", key="input_cedula_excluir",
                                   placeholder="Ingresa la cédula", label_visibility="collapsed")
    with col_add:
        if st.button("➕ Agregar", key="btn_agregar_excluir", use_container_width=True):
            if nueva_exc.strip() and nueva_exc.strip() not in st.session_state["cedulas_excluidas_planos"]:
                st.session_state["cedulas_excluidas_planos"].append(nueva_exc.strip())
                st.rerun()

    if st.session_state["cedulas_excluidas_planos"]:
        for i, ced in enumerate(st.session_state["cedulas_excluidas_planos"]):
            col_c, col_x = st.columns([4, 1])
            with col_c: st.caption(f"🚫 {ced}")
            with col_x:
                if st.button("✕", key=f"del_exc_{i}"):
                    st.session_state["cedulas_excluidas_planos"].pop(i); st.rerun()

    st.markdown("---")

    # ── 2. Pago solo a cuota ─────────────────────────────────────────────
    st.markdown("#### 💳 Pago solo a cuota")
    st.caption("Estas cédulas tendrán aplicaInteresMoratorio, aplicaDescuentoProntoPago y aplicaGestionCobranza = 0.")

    col_sc, col_sc_add = st.columns([3, 1])
    with col_sc:
        nueva_sc = st.text_input("Cédula pago solo cuota:", key="input_solo_cuota",
                                  placeholder="Ingresa la cédula", label_visibility="collapsed")
    with col_sc_add:
        if st.button("➕ Agregar", key="btn_agregar_sc", use_container_width=True):
            if nueva_sc.strip() and nueva_sc.strip() not in st.session_state["cedulas_solo_cuota"]:
                st.session_state["cedulas_solo_cuota"].append(nueva_sc.strip()); st.rerun()

    if st.session_state["cedulas_solo_cuota"]:
        for i, ced in enumerate(st.session_state["cedulas_solo_cuota"]):
            col_c, col_x = st.columns([4, 1])
            with col_c: st.caption(f"💳 {ced}")
            with col_x:
                if st.button("✕", key=f"del_sc_{i}"):
                    st.session_state["cedulas_solo_cuota"].pop(i); st.rerun()

    st.markdown("---")

    # ── 3. Inmovilizadas ─────────────────────────────────────────────────
    st.markdown("#### 🔒 Inmovilizadas")
    st.caption("Aplica las 3 columnas en 0 y agrega servicios a la hoja Services.")

    SERVICIOS_INMOV = [
        ("Intereses moratorios",        "8888"),
        ("Gestión Cobranza",            "919302"),
        ("Inmovilización",              "19051"),
        ("Parqueo",                     "19053"),
        ("Inmovilización A-P",          "19052"),
        ("Trasporte Inmovilización",    "19054"),
    ]

    col_im, col_im_add = st.columns([3, 1])
    with col_im:
        nueva_im = st.text_input("Cédula inmovilizada:", key="input_inmov",
                                  placeholder="Ingresa la cédula", label_visibility="collapsed")
    with col_im_add:
        if st.button("➕ Agregar", key="btn_agregar_im", use_container_width=True):
            ced_im = nueva_im.strip()
            if ced_im and ced_im not in st.session_state["cedulas_inmovilizadas"]:
                cuota_total = 0
                if st.session_state.get("df_sheet1") is not None:
                    df_s = st.session_state.df_sheet1
                    filas = df_s[df_s["IDEN"].astype(str).str.strip() == ced_im]
                    if not filas.empty and "CUOTA" in df_s.columns:
                        cuota_total = float(filas["CUOTA"].iloc[0] or 0)
                st.session_state["cedulas_inmovilizadas"][ced_im] = {
                    "cuota_total": cuota_total,
                    "servicios": {cod: 0 for _, cod in SERVICIOS_INMOV}
                }
                st.rerun()

    for ced_im, data_im in list(st.session_state["cedulas_inmovilizadas"].items()):
        with st.expander(f"🔒 Cédula {ced_im}", expanded=True):
            col_x2 = st.columns([5, 1])
            with col_x2[1]:
                if st.button("✕ Quitar", key=f"del_im_{ced_im}"):
                    del st.session_state["cedulas_inmovilizadas"][ced_im]; st.rerun()

            st.markdown("**Servicios:**")
            cols_serv = st.columns(2)
            suma_servicios = 0
            for idx_s, (nombre_s, cod_s) in enumerate(SERVICIOS_INMOV):
                with cols_serv[idx_s % 2]:
                    val_s = st.number_input(
                        f"{nombre_s} ({cod_s})",
                        min_value=0.0, value=float(data_im["servicios"].get(cod_s, 0)),
                        step=1000.0, format="%.0f",
                        key=f"serv_{ced_im}_{cod_s}"
                    )
                    st.session_state["cedulas_inmovilizadas"][ced_im]["servicios"][cod_s] = val_s
                    suma_servicios += val_s

            cuota_total = data_im["cuota_total"]
            valor_cuota = cuota_total - suma_servicios

            st.markdown("---")
            col_t, col_v = st.columns(2)
            with col_t:
                st.metric("Total cuota", f"${cuota_total:,.0f}")
            with col_v:
                st.metric("Valor cuota (CashReceipt)",
                          f"${valor_cuota:,.0f}",
                          delta=f"-${suma_servicios:,.0f}" if suma_servicios > 0 else None)

    st.markdown("---")

    # ── 4. Validación: cargar archivo de cédulas inmovilizadas ───────────
    st.markdown("#### 📋 Validar cédulas inmovilizadas (archivo)")
    st.caption("Sube el archivo de cédulas inmovilizadas (columna A: Cédula titular, columna B: Valor). "
               "Se cruzará contra la tabla de Alistar Información para validar antes de generar planos.")

    archivo_inmov = st.file_uploader(
        "Archivo de cédulas inmovilizadas",
        type=["xlsx", "xls"],
        key="archivo_cedulas_inmov_upload",
        label_visibility="collapsed"
    )

    if st.button("🔍 Cruzar cédulas", key="btn_cruzar_inmov", use_container_width=True):
        if archivo_inmov is None:
            st.warning("⚠️ Primero sube el archivo de cédulas inmovilizadas.")
        else:
            try:
                df_inmov_arch = pd.read_excel(archivo_inmov, header=0)
                col_ced = df_inmov_arch.columns[0]
                col_val = df_inmov_arch.columns[1] if df_inmov_arch.shape[1] > 1 else None

                df_inmov_arch = df_inmov_arch.rename(columns={col_ced: "CEDULA_TITULAR"})
                if col_val:
                    df_inmov_arch = df_inmov_arch.rename(columns={col_val: "VALOR"})
                else:
                    df_inmov_arch["VALOR"] = 0

                df_inmov_arch["CEDULA_TITULAR"] = df_inmov_arch["CEDULA_TITULAR"].astype(str).str.strip()
                df_inmov_arch = df_inmov_arch[df_inmov_arch["CEDULA_TITULAR"].notna() &
                                               (df_inmov_arch["CEDULA_TITULAR"] != "") &
                                               (df_inmov_arch["CEDULA_TITULAR"] != "nan")]

                df_s1 = st.session_state.df_sheet1.copy()
                df_s1["IDEN_STR"] = df_s1["IDEN"].astype(str).str.strip()

                cedulas_archivo = set(df_inmov_arch["CEDULA_TITULAR"])
                df_encontradas = df_s1[df_s1["IDEN_STR"].isin(cedulas_archivo)].copy()

                # Traer el valor del archivo a cada fila encontrada
                valores_map = dict(zip(df_inmov_arch["CEDULA_TITULAR"], df_inmov_arch["VALOR"]))
                df_encontradas["VALOR_ARCHIVO"] = df_encontradas["IDEN_STR"].map(valores_map)

                cols_mostrar = [c for c in ["IDEN_STR", "COMPANY", "NUM_FACTURA", "CUOTA", "VALOR_ARCHIVO"]
                                if c in df_encontradas.columns]
                df_resultado = df_encontradas[cols_mostrar].rename(columns={"IDEN_STR": "CEDULA"})

                st.session_state["df_inmov_cruce"] = df_resultado
                st.session_state["inmov_cruce_stats"] = {
                    "total_archivo": len(cedulas_archivo),
                    "encontradas":   df_resultado["CEDULA"].nunique() if not df_resultado.empty else 0,
                }
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error al procesar el archivo: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.get("df_inmov_cruce") is not None:
        stats = st.session_state.get("inmov_cruce_stats", {})
        df_resultado = st.session_state["df_inmov_cruce"]

        c1, c2, c3 = st.columns(3)
        c1.metric("Cédulas en archivo", stats.get("total_archivo", 0))
        c2.metric("Encontradas", stats.get("encontradas", 0))
        c3.metric("No encontradas", stats.get("total_archivo", 0) - stats.get("encontradas", 0))

        if df_resultado.empty:
            st.warning("⚠️ Ninguna cédula del archivo coincide con la tabla de Alistar Información.")
        else:
            st.dataframe(df_resultado, use_container_width=True)

    st.markdown("---")

    # ── 5. Pagos por Error de Causaciones ────────────────────────────────
    st.markdown("#### ⚠️ Pagos por Error de Causaciones")
    st.caption(
        "Pega las cédulas que deben excluirse de la compensación normal. Se "
        "generará un archivo de compensación aparte solo con esas cédulas, "
        "con la fecha de documento que definas aquí."
    )

    col_fdoc_causa, col_btn_causa = st.columns([3, 1])
    with col_fdoc_causa:
        fecha_doc_causaciones = st.text_input(
            "Fecha Documento para aplicar a toda la tabla",
            key="causaciones_fecha_doc_masiva",
            placeholder="Ej: 20/07/2026"
        )
    with col_btn_causa:
        st.markdown("<br>", unsafe_allow_html=True)
        aplicar_fecha_causa = st.button(
            "📅  Aplicar a toda la tabla", use_container_width=True, key="btn_aplicar_fecha_causa"
        )

    columnas_causa = ["CEDULA", "FECHA_DOCUMENTO"]

    if "causaciones_editor_version" not in st.session_state:
        st.session_state["causaciones_editor_version"] = 0
    causa_editor_key = f"editor_causaciones_{st.session_state['causaciones_editor_version']}"

    if "causaciones_seed_df" not in st.session_state:
        st.session_state["causaciones_seed_df"] = pd.DataFrame(
            {c: pd.Series(dtype="str") for c in columnas_causa}
        )
    plantilla_causa = st.session_state["causaciones_seed_df"]

    st.markdown("**⬆️ Pega aquí las cédulas copiadas desde Excel:**")
    df_causaciones_editado = st.data_editor(
        plantilla_causa,
        num_rows="dynamic",
        use_container_width=True,
        key=causa_editor_key,
        column_config={
            "CEDULA":          st.column_config.TextColumn("Cédula"),
            "FECHA_DOCUMENTO": st.column_config.TextColumn("Fecha Documento (tal cual la pegues)"),
        }
    )

    if aplicar_fecha_causa:
        if not fecha_doc_causaciones.strip():
            st.warning("⚠️ Escribe primero una fecha en el campo de arriba.")
        else:
            df_con_datos_causa = df_causaciones_editado.copy()
            mask_con_datos_causa = df_con_datos_causa["CEDULA"].astype(str).str.strip().ne("")
            df_con_datos_causa.loc[mask_con_datos_causa, "FECHA_DOCUMENTO"] = fecha_doc_causaciones.strip()
            st.session_state["causaciones_seed_df"] = df_con_datos_causa
            st.session_state["causaciones_editor_version"] += 1
            st.rerun()

    if st.button("🔄  Limpiar tabla de causaciones", key="limpiar_causaciones"):
        st.session_state["causaciones_editor_version"] += 1
        st.session_state["causaciones_seed_df"] = pd.DataFrame(
            {c: pd.Series(dtype="str") for c in columnas_causa}
        )
        st.rerun()

    st.session_state["df_causaciones_actual"] = df_causaciones_editado

    st.markdown("---")

    with col1:
        st.markdown("#### 📄 Planos por empresa")
        st.caption("Genera CashReceipt, Services y PaymentMethod para cada empresa.")
        if st.button("📄  Crear Planos", type="primary", use_container_width=True, key="btn_planos"):
            st.session_state["planos_generados"] = []
            from generar_planos import crear_planos
            with st.spinner("Generando planos..."):
                try:
                    resultado = crear_planos(
                        st.session_state.df_sheet1,
                        st.session_state.config,
                        st.session_state.df_area_banco,
                        st.session_state.tipo_pago,
                        st.session_state.get("cedulas_excluidas_planos", []),
                        st.session_state.get("cedulas_solo_cuota", []),
                        st.session_state.get("cedulas_inmovilizadas", {})
                    )
                    st.success(f"✅ {resultado}")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")

    planos = st.session_state.get("planos_generados", [])
    if planos:
        import zipfile as zf
        import io as _io

        st.markdown("---")
        st.markdown("#### 📥 Descargar planos generados:")

        if "planos_zip" not in st.session_state or st.session_state.get("planos_zip_count") != len(planos):
            zip_buf = _io.BytesIO()
            with zf.ZipFile(zip_buf, "w", zf.ZIP_DEFLATED) as zfile:
                for arch in planos:
                    zfile.writestr(arch["nombre"], bytes(arch["bytes"]))
            st.session_state["planos_zip"] = zip_buf.getvalue()
            st.session_state["planos_zip_count"] = len(planos)
            st.session_state["planos_zip_nombre"] = f"PLANOS_{planos[0]['nombre'].split('_')[2]}.zip"

        st.download_button(
            label=f"⬇️  Descargar todos los planos ({len(planos)} archivos) — ZIP",
            data=st.session_state["planos_zip"],
            file_name=st.session_state["planos_zip_nombre"],
            mime="application/zip",
            key="dl_zip_planos",
            type="primary",
            use_container_width=True
        )

    with col2:
        st.markdown("#### 📊 Compensaciones por fecha")
        st.caption("Genera un archivo de compensación por cada fecha única en Área de Banco.")
        if st.button("📊  Crear Compensaciones", type="primary", use_container_width=True, key="btn_comp"):
            from generar_compensaciones import crear_compensaciones
            with st.spinner("Generando compensaciones..."):
                try:
                    df_banco_comp = st.session_state.df_area_banco.copy()

                    # Separar cédulas de "Pagos por Error de Causaciones"
                    df_causa = st.session_state.get("df_causaciones_actual")
                    cedulas_causa_map = {}
                    if df_causa is not None and not df_causa.empty:
                        for _, r in df_causa.iterrows():
                            ced = str(r.get("CEDULA", "")).strip()
                            fdoc = str(r.get("FECHA_DOCUMENTO", "")).strip()
                            if ced:
                                cedulas_causa_map[ced] = fdoc

                    if cedulas_causa_map:
                        mask_causa = df_banco_comp["CEDULA"].astype(str).str.strip().isin(cedulas_causa_map.keys())

                        # ── El archivo "aparte" de causaciones es igual en ambos casos:
                        # solo las cédulas de esta tabla, una a una (estilo Pendiente por ID)
                        df_causaciones = df_banco_comp[mask_causa].copy()
                        df_causaciones["FECHA_DOCUMENTO"] = df_causaciones["FECHA_DOCUMENTO"].astype(object)
                        for ced, fdoc in cedulas_causa_map.items():
                            if fdoc:
                                mask_ced = df_causaciones["CEDULA"].astype(str).str.strip() == ced
                                df_causaciones.loc[mask_ced, "FECHA_DOCUMENTO"] = fdoc
                        from generar_gastos_bancarios import _parsear_fecha_mixta
                        df_causaciones["FECHA_DOCUMENTO"] = _parsear_fecha_mixta(df_causaciones["FECHA_DOCUMENTO"])

                        if st.session_state.tipo_pago == "bancarios":
                            # Bancarios: se excluyen del archivo normal
                            df_normal = df_banco_comp[~mask_causa].copy()
                        else:
                            # Recaudos: NO se excluyen del archivo normal, se marcan
                            # "NO EXISTE" para que _generar_recaudos() las lleve a la
                            # cuenta 141299011 en vez de acreditarlas a la cédula
                            # (aunque sí sabemos quién es).
                            df_normal = df_banco_comp.copy()
                            if "RECIBOS" not in df_normal.columns:
                                df_normal["RECIBOS"] = None
                            df_normal["RECIBOS"] = df_normal["RECIBOS"].astype(object)
                            df_normal.loc[mask_causa, "RECIBOS"] = "NO EXISTE"
                    else:
                        df_normal = df_banco_comp
                        df_causaciones = None

                    resultado = crear_compensaciones(
                        df_normal,
                        st.session_state.config,
                        st.session_state.tipo_pago
                    )
                    st.success(f"✅ {resultado}")

                    if df_causaciones is not None and not df_causaciones.empty:
                        st.markdown("---")
                        # El archivo de causaciones siempre se genera "una a una"
                        # (estilo bancarios/Pendiente por ID), sin importar si el
                        # trabajo actual es por bancarios o por recaudos.
                        resultado_causa = crear_compensaciones(
                            df_causaciones,
                            st.session_state.config,
                            "bancarios",
                            prefijo="COMPENSACION_CAUSACIONES"
                        )
                        st.success(f"✅ Causaciones: {resultado_causa}")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

    # ── Consecutivos de compensación ─────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🔢 Registrar consecutivos de compensación")
    st.caption("Ingresa el consecutivo que generó el sistema por cada fecha de compensación.")

    df_banco = st.session_state.df_area_banco
    if df_banco is not None and "FECHA_DOCUMENTO" in df_banco.columns:
        fechas_doc = pd.to_datetime(df_banco["FECHA_DOCUMENTO"], errors="coerce").dt.normalize().dropna().unique()

        if len(fechas_doc) > 0:
            if "consecutivos_comp" not in st.session_state:
                st.session_state.consecutivos_comp = {}

            with st.form("form_consecutivos"):
                cols = st.columns(min(len(fechas_doc), 3))
                for i, fecha in enumerate(sorted(fechas_doc)):
                    fecha_str = pd.Timestamp(fecha).strftime("%d/%m/%Y")
                    with cols[i % len(cols)]:
                        val = st.text_input(
                            f"Compensación {fecha_str}",
                            value=st.session_state.consecutivos_comp.get(fecha_str, ""),
                            key=f"consec_{i}"
                        )
                        st.session_state.consecutivos_comp[fecha_str] = val

                aplicar = st.form_submit_button(
                    "✅ Aplicar consecutivos a tabla de pagos",
                    use_container_width=True,
                    type="primary"
                )

            if aplicar:
                df_banco = st.session_state.df_area_banco.copy()
                df_banco["FECHA_DOC_STR"] = pd.to_datetime(
                    df_banco["FECHA_DOCUMENTO"], errors="coerce"
                ).dt.strftime("%d/%m/%Y")

                for fecha_str, consec in st.session_state.consecutivos_comp.items():
                    if consec.strip():
                        mask = df_banco["FECHA_DOC_STR"] == fecha_str
                        df_banco.loc[mask, "COMPENSACION"] = consec.strip()

                df_banco = df_banco.drop(columns=["FECHA_DOC_STR"])

                mask_aplicado = (
                    df_banco["RECIBOS"].isna() |
                    (df_banco["RECIBOS"].astype(str).str.strip() == "") |
                    (df_banco["RECIBOS"].astype(str).str.strip() == "None")
                )
                df_banco.loc[mask_aplicado, "RECIBOS"] = "APLICADO"

                st.session_state.df_area_banco = df_banco
                st.success("✅ Consecutivos aplicados. Registros marcados como APLICADO en columna RECIBOS.")

    # ── Exportar tabla de pagos ───────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 💾 Exportar tabla de pagos")
    st.caption("Descarga la tabla de pagos con los consecutivos de compensación aplicados.")

    if st.button("💾  Generar Excel tabla de pagos", use_container_width=True, key="btn_export_tabla"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import io as _io

            df_export = st.session_state.df_area_banco.copy()

            cols_visibles = ["FECHA", "ENTIDAD", "CEDULA", "VALOR", "FRA",
                             "RECIBOS", "FECHA_DOCUMENTO", "REINCIDENTES_CB", "COMPENSACION"]
            cols_export = [c for c in cols_visibles if c in df_export.columns]
            df_export = df_export[cols_export]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AREA DE BANCO"

            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for col_idx, col_name in enumerate(df_export.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center")

            cols_fecha = ["FECHA", "FECHA_DOCUMENTO"]
            idx_fechas = [list(df_export.columns).index(c)+1 for c in cols_fecha if c in df_export.columns]

            for row_idx, row in df_export.iterrows():
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx+2, column=col_idx, value=value)
                    if col_idx in idx_fechas and value:
                        try:
                            cell.value = pd.Timestamp(value).to_pydatetime()
                            cell.number_format = "DD/MM/YYYY"
                        except Exception:
                            pass

            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

            buffer = _io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            from datetime import datetime as _dt
            nombre = f"TABLA_PAGOS_{_dt.now().strftime('%d_%m_%Y_%H_%M_%S')}.xlsx"

            # ── Guardar automáticamente en ruta configurada ──────────────
            ruta_auto = st.session_state.config.get("ruta_tabla_pagos", "")
            if ruta_auto:
                try:
                    os.makedirs(ruta_auto, exist_ok=True)
                    ruta_completa = os.path.join(ruta_auto, nombre)
                    buffer.seek(0)
                    with open(ruta_completa, "wb") as f:
                        f.write(buffer.read())
                    st.success(f"💾 Guardado en: {ruta_completa}")
                except Exception as e:
                    st.warning(f"⚠️ No se pudo guardar automáticamente: {str(e)}")
                buffer.seek(0)

            st.download_button(
                label="⬇️  Descargar tabla de pagos",
                data=buffer,
                file_name=nombre,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tabla_pagos"
            )
        except Exception as e:
            st.error(f"❌ Error al exportar: {str(e)}")

# ══════════════════════════════════════════════════════════════════════════
# PENDIENTE POR ID
# ══════════════════════════════════════════════════════════════════════════
def _formato_cop_pid(valor):
    """Formatea un número como moneda colombiana: $5.000.025,25"""
    try:
        s = f"{float(valor):,.2f}"
    except Exception:
        return "$0,00"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"${s}"


def render_pendiente_por_id():
    from alistar import alistar_informacion, _resolver_escenarios_multifactura
    from generar_planos import crear_planos
    from generar_compensaciones import crear_compensaciones
    from generar_gastos_bancarios import _num, _parsear_fecha_mixta

    st.markdown("### 🆔 Aplicación y Compensación Pendiente por ID")
    st.caption(
        "Pega los pagos que antes quedaron marcados como NO EXISTE y que ya se "
        "identificaron. Columnas: VALOR, FECHA, ENTIDAD, CEDULA, FRA. Incluye "
        "Bancolombia, Davivienda, PSE, Efecty y Récord — la compensación se "
        "genera uno a uno por registro (misma cuenta débito que Bancolombia/"
        "Davivienda), sin sumar por entidad."
    )

    if not st.session_state.get("archivo_clientes"):
        archivo = _cargar_desde_github("CLIENTES_ACTIVOS.xlsx")
        if archivo:
            st.session_state.archivo_clientes = archivo

    if not st.session_state.get("archivo_clientes"):
        st.warning(
            "⚠️ No se pudo cargar el archivo de Clientes Activos desde el "
            "repositorio. Verifica tu conexión o inténtalo de nuevo."
        )
        return

    col_fdoc, col_btn_fdoc = st.columns([3, 1])
    with col_fdoc:
        fecha_doc_masiva = st.text_input(
            "Fecha Documento para aplicar a toda la tabla (déjalo vacío para usar la misma Fecha de cada fila)",
            key="pid_fecha_doc_masiva",
            placeholder="Ej: 20/07/2026"
        )
    with col_btn_fdoc:
        st.markdown("<br>", unsafe_allow_html=True)
        aplicar_fecha_doc = st.button(
            "📅  Aplicar a toda la tabla", use_container_width=True, key="btn_aplicar_fecha_doc"
        )

    columnas_base = ["FECHA", "ENTIDAD", "CEDULA", "VALOR", "FRA", "FECHA_DOCUMENTO"]

    metric_placeholder = st.empty()

    if "pid_editor_version" not in st.session_state:
        st.session_state["pid_editor_version"] = 0
    editor_key = f"editor_pid_{st.session_state['pid_editor_version']}"

    if "pid_seed_df" not in st.session_state:
        st.session_state["pid_seed_df"] = pd.DataFrame(
            {c: pd.Series(dtype="str") for c in columnas_base}
        )
    plantilla_inicial = st.session_state["pid_seed_df"]

    st.markdown("**⬆️ Pega aquí las filas copiadas desde Excel:**")
    df_editado = st.data_editor(
        plantilla_inicial,
        num_rows="dynamic",
        use_container_width=True,
        key=editor_key,
        column_config={
            "FECHA":           st.column_config.TextColumn("Fecha (tal cual la pegues)"),
            "ENTIDAD":         st.column_config.TextColumn("Entidad"),
            "CEDULA":          st.column_config.TextColumn("Cédula"),
            "VALOR":           st.column_config.TextColumn("Valor (tal cual lo pegues)"),
            "FRA":             st.column_config.TextColumn("Factura (FRA)"),
            "FECHA_DOCUMENTO": st.column_config.TextColumn("Fecha Documento (vacío = igual a Fecha)"),
        }
    )

    if aplicar_fecha_doc:
        if not fecha_doc_masiva.strip():
            st.warning("⚠️ Escribe primero una fecha en el campo de arriba.")
        else:
            df_con_datos = df_editado.copy()
            mask_con_datos = (
                df_con_datos["FECHA"].astype(str).str.strip().ne("") |
                df_con_datos["VALOR"].astype(str).str.strip().ne("") |
                df_con_datos["CEDULA"].astype(str).str.strip().ne("")
            )
            df_con_datos.loc[mask_con_datos, "FECHA_DOCUMENTO"] = fecha_doc_masiva.strip()
            st.session_state["pid_seed_df"] = df_con_datos
            st.session_state["pid_editor_version"] += 1
            st.rerun()

    total_valor = sum(_num(v) for v in df_editado["VALOR"] if str(v).strip() != "")
    metric_placeholder.markdown(f"""
    <div style="background-color:#0f2a1a;border:1px solid #2ecc71;border-radius:10px;
                padding:0.4rem 1.2rem;margin-bottom:1rem;width:fit-content;">
        <div style="color:#2ecc71;font-size:0.85rem;font-weight:600;letter-spacing:0.5px;">
            🆔 VALOR PENDIENTE POR ID
        </div>
        <div style="color:#2ecc71;font-size:1.4rem;font-weight:700;">
            {_formato_cop_pid(total_valor)}
        </div>
    </div>
    """, unsafe_allow_html=True)

    col_alistar, col_lim = st.columns([3, 1])
    with col_lim:
        if st.button("🔄  Limpiar tabla", use_container_width=True, key="limpiar_pid"):
            st.session_state["pid_editor_version"] += 1
            st.session_state["pid_seed_df"] = pd.DataFrame(
                {c: pd.Series(dtype="str") for c in columnas_base}
            )
            for k in ["pid_df_area_banco", "pid_df_sheet1", "pid_df_sheet1_base",
                      "pid_df_sheet1_alertas"]:
                st.session_state.pop(k, None)
            st.session_state["distribuciones_confirmadas"] = None
            st.rerun()
    with col_alistar:
        ejecutar = st.button(
            "🔗  Alistar Información", type="primary",
            use_container_width=True, key="btn_alistar_pid"
        )

    if ejecutar:
        with st.spinner("Cruzando información con Clientes Activos..."):
            try:
                df = df_editado[columnas_base].copy()
                df = df[df["CEDULA"].astype(str).str.strip() != ""].reset_index(drop=True)

                if df.empty:
                    st.warning("⚠️ No hay filas con cédula para procesar.")
                else:
                    df["FECHA"] = _parsear_fecha_mixta(df["FECHA"])
                    df["VALOR"] = df["VALOR"].apply(_num)

                    fecha_doc_pegada = _parsear_fecha_mixta(df["FECHA_DOCUMENTO"])
                    df["FECHA_DOCUMENTO"] = fecha_doc_pegada.where(fecha_doc_pegada.notna(), df["FECHA"])
                    for col in ["RECIBO", "T_TRANSACCION", "INMOVILIZACION",
                                "OTROS_GASTOS", "OBSERVACION"]:
                        df[col] = None

                    st.session_state["pid_df_area_banco"] = df
                    st.session_state["no_encontradas"] = []

                    df_resultado, alertas = alistar_informacion(
                        df, st.session_state.archivo_clientes
                    )
                    st.session_state["pid_df_sheet1_base"]    = df_resultado
                    st.session_state["pid_df_sheet1_alertas"] = alertas
                    st.session_state["distribuciones_confirmadas"] = None

                    if not alertas:
                        st.session_state["pid_df_sheet1"] = df_resultado
                        st.success(f"✅ Información alistada. {len(df_resultado)} registro(s) procesados.")

                    no_enc = st.session_state.get("no_encontradas", [])
                    if no_enc:
                        st.warning(
                            f"⚠️ {len(no_enc)} cédula(s) todavía no se encuentran en "
                            "Clientes Activos. Verifica que la cédula esté bien escrita "
                            "o que el cliente ya esté registrado allí, y vuelve a intentar."
                        )
            except Exception as e:
                st.error(f"❌ Error al alistar: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.get("pid_df_sheet1_alertas"):
        alertas = st.session_state["pid_df_sheet1_alertas"]
        try:
            _resolver_escenarios_multifactura(alertas)
        except Exception:
            pass

        if st.session_state.get("distribuciones_confirmadas") is not None:
            df_base  = st.session_state.get("pid_df_sheet1_base", pd.DataFrame())
            df_extra = st.session_state["distribuciones_confirmadas"]
            df_final = pd.concat([df_base, df_extra], ignore_index=True)
            st.session_state["pid_df_sheet1"] = df_final
            st.success(f"✅ Información alistada. {len(df_final)} registro(s) procesados.")

    if st.session_state.get("pid_df_sheet1") is not None:
        st.markdown("---")
        st.markdown("#### Vista previa de información alistada")
        st.dataframe(st.session_state["pid_df_sheet1"], use_container_width=True)

        st.markdown("---")
        if st.button(
            "📤  Generar Planos y Compensación", type="primary",
            use_container_width=True, key="btn_generar_pid"
        ):
            with st.spinner("Generando archivos..."):
                try:
                    resultado_planos = crear_planos(
                        st.session_state["pid_df_sheet1"],
                        st.session_state.config,
                        st.session_state["pid_df_area_banco"],
                        "bancarios"
                    )
                    st.success(f"✅ Planos: {resultado_planos}")
                except Exception as e:
                    st.error(f"❌ Error generando planos: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

                try:
                    resultado_comp = crear_compensaciones(
                        st.session_state["pid_df_area_banco"],
                        st.session_state.config,
                        "bancarios"
                    )
                    st.success(f"✅ Compensación: {resultado_comp}")
                except Exception as e:
                    st.error(f"❌ Error generando compensación: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())