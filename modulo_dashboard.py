import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

def render():
    st.markdown("""
    <div class="module-header">
        <div class="module-icon">📊</div>
        <div>
            <h1>Dashboard</h1>
            <p>Informes y análisis gerencial</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Menú de informes
    if "submodulo_dashboard" not in st.session_state:
        st.session_state.submodulo_dashboard = None

    sub = st.session_state.submodulo_dashboard

    if sub is None:
        _render_menu()
    elif sub == "corresponsal":
        if st.button("← Volver", key="btn_volver_dash"):
            st.session_state.submodulo_dashboard = None
            st.rerun()
        st.markdown("<br>", unsafe_allow_html=True)
        render_corresponsal()
    elif sub == "cartera":
        if st.button("← Volver", key="btn_volver_dash2"):
            st.session_state.submodulo_dashboard = None
            st.rerun()
        st.markdown("<br>", unsafe_allow_html=True)
        st.info("🚧 Informe Planilla Cartera — próximamente.")


def _render_menu():
    st.markdown("### Selecciona el informe")
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">🏦</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem">
                Informe Corresponsal</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Bancolombia — actualización mensual</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_corresponsal", use_container_width=True, type="primary"):
            st.session_state.submodulo_dashboard = "corresponsal"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:12px;
                    padding:1.5rem;text-align:center;">
            <div style="font-size:2.5rem">📋</div>
            <div style="font-weight:700;color:#fff;margin-top:0.5rem">
                Informe Planilla Cartera</div>
            <div style="color:#64748b;font-size:0.8rem;margin-top:0.3rem">
                Próximamente</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar →", key="btn_cartera", use_container_width=True, type="primary"):
            st.session_state.submodulo_dashboard = "cartera"
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════
# INFORME CORRESPONSAL
# ══════════════════════════════════════════════════════════════════════════
def render_corresponsal():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);border-radius:10px;
                padding:1rem 1.5rem;margin-bottom:1rem;">
        <h3 style="color:#fff;margin:0">🏦 Informe Corresponsal Bancolombia</h3>
        <p style="color:#93c5fd;margin:0.2rem 0 0 0;font-size:0.85rem">
            Actualización mensual del histórico de transferencias</p>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📂 1. Cargar y Procesar", "📊 2. Dashboard"])

    with tab1:
        _render_procesar_corresponsal()
    with tab2:
        _render_dashboard_corresponsal()


def _render_procesar_corresponsal():
    st.markdown("#### 📂 Archivos necesarios")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**📖 Libro de Banco**")
        libro = st.file_uploader("Libro de Banco", type=["xlsx","xlsm"],
                                  key="up_libro_corresponsal", label_visibility="collapsed")
        if libro:
            st.session_state["corr_libro"] = libro
            st.success(f"✅ {libro.name}")
        elif st.session_state.get("corr_libro"):
            st.success("✅ Ya cargado")

    with col2:
        st.markdown("**📊 Informe Corresponsal**")
        informe = st.file_uploader("Informe Corresponsal", type=["xlsx"],
                                    key="up_informe_corresponsal", label_visibility="collapsed")
        if informe:
            st.session_state["corr_informe"] = informe
            st.success(f"✅ {informe.name}")
        elif st.session_state.get("corr_informe"):
            st.success("✅ Ya cargado")

    st.markdown("---")
    st.markdown("#### 📅 Mes a trabajar")
    col_m, col_a = st.columns(2)
    with col_m:
        mes_sel = st.selectbox("Mes", MESES, key="corr_mes")
    with col_a:
        anio_sel = st.number_input("Año", min_value=2020, max_value=2035,
                                    value=datetime.now().year, key="corr_anio")

    st.markdown("<br>", unsafe_allow_html=True)

    # Botón procesar
    col_btn, col_lim = st.columns([3,1])
    with col_btn:
        procesar = st.button("⚙️ Procesar datos", type="primary",
                              use_container_width=True, key="btn_procesar_corr")
    with col_lim:
        if st.button("🔄 Limpiar", use_container_width=True, key="btn_limpiar_corr"):
            for k in ["corr_resultado","corr_stats","corr_libro","corr_informe"]:
                st.session_state.pop(k, None)
            st.rerun()

    if procesar:
        if not st.session_state.get("corr_libro") or not st.session_state.get("corr_informe"):
            st.error("❌ Debes cargar ambos archivos.")
        else:
            with st.spinner("Procesando..."):
                try:
                    stats = _procesar_corresponsal(
                        st.session_state["corr_libro"],
                        st.session_state["corr_informe"],
                        mes_sel, int(anio_sel)
                    )
                    st.session_state["corr_stats"]  = stats
                    st.session_state["corr_mes_sel"]  = mes_sel
                    st.session_state["corr_anio_sel"] = int(anio_sel)
                    st.success("✅ Datos procesados correctamente.")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

    # Mostrar resumen y campo comisión
    stats = st.session_state.get("corr_stats")
    if stats:
        st.markdown("---")
        st.markdown("#### 📊 Resumen del mes")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total transacciones", f"{stats['transacciones']:,}")
        c2.metric("Sin identificar",     f"{stats['sin_identificar']:,}")
        c3.metric("Identificadas",        f"{stats['identificadas']:,}")
        c4.metric("Reincidentes",         f"{stats['reincidentes']:,}")
        c5.metric("Nuevos",               f"{stats['nuevos']:,}")

        st.markdown("---")
        st.markdown("#### 💰 Valor de la comisión")
        comision = st.number_input(
            "Ingresa el valor de la comisión del mes ($):",
            min_value=0, value=0, step=10000,
            key="corr_comision", format="%d"
        )

        if st.button("✅ Generar Informe Excel", type="primary",
                     use_container_width=True, key="btn_generar_corr"):
            if comision <= 0:
                st.warning("⚠️ Ingresa un valor de comisión mayor a 0.")
            else:
                with st.spinner("Generando Excel..."):
                    try:
                        buf = _generar_excel_corresponsal(
                            st.session_state["corr_informe"],
                            stats, comision,
                            st.session_state["corr_mes_sel"],
                            st.session_state["corr_anio_sel"]
                        )
                        st.session_state["corr_excel"] = buf
                        st.session_state["corr_comision_val"] = comision
                        st.success("✅ Excel generado correctamente.")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())

        if st.session_state.get("corr_excel"):
            mes = st.session_state.get("corr_mes_sel","")
            anio = st.session_state.get("corr_anio_sel","")
            st.download_button(
                label="⬇️  Descargar Informe Corresponsal actualizado",
                data=st.session_state["corr_excel"],
                file_name=f"INFORME_CORRESPONSAL_BANCOLOMBIA_{mes.upper()}_{anio}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_corr_excel"
            )


def _procesar_corresponsal(libro, informe, mes, anio):
    """Extrae movimientos del libro de banco y cruza con histórico."""
    # ── Leer libro de banco ──────────────────────────────────────────────
    df_banco = pd.read_excel(libro, sheet_name="BANCOLOMBIA SUPRECREDITO", header=0)
    df_banco.columns = [f"COL_{i}" for i in range(df_banco.shape[1])]

    # Col A = fecha, Col D = cédula, Col F = tipo transacción
    df_banco["COL_0"] = pd.to_datetime(df_banco["COL_0"], errors="coerce")

    # Filtrar por mes y año
    df_mes = df_banco[
        (df_banco["COL_0"].dt.month == MESES.index(mes) + 1) &
        (df_banco["COL_0"].dt.year  == anio)
    ].copy()

    # Filtrar por CONSIGNACION CORRESPONSAL CB en col F
    df_corr = df_mes[
        df_mes["COL_5"].astype(str).str.strip().str.upper() == "CONSIGNACION CORRESPONSAL CB"
    ].copy()

    total_transacciones = len(df_corr)

    # Sin identificar = col D vacía
    sin_iden = df_corr[
        df_corr["COL_3"].isna() | (df_corr["COL_3"].astype(str).str.strip().isin(["","nan"]))
    ]
    con_iden = df_corr[
        ~df_corr["COL_3"].isna() & (~df_corr["COL_3"].astype(str).str.strip().isin(["","nan"]))
    ]

    sin_identificar = len(sin_iden)
    identificadas   = len(con_iden)

    # Consolidar cédulas del mes (cuántos movimientos tuvo cada una)
    cedulas_mes = con_iden["COL_3"].astype(str).str.strip().value_counts().to_dict()

    # ── Leer histórico ───────────────────────────────────────────────────
    df_hist = pd.read_excel(informe, sheet_name="TRANSFERENCIAS CORRESPONSAL", header=0)
    df_hist.columns = ["CEDULA","HIST_ANT","TRANS_MES","HIST_ACT","OBS",
                       "VACIO"] + [f"X{i}" for i in range(df_hist.shape[1]-6)]
    df_hist["CEDULA"] = df_hist["CEDULA"].astype(str).str.strip()

    cedulas_historico = set(df_hist["CEDULA"].tolist())

    reincidentes = 0
    nuevos       = 0
    actualizaciones = {}  # cedula → {hist_ant, trans_mes, hist_act, obs}

    for ced, movs in cedulas_mes.items():
        if ced in cedulas_historico:
            # Contar movimientos (no cédulas) para cuadrar con identificadas
            reincidentes += movs
            fila = df_hist[df_hist["CEDULA"] == ced].iloc[0]
            hist_ant = int(fila["HIST_ACT"]) if pd.notna(fila["HIST_ACT"]) else 0
            actualizaciones[ced] = {
                "hist_ant":  hist_ant,
                "trans_mes": movs,
                "hist_act":  hist_ant + movs,
                "obs":       "REICIDENTE"
            }
        else:
            nuevos += movs
            actualizaciones[ced] = {
                "hist_ant":  0,
                "trans_mes": movs,
                "hist_act":  movs,
                "obs":       "NUEVO"
            }

    return {
        "transacciones":   total_transacciones,
        "sin_identificar": sin_identificar,
        "identificadas":   identificadas,
        "reincidentes":    reincidentes,
        "nuevos":          nuevos,
        "actualizaciones": actualizaciones,
        "cedulas_mes":     cedulas_mes,
        "mes":             mes,
        "anio":            anio,
    }


def _generar_excel_corresponsal(informe_file, stats, comision, mes, anio):
    """Actualiza el Excel del informe corresponsal."""
    # Leer con estilos preservando formatos pero sin evaluar fórmulas
    informe_file.seek(0)
    wb = openpyxl.load_workbook(informe_file, data_only=False, keep_vba=False)

    # ── HOJA TRANSFERENCIAS CORRESPONSAL ────────────────────────────────
    ws_trans = wb["TRANSFERENCIAS CORRESPONSAL"]

    # Actualizar encabezados fila 1 col C y D con nuevo mes
    mes_anterior = ws_trans["C1"].value or ""
    ws_trans["B1"] = mes_anterior  # El anterior pasa a B
    ws_trans["C1"] = f"TRANSACCIONES {mes.upper()} {anio}"
    ws_trans["D1"] = f"HISTORICO A {_ultimo_dia_mes(mes, anio)}"

    # Actualizar encabezado tabla resumen col G
    ws_trans["G1"] = f"TRANSACCIONES CORRESPONSAL MES DE {mes.upper()} {anio}"

    actualizaciones = stats["actualizaciones"]
    cedulas_en_hist = set()

    # Actualizar cédulas existentes
    for row in ws_trans.iter_rows(min_row=2):
        ced = str(row[0].value).strip() if row[0].value else ""
        if not ced or ced == "None":
            continue
        cedulas_en_hist.add(ced)

        if ced in actualizaciones:
            act = actualizaciones[ced]
            # B = historico anterior (lo que era C antes)
            row[1].value = act["hist_ant"]
            # C = transacciones del mes
            row[2].value = act["trans_mes"]
            # D = historico actual
            row[3].value = act["hist_act"]
            # E = observación
            row[4].value = act["obs"]
        else:
            # No tuvo movimientos este mes
            hist_act = row[3].value or 0
            row[1].value = hist_act  # B = lo que era D
            row[2].value = 0         # C = 0
            # D permanece igual
            row[4].value = None      # Limpiar observación

    # Agregar cédulas nuevas copiando formato exacto de fila 2
    from copy import copy
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
    # Leer estilos de referencia de fila 2
    ref_styles = {}
    for col_idx in range(1, 6):
        ref_cell = ws_trans.cell(row=2, column=col_idx)
        ref_styles[col_idx] = {
            "font":           copy(ref_cell.font),
            "border":         copy(ref_cell.border),
            "fill":           copy(ref_cell.fill),
            "number_format":  ref_cell.number_format,
            "alignment":      copy(ref_cell.alignment),
        }

    ultima_fila = ws_trans.max_row
    for ced, act in actualizaciones.items():
        if ced not in cedulas_en_hist:
            ultima_fila += 1
            valores = [
                int(ced) if ced.isdigit() else ced,
                act["hist_ant"],
                act["trans_mes"],
                act["hist_act"],
                act["obs"]
            ]
            for col_idx, val in enumerate(valores, start=1):
                new_cell = ws_trans.cell(row=ultima_fila, column=col_idx, value=val)
                s = ref_styles[col_idx]
                new_cell.font          = copy(s["font"])
                new_cell.border        = copy(s["border"])
                new_cell.fill          = copy(s["fill"])
                new_cell.number_format = s["number_format"]
                new_cell.alignment     = copy(s["alignment"])

    # Actualizar tabla resumen fila 3 cols G-L
    ws_trans["G3"] = stats["transacciones"]
    ws_trans["H3"] = stats["sin_identificar"]
    ws_trans["I3"] = stats["identificadas"]
    ws_trans["J3"] = stats["reincidentes"]
    ws_trans["K3"] = stats["nuevos"]
    ws_trans["L3"] = comision

    # ── HOJA COMISIÓN CORRESPONSAL ───────────────────────────────────────
    ws_com = wb["COMISIÓN CORRESPONSAL"]
    from copy import copy

    # Encontrar fila TOTAL y última fila con datos
    fila_total = None
    fila_ant   = None
    for row in ws_com.iter_rows(min_row=2):
        val = str(row[0].value).strip() if row[0].value else ""
        if "TOTAL" in val.upper():
            fila_total = row[0].row
            break
        if row[0].value and val not in ["","None"]:
            fila_ant = row

    if fila_total is None:
        fila_total = (fila_ant[0].row + 1) if fila_ant else ws_com.max_row + 1

    # Leer comisión anterior para variaciones
    com_ant = 0
    if fila_ant and fila_ant[3].value:
        try:
            # Leer valor numérico del workbook separado con data_only
            informe_file.seek(0)
            wb2 = openpyxl.load_workbook(informe_file, data_only=True)
            ws_com2 = wb2["COMISIÓN CORRESPONSAL"]
            com_ant = float(ws_com2.cell(row=fila_ant[0].row, column=4).value or 0)
            # También leer comisión año anterior
            com_anio_ant = 0
            for row2 in ws_com2.iter_rows(min_row=2):
                if (str(row2[0].value) == str(anio-1) and
                    row2[1].value and str(row2[1].value).strip().lower().startswith(mes.lower()[:3])):
                    com_anio_ant = float(row2[3].value or 0)
                    break
            wb2.close()
        except Exception:
            com_ant = 0
            com_anio_ant = 0

    var_mes_ant_val = comision - com_ant
    var_mes_ant_pct = round(var_mes_ant_val / com_ant, 4) if com_ant else 0
    var_anio_val    = comision - com_anio_ant
    var_anio_pct    = round(var_anio_val / com_anio_ant, 4) if com_anio_ant else 0

    # Insertar nueva fila antes del TOTAL
    nueva_fila    = fila_total
    fila_ref_com  = nueva_fila - 1
    ws_com.insert_rows(nueva_fila)

    # Fórmulas: G y H siempre 12 filas atrás (mismo mes año anterior)
    fila_anio_ant = nueva_fila - 12
    valores_com = {
        1: anio,
        2: mes,
        3: stats["transacciones"],
        4: comision,
        5: f"=+D{nueva_fila}-D{nueva_fila-1}",
        6: f"=+(D{nueva_fila}-D{nueva_fila-1})/D{nueva_fila-1}",
        7: f"=+D{nueva_fila}-D{fila_anio_ant}",
        8: f"=+(D{nueva_fila}-D{fila_anio_ant})/D{fila_anio_ant}",
    }
    for col_idx, val in valores_com.items():
        new_cell = ws_com.cell(row=nueva_fila, column=col_idx, value=val)
        ref_cell = ws_com.cell(row=fila_ref_com, column=col_idx)
        if ref_cell.has_style:
            new_cell.font          = copy(ref_cell.font)
            new_cell.border        = copy(ref_cell.border)
            new_cell.fill          = copy(ref_cell.fill)
            new_cell.number_format = ref_cell.number_format
            new_cell.alignment     = copy(ref_cell.alignment)

    # Actualizar fórmulas del TOTAL para incluir nueva fila
    fila_total_nueva = fila_total + 1  # se movió por insert_rows
    for col_letra in ["C","D","E","G"]:
        cell_total = ws_com[f"{col_letra}{fila_total_nueva}"]
        if cell_total.value and "SUM" in str(cell_total.value):
            # Actualizar rango del SUM para incluir nueva fila
            cell_total.value = f"=SUM({col_letra}2:{col_letra}{nueva_fila})" 

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _ultimo_dia_mes(mes, anio):
    """Retorna el último día del mes como string DD-MM-YYYY."""
    import calendar
    mes_idx = MESES.index(mes) + 1
    ultimo  = calendar.monthrange(anio, mes_idx)[1]
    return f"{ultimo:02d}-{mes_idx:02d}-{anio}"


# ══════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════
# DASHBOARD VISUAL
# ══════════════════════════════════════════════════════════════════════════
def _render_dashboard_corresponsal():
    stats = st.session_state.get("corr_stats")
    if not stats:
        st.warning("⚠️ Primero procesa los datos en la pestaña **1. Cargar y Procesar**.")
        return

    mes  = stats["mes"]
    anio = stats["anio"]
    com  = st.session_state.get("corr_comision_val", 0)

    if not com:
        st.warning("⚠️ Ingresa la comisión en la pestaña 1 antes de ver el dashboard.")
        return

    st.markdown(f"### Corresponsal Bancolombia — {mes} {anio}")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total transacciones", f"{stats['transacciones']:,}")
    c2.metric("Sin identificar",     f"{stats['sin_identificar']:,}")
    c3.metric("Identificadas",       f"{stats['identificadas']:,}")
    c4.metric("Reincidentes",        f"{stats['reincidentes']:,}")
    c5.metric("Nuevos",              f"{stats['nuevos']:,}")
    c6.metric("Comisión",            f"${com:,.0f}")

    st.markdown("---")
    st.markdown("#### 📋 Cédulas del mes")
    actualizaciones = stats.get("actualizaciones", {})
    if actualizaciones:
        df_tabla = pd.DataFrame([
            {
                "Cédula":         ced,
                "Hist. anterior": act["hist_ant"],
                "Trans. mes":     act["trans_mes"],
                "Hist. actual":   act["hist_act"],
                "Observación":    act["obs"]
            }
            for ced, act in actualizaciones.items()
        ])
        st.dataframe(df_tabla, use_container_width=True, height=400)

    st.markdown("---")
    st.markdown("#### 🌐 Dashboard HTML")
    st.caption("Genera el dashboard visual para compartir como archivo HTML.")

    if st.button("🌐 Generar Dashboard HTML", type="primary",
                 use_container_width=True, key="btn_html_corr"):
        html = _generar_html_corresponsal(stats, com)
        st.session_state["corr_html"] = html

    if st.session_state.get("corr_html"):
        st.download_button(
            label="⬇️  Descargar Dashboard HTML",
            data=st.session_state["corr_html"],
            file_name=f"Dashboard_Corresponsal_{mes}_{anio}.html",
            mime="text/html",
            key="dl_corr_html"
        )
        st.success("✅ Dashboard listo. Descárgalo y ábrelo en tu navegador.")
        st.info("📤 **Cómo publicar:** Ve a [app.netlify.com](https://app.netlify.com) → *Add new site* → *Deploy manually* → arrastra el archivo HTML → Netlify genera un link público para compartir ✅")



def _generar_html_corresponsal(stats, comision):
    import json
    import io as _io
    mes    = stats["mes"]
    anio   = stats["anio"]
    trans  = stats["transacciones"]
    sin_id = stats["sin_identificar"]
    ident  = stats["identificadas"]
    reinc  = stats["reincidentes"]
    nuevos = stats["nuevos"]
    # % sobre identificados — garantizar que sumen 100%
    pct_reinc = round(reinc/ident*100, 1) if ident else 0
    pct_nuevo = round(100 - pct_reinc, 1) if ident else 0
    pct_ident = round(ident/trans*100, 1) if trans else 0
    com_fmt   = "${:,.0f}".format(comision).replace(",",".")
    mes_idx   = MESES.index(mes)  # 0-based

    # Leer historico del Excel generado
    informe_bytes = st.session_state.get("corr_excel")
    hist_data     = []
    cedulas_all   = []

    if informe_bytes:
        wb_h = openpyxl.load_workbook(_io.BytesIO(informe_bytes), data_only=True)
        hoja_com = "COMISIÓN CORRESPONSAL" if "COMISIÓN CORRESPONSAL" in wb_h.sheetnames else "COMISION CORRESPONSAL"
        ws_c = wb_h[hoja_com]
        # Leer todos los valores numéricos
        filas_com = []
        for row in ws_c.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            filas_com.append(row)

        # Calcular variaciones en Python (no depender de fórmulas Excel)
        for idx, row in enumerate(filas_com):
            anio_r = str(row[0]).strip()
            mes_r  = str(row[1]).strip() if row[1] else ""
            if "TOTAL" in anio_r.upper():
                hist_data.append({"anio":"TOTAL","mes":"",
                    "trans": row[2] or 0, "com": row[3] or 0,
                    "var_e":0,"var_pct":0,"var_anio":0,"var_anio_pct":0})
                continue
            com_r = float(row[3] or 0)
            # Var mes anterior
            com_ant = float(filas_com[idx-1][3] or 0) if idx > 0 and "TOTAL" not in str(filas_com[idx-1][0]).upper() else 0
            var_e   = com_r - com_ant if com_ant else 0
            var_pct = var_e / com_ant if com_ant else 0
            # Var año anterior (12 filas atrás)
            com_anio = float(filas_com[idx-12][3] or 0) if idx >= 12 else 0
            var_anio     = com_r - com_anio if com_anio else 0
            var_anio_pct = var_anio / com_anio if com_anio else 0
            hist_data.append({
                "anio": anio_r, "mes": mes_r,
                "trans": row[2] or 0, "com": com_r,
                "var_e": var_e, "var_pct": var_pct,
                "var_anio": var_anio, "var_anio_pct": var_anio_pct
            })

        # Leer cedulas
        ws_t = wb_h["TRANSFERENCIAS CORRESPONSAL"]
        for i, row in enumerate(ws_t.iter_rows(min_row=2, values_only=True), start=1):
            if not row[0]: continue
            cedulas_all.append({
                "n": i, "ced": str(row[0]),
                "hist_ant": row[1] or 0,
                "trans_mes": row[2] or 0,
                "hist_act": row[3] or 0,
                "obs": str(row[4]) if row[4] else ""
            })
        wb_h.close()

    # Variación comisión mes actual vs anterior
    var_mes_ant_val = 0
    for d in hist_data:
        if d["anio"] == str(anio) and d["mes"].strip().lower().startswith(mes.lower()[:3]):
            var_mes_ant_val = d["var_e"]
            break

    if var_mes_ant_val > 0:
        var_mes_str = '<p class="kpi-sub kpi-up">&#9650; ${:,.0f} vs {}</p>'.format(abs(var_mes_ant_val), _mes_anterior(mes))
    elif var_mes_ant_val < 0:
        var_mes_str = '<p class="kpi-sub kpi-dn">&#9660; ${:,.0f} vs {}</p>'.format(abs(var_mes_ant_val), _mes_anterior(mes))
    else:
        var_mes_str = ""

    # Datos para graficas
    ml = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    d25t=[0]*12; d26t=["null"]*12; d25c=[0]*12; d26c=[0]*12
    dv6=[0]*12;  dvc=["#1D9E75"]*12

    for d in hist_data:
        if "TOTAL" in str(d["anio"]).upper(): continue
        try:
            a = int(d["anio"])
            m = next((i for i,x in enumerate(ml) if d["mes"].strip().lower().startswith(x.lower())), -1)
            if m < 0: continue
            if a == 2025:
                d25t[m] = d["trans"]; d25c[m] = d["com"]
            elif a == 2026:
                # Solo hasta el mes trabajado
                if m <= mes_idx:
                    d26t[m] = d["trans"]; d26c[m] = d["com"]
                    # Variación: solo meses activos, resto queda null
                    dv6[m]  = d["var_e"]
                    dvc[m]  = "#1D9E75" if d["var_e"] >= 0 else "#D85A30"
        except Exception:
            pass

    # Variación: convertir meses futuros a null
    dv6_js = ["null" if i > mes_idx else str(v) for i, v in enumerate(dv6)]

    # Tabla historica (desc)
    hist_rows = [d for d in hist_data if "TOTAL" not in str(d["anio"]).upper()]
    hist_rows = list(reversed(hist_rows))
    tot_row   = next((d for d in hist_data if "TOTAL" in str(d["anio"]).upper()), None)

    def pill_var(val, is_pct=False):
        try: v = float(val or 0)
        except Exception: return ""
        if abs(v) < 0.0001: return '<span style="color:#aaa">$0</span>' if not is_pct else '<span style="color:#aaa">0.0%</span>'
        c = "kpi-up" if v >= 0 else "kpi-dn"
        s = "&#9650;" if v >= 0 else "&#9660;"
        if is_pct:
            return '<span class="{}">{} {:.1f}%</span>'.format(c, s, abs(v)*100)
        return '<span class="{}">{} ${:,.0f}</span>'.format(c, s, abs(v))

    fh = ""
    for d in hist_rows:
        fh += "<tr><td>{}</td><td>{}</td><td>{:,}</td><td>${:,.0f}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
            d["anio"], d["mes"], int(d["trans"]), d["com"],
            pill_var(d["var_e"]),
            pill_var(d["var_pct"], True),
            pill_var(d["var_anio_pct"], True)
        )
    if tot_row:
        fh += "<tr style='font-weight:600;background:#f5f5f2'><td>TOTAL</td><td></td><td>{:,}</td><td>${:,.0f}</td><td colspan='3' style='color:#888'>{} periodos</td></tr>".format(
            int(tot_row["trans"]), tot_row["com"], len(hist_rows))

    # Tabla cedulas
    fc = ""
    for d in cedulas_all:
        obs = d["obs"]
        if obs == "REICIDENTE":
            pill = "<span class='pill pill-rei'>REICIDENTE</span>"
        elif obs == "NUEVO":
            pill = "<span class='pill pill-nuevo'>NUEVO</span>"
        else:
            pill = "&mdash;"
        fc += "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
            d["n"], d["ced"], d["hist_ant"], d["trans_mes"] or 0, d["hist_act"], pill)

    n_ced = len(cedulas_all)
    jml   = json.dumps(ml)
    jd25t = json.dumps(d25t)
    jd26t = "[{}]".format(",".join(str(x) for x in d26t))
    jd25c = json.dumps(d25c)
    jd26c = json.dumps(d26c)
    # Variación: solo meses hasta mes_idx, el resto null
    dv6_js = ["null" if i > mes_idx else str(int(v)) for i, v in enumerate(dv6)]
    jdv6   = "[{}]".format(",".join(dv6_js))
    jdvc  = json.dumps(dvc)

    css = """*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f2;color:#1a1a1a;padding:24px 16px}
.db-wrap{max-width:960px;margin:0 auto}
.db-header{display:flex;align-items:baseline;gap:12px;margin-bottom:1.5rem;flex-wrap:wrap}
.db-title{font-size:20px;font-weight:500}
.db-badge{font-size:11px;padding:3px 10px;border-radius:20px;background:#eaf3de;color:#3B6D11;font-weight:500}
.db-section{font-size:11px;font-weight:500;color:#888;text-transform:uppercase;letter-spacing:.06em;margin:1.75rem 0 .75rem;border-bottom:0.5px solid rgba(0,0,0,0.1);padding-bottom:6px}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:10px;margin-bottom:1rem}
.kpi{background:#fff;border:0.5px solid rgba(0,0,0,0.1);border-radius:10px;padding:14px 16px}
.kpi-label{font-size:12px;color:#777;margin:0 0 5px}
.kpi-val{font-size:24px;font-weight:500;margin:0;line-height:1.2;color:#1a1a1a}
.kpi-sub{font-size:11px;color:#aaa;margin:4px 0 0}
.kpi-up{color:#1D9E75}.kpi-dn{color:#D85A30}
.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:1rem}
.chart-card{background:#fff;border:0.5px solid rgba(0,0,0,0.1);border-radius:12px;padding:1rem 1.25rem;margin-bottom:1rem}
.chart-title{font-size:14px;font-weight:500;margin:0 0 14px;color:#1a1a1a}
.legend{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:10px;font-size:12px;color:#666}
.leg-dot{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:4px;vertical-align:middle}
.donut-wrap{display:flex;align-items:center;gap:20px;padding:4px 0}
.table-wrap{overflow-x:auto}
table.dt{width:100%;border-collapse:collapse;font-size:13px}
table.dt th{text-align:left;font-weight:500;color:#555;padding:8px 10px;border-bottom:1px solid rgba(0,0,0,0.1);white-space:nowrap;background:#fafaf8;position:sticky;top:0;z-index:1}
table.dt td{padding:6px 10px;border-bottom:0.5px solid rgba(0,0,0,0.06);color:#1a1a1a}
table.dt tr:hover td{background:#fafaf8}
.pill{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500}
.pill-rei{background:#e6f1fb;color:#185FA5}.pill-nuevo{background:#eaf3de;color:#3B6D11}
.controls{display:flex;gap:10px;align-items:center;margin-bottom:12px;flex-wrap:wrap}
.controls input,.controls select{padding:7px 12px;border:0.5px solid rgba(0,0,0,0.2);border-radius:8px;font-size:13px;outline:none;background:#fff}
.controls input{width:220px}
.btn-excel{display:inline-flex;align-items:center;gap:6px;padding:7px 14px;border:0.5px solid rgba(0,0,0,0.2);border-radius:8px;font-size:13px;background:#fff;cursor:pointer;color:#1a1a1a;font-weight:500;margin-left:auto;text-decoration:none}
.btn-excel:hover{background:#f0f0ec}
.scroll{max-height:460px;overflow-y:auto;border-radius:8px;border:0.5px solid rgba(0,0,0,0.1)}
.pag-info{font-size:12px;color:#999;margin-top:8px;text-align:right}
.note-box{background:#e6f1fb;border:0.5px solid #b5d4f4;border-radius:8px;padding:10px 14px;font-size:12px;color:#0C447C;margin-top:1rem}"""

    html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Corresponsal Bancolombia &mdash; {mes} {anio}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>{css}</style>
</head>
<body>
<div class="db-wrap">
  <div class="db-header">
    <p class="db-title">Corresponsal Bancolombia</p>
    <span class="db-badge">{mes} {anio}</span>
  </div>

  <p class="db-section">Transferencias corresponsal &mdash; {mes} {anio}</p>
  <div class="kpi-grid">
    <div class="kpi"><p class="kpi-label">Total transacciones</p><p class="kpi-val">{trans}</p><p class="kpi-sub">mes de {mes} {anio}</p></div>
    <div class="kpi"><p class="kpi-label">Sin identificar</p><p class="kpi-val">{sin_id}</p><p class="kpi-sub">de {trans} transacciones</p></div>
    <div class="kpi"><p class="kpi-label">Identificadas</p><p class="kpi-val">{ident}</p><p class="kpi-sub">{pct_ident}% del total</p></div>
    <div class="kpi"><p class="kpi-label">Reincidentes</p><p class="kpi-val kpi-up">{reinc}</p><p class="kpi-sub kpi-up">{pct_reinc}% de identificados</p></div>
    <div class="kpi"><p class="kpi-label">Nuevos</p><p class="kpi-val">{nuevos}</p><p class="kpi-sub">{pct_nuevo}% de identificados</p></div>
    <div class="kpi"><p class="kpi-label">Comisi&oacute;n {mes}</p><p class="kpi-val">{com_fmt}</p>{var_mes_str}</div>
  </div>

  <div class="chart-row">
    <div class="chart-card">
      <p class="chart-title">Nuevos vs reincidentes ({mes})</p>
      <div class="donut-wrap">
        <div style="position:relative;width:150px;height:150px;flex-shrink:0"><canvas id="c3"></canvas></div>
        <div style="font-size:13px;line-height:2">
          <div><span class="leg-dot" style="background:#185FA5"></span><strong>{pct_reinc}%</strong> Reincidentes &mdash; {reinc}</div>
          <div><span class="leg-dot" style="background:#1D9E75"></span><strong>{pct_nuevo}%</strong> Nuevos &mdash; {nuevos}</div>
          <div style="margin-top:10px;font-size:12px;color:#999">Total identificados: {ident}</div>
        </div>
      </div>
    </div>
    <div class="chart-card">
      <p class="chart-title">Transacciones por mes</p>
      <div class="legend"><span><span class="leg-dot" style="background:#185FA5"></span>2025</span><span><span class="leg-dot" style="background:#1D9E75"></span>2026</span></div>
      <div style="position:relative;width:100%;height:160px"><canvas id="c2"></canvas></div>
    </div>
  </div>

  <p class="db-section">Comisi&oacute;n mensual &mdash; evoluci&oacute;n hist&oacute;rica</p>
  <div class="chart-card">
    <p class="chart-title">Valor comisi&oacute;n por mes (COP)</p>
    <div class="legend"><span><span class="leg-dot" style="background:#185FA5"></span>2025</span><span><span class="leg-dot" style="background:#1D9E75"></span>2026</span></div>
    <div style="position:relative;width:100%;height:260px"><canvas id="c1"></canvas></div>
  </div>
  <div class="chart-card">
    <p class="chart-title">Variaci&oacute;n $ vs mes anterior &mdash; {anio}</p>
    <div style="position:relative;width:100%;height:180px"><canvas id="c4"></canvas></div>
  </div>

  <p class="db-section">Detalle hist&oacute;rico completo</p>
  <div class="chart-card">
    <div class="table-wrap">
      <table class="dt">
        <thead><tr><th>A&ntilde;o</th><th>Mes</th><th>Transacciones</th><th>Comisi&oacute;n (COP)</th><th>Var. $ mes ant.</th><th>Var. % mes ant.</th><th>Var. vs a&ntilde;o ant.</th></tr></thead>
        <tbody>{fh}</tbody>
      </table>
    </div>
  </div>

  <p class="db-section">Base de clientes ({n_ced} registros)</p>
  <div class="chart-card">
    <div class="controls">
      <input type="text" id="s" placeholder="Buscar c&eacute;dula..." oninput="filtrar()">
      <select id="o" onchange="filtrar()">
        <option value="">Todos</option>
        <option value="REICIDENTE">Solo reincidentes</option>
        <option value="NUEVO">Solo nuevos</option>
      </select>
      <a class="btn-excel" onclick="descargarExcel();return false;" href="#">&#128196; Descargar Excel</a>
    </div>
    <div class="scroll">
      <table class="dt" id="tbl">
        <thead><tr><th>#</th><th>C&eacute;dula</th><th>Hist. anterior</th><th>Trans. {mes}</th><th>Hist. actual</th><th>Observaci&oacute;n</th></tr></thead>
        <tbody id="tb">{fc}</tbody>
      </table>
    </div>
    <p class="pag-info" id="pi"></p>
  </div>

  <div class="note-box"><strong>Nota:</strong> Del total de recaudo por corresponsal se est&aacute; exento de las primeras 50 transacciones.</div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
var ML={jml};
var D25T={jd25t}, D26T={jd26t};
var D25C={jd25c}, D26C={jd26c};
var DV6={jdv6},   DVC={jdvc};
var CFG={{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}}}};
var fmtK=function(v){{return v>=1e6?'$'+v/1e6+'M':v>=1e3?'$'+Math.round(v/1e3)+'K':'$'+v;}};

new Chart(document.getElementById('c3'),{{
  type:'doughnut',
  data:{{labels:['Reincidentes','Nuevos'],datasets:[{{data:[{reinc},{nuevos}],backgroundColor:['#185FA5','#1D9E75'],borderWidth:0}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},cutout:'70%'}}
}});

new Chart(document.getElementById('c2'),{{
  type:'line',
  data:{{labels:ML,datasets:[
    {{label:'2025',data:D25T,borderColor:'#185FA5',backgroundColor:'rgba(24,95,165,0.08)',tension:0.4,fill:true}},
    {{label:'2026',data:D26T,borderColor:'#1D9E75',tension:0.4,borderDash:[5,3],fill:false,spanGaps:false}}
  ]}},
  options:{{...CFG,scales:{{y:{{beginAtZero:false}}}}}}
}});

new Chart(document.getElementById('c1'),{{
  type:'bar',
  data:{{labels:ML,datasets:[
    {{label:'2025',data:D25C,backgroundColor:'#185FA5'}},
    {{label:'2026',data:D26C,backgroundColor:'#1D9E75'}}
  ]}},
  options:{{...CFG,scales:{{y:{{ticks:{{callback:fmtK}}}}}}}}
}});

var ML4=ML.slice(0,{mesidx});
var DV6_4=DV6.slice(0,{mesidx});
var DVC_4=DVC.slice(0,{mesidx});
new Chart(document.getElementById('c4'),{{
  type:'bar',
  data:{{labels:ML4,datasets:[{{data:DV6_4,backgroundColor:DVC_4}}]}},
  options:{{...CFG,scales:{{y:{{ticks:{{callback:fmtK}}}}}}}}
}});

var rows=[].slice.call(document.querySelectorAll('#tb tr'));
function filtrar(){{
  var s=document.getElementById('s').value.toLowerCase();
  var o=document.getElementById('o').value;
  var vis=0;
  rows.forEach(function(r){{
    var c=r.cells[1].textContent.toLowerCase();
    var v=r.cells[5].textContent.trim();
    var ok=c.includes(s)&&(!o||v.indexOf(o)>=0);
    r.style.display=ok?'':'none';
    if(ok)vis++;
  }});
  document.getElementById('pi').textContent='Mostrando '+vis+' de {n_ced} registros';
}}
filtrar();

function descargarExcel(){{
  var wb=XLSX.utils.book_new();
  var data=[['#','Cedula','Hist. anterior','Trans. {mes}','Hist. actual','Observacion']];
  rows.forEach(function(r){{
    data.push([r.cells[0].textContent,r.cells[1].textContent,
               r.cells[2].textContent,r.cells[3].textContent,
               r.cells[4].textContent,r.cells[5].textContent.trim()]);
  }});
  var ws=XLSX.utils.aoa_to_sheet(data);
  XLSX.utils.book_append_sheet(wb,ws,'Cedulas');
  XLSX.writeFile(wb,'Cedulas_Corresponsal_{mes}_{anio}.xlsx');
}}
</script>
</body>
</html>""".format(
        mes=mes, anio=anio, css=css, trans=trans, sin_id=sin_id,
        ident=ident, reinc=reinc, nuevos=nuevos,
        pct_ident=pct_ident, pct_reinc=pct_reinc, pct_nuevo=pct_nuevo,
        com_fmt=com_fmt, var_mes_str=var_mes_str,
        fh=fh, fc=fc, n_ced=n_ced,
        jml=jml, jd25t=jd25t, jd26t=jd26t,
        jd25c=jd25c, jd26c=jd26c, jdv6=jdv6, jdvc=jdvc,
        mesidx=mes_idx + 1
    )
    return html.encode("utf-8")



def _mes_anterior(mes):
    idx = MESES.index(mes)
    return MESES[idx-1] if idx > 0 else MESES[11]


    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Corresponsal Bancolombia &mdash; {mes} {anio}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f2;color:#1a1a1a;padding:24px 16px}}
.db-wrap{{max-width:960px;margin:0 auto}}
.db-header{{display:flex;align-items:baseline;gap:12px;margin-bottom:1.5rem;flex-wrap:wrap}}
.db-title{{font-size:20px;font-weight:500}}
.db-badge{{font-size:11px;padding:3px 10px;border-radius:20px;background:#eaf3de;color:#3B6D11;font-weight:500}}
.db-section{{font-size:11px;font-weight:500;color:#888;text-transform:uppercase;letter-spacing:.06em;margin:1.75rem 0 .75rem;border-bottom:0.5px solid rgba(0,0,0,0.1);padding-bottom:6px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:10px;margin-bottom:1rem}}
.kpi{{background:#fff;border:0.5px solid rgba(0,0,0,0.1);border-radius:10px;padding:14px 16px}}
.kpi-label{{font-size:12px;color:#777;margin:0 0 5px}}
.kpi-val{{font-size:24px;font-weight:500;margin:0;line-height:1.2}}
.kpi-sub{{font-size:11px;color:#aaa;margin:4px 0 0}}
.chart-card{{background:#fff;border:0.5px solid rgba(0,0,0,0.1);border-radius:12px;padding:1rem 1.25rem;margin-bottom:1rem}}
.chart-title{{font-size:14px;font-weight:500;margin:0 0 14px}}
table.dt{{width:100%;border-collapse:collapse;font-size:13px}}
table.dt th{{text-align:left;font-weight:500;color:#555;padding:8px 10px;border-bottom:1px solid rgba(0,0,0,0.1);white-space:nowrap;background:#fafaf8;position:sticky;top:0}}
table.dt td{{padding:6px 10px;border-bottom:0.5px solid rgba(0,0,0,0.06)}}
table.dt tr:hover td{{background:#fafaf8}}
.pill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500}}
.pill-rei{{background:#e6f1fb;color:#185FA5}}
.pill-nuevo{{background:#eaf3de;color:#3B6D11}}
.controls{{display:flex;gap:10px;align-items:center;margin-bottom:12px;flex-wrap:wrap}}
.controls input,.controls select{{padding:7px 12px;border:0.5px solid rgba(0,0,0,0.2);border-radius:8px;font-size:13px;outline:none;background:#fff}}
.controls input{{width:220px}}
.scroll{{max-height:460px;overflow-y:auto;border-radius:8px;border:0.5px solid rgba(0,0,0,0.1)}}
.note-box{{background:#e6f1fb;border:0.5px solid #b5d4f4;border-radius:8px;padding:10px 14px;font-size:12px;color:#0C447C;margin-top:1rem}}
</style>
</head>
<body>
<div class="db-wrap">
  <div class="db-header">
    <p class="db-title">Corresponsal Bancolombia</p>
    <span class="db-badge">{mes} {anio}</span>
  </div>
  <p class="db-section">Transferencias corresponsal &mdash; {mes} {anio}</p>
  <div class="kpi-grid">
    <div class="kpi"><p class="kpi-label">Total transacciones</p><p class="kpi-val">{trans}</p><p class="kpi-sub">mes de {mes} {anio}</p></div>
    <div class="kpi"><p class="kpi-label">Sin identificar</p><p class="kpi-val">{sin_id}</p><p class="kpi-sub">de {trans} transacciones</p></div>
    <div class="kpi"><p class="kpi-label">Identificadas</p><p class="kpi-val">{ident}</p><p class="kpi-sub">{pct_ident}% del total</p></div>
    <div class="kpi"><p class="kpi-label">Reincidentes</p><p class="kpi-val">{reinc}</p><p class="kpi-sub">{pct_reinc}% de identificados</p></div>
    <div class="kpi"><p class="kpi-label">Nuevos</p><p class="kpi-val">{nuevos}</p><p class="kpi-sub">{pct_nuevo}% de identificados</p></div>
    <div class="kpi"><p class="kpi-label">Comisi&oacute;n {mes}</p><p class="kpi-val">{com_fmt}</p></div>
  </div>
  <p class="db-section">Base de clientes ({len(stats['actualizaciones'])} registros)</p>
  <div class="chart-card">
    <div class="controls">
      <input type="text" id="s" placeholder="Buscar c&eacute;dula..." oninput="f()">
      <select id="o" onchange="f()">
        <option value="">Todos</option>
        <option value="REICIDENTE">Solo reincidentes</option>
        <option value="NUEVO">Solo nuevos</option>
      </select>
    </div>
    <div class="scroll">
      <table class="dt">
        <thead><tr><th>#</th><th>C&eacute;dula</th><th>Hist. anterior</th><th>Trans. {mes}</th><th>Hist. actual</th><th>Observaci&oacute;n</th></tr></thead>
        <tbody id="tb">{filas_ced}</tbody>
      </table>
    </div>
  </div>
  <div class="note-box"><strong>Nota:</strong> Del total de recaudo por corresponsal se est&aacute; exento de las primeras 50 transacciones.</div>
</div>
<script>
function f(){{
  var s=document.getElementById('s').value.toLowerCase();
  var o=document.getElementById('o').value;
  document.querySelectorAll('#tb tr').forEach(function(r){{
    var c=r.cells[1].textContent.toLowerCase();
    var v=r.cells[5].textContent.trim();
    r.style.display=(c.includes(s)&&(!o||v===o))?'':'none';
  }});
}}
</script>
</body>
</html>"""
    return html.encode("utf-8")
